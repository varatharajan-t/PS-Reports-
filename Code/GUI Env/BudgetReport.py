"""
************************************************************************************************************
SAP PROJECT SYSTEM - BUDGET REPORT PROCESSOR
************************************************************************************************************

SYSTEM REQUIREMENTS:
- Python 3.7+
- pandas >= 1.3.0
- openpyxl >= 3.0.0
- PySide6 >= 6.0.0
- re (built-in)
- os (built-in)
- sys (built-in)
- pathlib (built-in)

PURPOSE:
This program processes SAP DAT files containing budget data and converts them to formatted Excel files.
It extracts Work Breakdown Structure (WBS) elements, applies data transformations, and creates
professionally formatted Excel reports with currency formatting, styling, and conditional highlighting.

INPUT REQUIREMENTS:
- DAT file: Tab-delimited file with multi-level headers from SAP
- WBS_NAMES.XLSX: Master file containing WBS element descriptions
- Files should be in the same directory as the selected DAT file

OUTPUT:
- Excel file with same name as input DAT file but with .XLSX extension
- Formatted with currency columns, color-coded headers, and summary WBS highlighting
- Contains processed WBS hierarchy with Level, Description, and ID columns

MAIN WORKFLOW:
1. File Selection: User selects DAT file via dialog
2. Data Cleaning: Removes unnecessary header/footer lines
3. Data Transformation: Splits WBS elements and maps descriptions
4. WBS Classification: Separates summary and transaction WBS elements
5. Excel Formatting: Applies professional styling and conditional formatting

AUTHOR: [Varatha Rajan T]
DATE: [12/03/2024]
LAST MODIFIED: [24/09/2025]
************************************************************************************************************
"""

import os, sys
import pandas as pd
import re
from pathlib import Path
from PySide6.QtWidgets import QFileDialog, QApplication

import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from config import Config
from error_handler import SAPReportLogger


class FileHandler:
    """
    Handles file selection operations using PySide6 QFileDialog.

    This class provides static methods for file operations including:
    - Opening file selection dialogs
    - Path and filename extraction
    - Directory management
    """

    @staticmethod
    def get_file_path_and_name():
        """
        Opens a file dialog to select a DAT file and extracts path components.

        Uses PySide6 QFileDialog to present a native file selection interface
        filtered specifically for DAT files (*.DAT). The dialog allows users
        to browse and select the input file for processing.

        Returns:
            tuple: (file_path, file_name) if file selected, (None, None) if cancelled
                - file_path (str): Directory path containing the selected file
                - file_name (str): Name of the selected file with extension

        Example:
            path, name = FileHandler.get_file_path_and_name()
            # Returns: ("/home/user/data", "budget_report.DAT")
        """
        logger = SAPReportLogger("FileHandler")
        filepath = QFileDialog.getOpenFileName(
            None, "Select a DAT file", "", "DAT Files(*.DAT)"
        )
        logger.log_info(f"File selected: {filepath[0]}")

        if filepath:
            file_path, file_name = os.path.split(filepath[0])
            return file_path, file_name
        else:
            return None, None


class DataProcessor:
    """
    Handles all data processing operations for SAP DAT files.

    This class manages the complete data transformation pipeline including:
    - Reading tab-delimited DAT files with multi-level headers
    - Data cleaning and preprocessing
    - WBS element extraction and classification
    - Data structure reorganization
    - Integration with master WBS description files

    Attributes:
        file_path (str): Directory path containing the input file
        file_name (str): Name of the input DAT file
        df (pd.DataFrame): Main dataframe holding processed data

    Processing Steps:
        1. Clean raw DAT file by removing header/footer lines
        2. Read cleaned file into pandas DataFrame
        3. Transform WBS element structure
        4. Map WBS descriptions from master file
        5. Classify WBS elements into summary/transaction categories
    """

    def __init__(self, file_full_path):
        """
        Initialize DataProcessor with file information.

        Args:
            file_full_path (str): The full path to the input DAT file to process.
        """
        self.file_full_path = file_full_path
        self.df = None
        self.summary = False
        self.summary_wbs_list = []
        self.transaction_wbs_list = []
        self.logger = SAPReportLogger("DataProcessor")

    def read_data(self, cleaned_DAT, delimiter="\t"):
        """
        Reads the cleaned DAT file into a pandas DataFrame with multi-level headers.

        The method handles SAP's multi-level header structure by reading two header rows
        and converting column names to tuple format for proper handling.

        Args:
            cleaned_DAT (str): Path to the cleaned DAT file
            delimiter (str): Field separator (default: tab "\t")

        Returns:
            pd.DataFrame: DataFrame with multi-level column headers

        Note:
            Uses ISO-8859-1 encoding to handle special characters in SAP exports
        """
        self.df = pd.read_csv(
            cleaned_DAT, sep=delimiter, header=[0, 1], encoding="iso-8859-1"
        )
        self.df.columns = self.df.columns.map(lambda x: tuple(map(str, x)))
        return self.df

    def clean_data(self, cleaned_DAT):
        """
        Cleans the raw DAT file by removing unnecessary header and footer lines.

        SAP DAT exports typically contain metadata lines that interfere with data processing.
        This method removes specific lines that don't contain useful data:
        - Lines 1, 2: Header metadata
        - Line 5: Additional header information
        - Last line: Footer/summary information

        Args:
            cleaned_DAT (str): Output filename for cleaned data

        Side Effects:
            Creates a new cleaned DAT file with unnecessary lines removed
        """
        # Read the file
        with open(self.file_full_path, "r") as file:
            lines = file.readlines()

        # Calculate total number of lines
        total_lines = len(lines)

        # The position of object varies for all the projects and individual project
        ctr = 0
        for line in lines:
            firstWord = line.split()[0]
            if firstWord == "Object":
                break
            else:
                # self.logger.log_info("Counter: ",ctr,line)
                ctr = ctr + 1

        self.logger.log_info(f"Object Found in line {ctr}")
        if (
            ctr == 2
        ):  # if object in line 2 this is individual project else this is a summary one
            # Remove the first, fourth, fifth, and last line
            lines_to_delete = [0, 3, 4, total_lines - 1]
            self.summary = False
        else:
            # Remove the first, second, fifth, and last line
            lines_to_delete = [0, 1, 4, total_lines - 1]
            self.summary = True

        lines = [line for i, line in enumerate(lines) if i not in lines_to_delete]

        # Write the modified content back to the file
        with open(cleaned_DAT, "w") as file:
            file.writelines(lines)

    @staticmethod
    def extract_parts(s):
        level = re.search(r"^(\*+|\d+\*)", s.strip())
        proj_id = re.search(r"[A-Z]{2}-[A-Z]-[A-Z]{3}-\d{3}", s)
        desc = re.sub(
            r"(^(\*+|\d+\*)|\bPRJ\b|[A-Z]{2}-[A-Z]-[A-Z]{3}-\d{3})", "", s
        ).strip()
        desc = re.sub(r"\s+", " ", desc)  # normalize spaces
        return [
            level.group(0) if level else None,
            desc,
            proj_id.group(0) if proj_id else None,
        ]

    def transform_data(self, master_file, output_file, cleaned_DAT):
        master_file = master_file
        output_file = output_file

        df_cleaned = self.read_data(cleaned_DAT)
        df_columns = list(df_cleaned.columns[1:])  ### New Line of Code

        # Rename the first column header
        df_cleaned = df_cleaned.rename(
            columns={"Unnamed: 0_level_0": "WBS_Elements_Info."}
        )

        # Split "WBS Element Details" column into separate columns
        split_details = df_cleaned[("WBS_Elements_Info.", "Object")].str.split()

        """Transforms data by splitting the WBS Element Details and renaming columns to a multi-index format."""
        level, description, id_no = [], [], []

        for details in split_details:
            newString = DataProcessor.extract_parts(" ".join(details))
            Level = newString[0]
            Description = newString[1]
            ID = newString[2]
            # self.logger.log_info('Level = ',Level,'\n Description =',Description,'\n ID =',ID)

            if Level not in ["*", "**", "***", "4*", "5*", "6*"]:
                if Level == None or Level == " ":
                    Description = Description
                else:
                    Description = Level + Description
                Level = " "

            level.append(Level)
            description.append(Description)
            id_no.append(ID)

        # Add the new columns to the DataFrame
        df_cleaned[("WBS_Elements_Info.", "Level")] = level
        df_cleaned[("WBS_Elements_Info.", "Description")] = description
        df_cleaned[("WBS_Elements_Info.", "ID_No")] = id_no

        # Re-arrange columns to place new columns at the beginning
        df_cleaned = df_cleaned[
            [
                ("WBS_Elements_Info.", "Level"),
                ("WBS_Elements_Info.", "Description"),
                ("WBS_Elements_Info.", "ID_No"),
            ]
            + df_columns
        ]  # Modified line of code

        # Updating the Description of WBS elements to full description
        master_df = pd.read_excel(master_file)
        transaction_df = df_cleaned

        master_df["WBS_element"] = master_df["WBS_element"].str.strip()
        transaction_df[("WBS_Elements_Info.", "ID_No")] = transaction_df[
            ("WBS_Elements_Info.", "ID_No")
        ].str.strip()

        transaction_df[("WBS_Elements_Info.", "Description")] = transaction_df[
            ("WBS_Elements_Info.", "ID_No")
        ].map(master_df.set_index("WBS_element")["Name"])

        transaction_df = transaction_df.dropna(how="all")

        # Process WBS classification before flattening columns
        self.summary_wbs_list, self.transaction_wbs_list = self.process_wbs(
            transaction_df[("WBS_Elements_Info.", "ID_No")].to_list()
        )

        # Flatten MultiIndex columns to strings for Excel compatibility
        transaction_df.columns = [' - '.join(col).strip() for col in transaction_df.columns]

        # Add serial number column at the beginning
        transaction_df.insert(0, 'Sl No.', range(1, len(transaction_df) + 1))

        # self.logger.log_info("Summary WBS Elements: ",self.summary_wbs_list,"\n\n Transaction WBS Elements : ",self.transaction_wbs_list)
        transaction_df.to_excel(output_file, index=False)
        self.logger.log_info(f"Updated {output_file} saved successfully!")

    def process_wbs(self, wbs_list):
        """Classifies WBS elements into summary and transaction WBS dynamically."""
        # Summary projects have no child wbs
        if self.summary:
            return [], wbs_list

        summary_wbs = []
        transaction_wbs = []

        for wbs in wbs_list:
            # Define a regular expression pattern to match a WBS element with child elements
            pattern = re.compile(re.escape(wbs) + r"-\d{2}")

            # Check if any WBS starts with the parent WBS followed by a hyphen and two digits
            has_child = any(pattern.fullmatch(item) for item in wbs_list if item != wbs)

            if has_child:
                summary_wbs.append(wbs)
            else:
                transaction_wbs.append(wbs)

        return summary_wbs, transaction_wbs

    def add_heading(self, output_file):
        """
        Placeholder method - serial number column is now added during data transformation.
        Kept for backward compatibility.
        """
        return output_file


class ExcelFormatter:
    """
    Handles comprehensive Excel formatting and styling operations.

    This class provides professional formatting capabilities including:
    - Font styling and currency formatting
    - Column width auto-adjustment
    - Freeze panes for navigation
    - Color-coded headers and alternating row colors
    - Conditional row highlighting based on data values
    - Table creation with predefined styles

    Attributes:
        output_file (str): Path to the Excel file being formatted
        workbook (openpyxl.Workbook): Excel workbook object
        worksheet (openpyxl.Worksheet): Active worksheet being formatted
        last_col (int): Last column index with data
        last_row (int): Last row index with data

    Formatting Features:
        - Headers: Yellow background with bold text
        - Data rows: Alternating sky blue and white backgrounds
        - Currency columns: Indian Rupee formatting with red negatives
        - Summary WBS: Light green highlighting for identification
    """

    def __init__(self, output_file):
        """
        Initialize ExcelFormatter with workbook loading and dimension calculation.

        Args:
            output_file (str): Path to Excel file to be formatted

        Side Effects:
            Loads existing Excel file and caches worksheet dimensions
        """
        self.output_file = output_file
        self.workbook = openpyxl.load_workbook(output_file)
        self.worksheet = self.workbook.active
        self.last_col = self.worksheet.max_column
        self.last_row = self.worksheet.max_row

    def apply_freeze_panes(self):
        """Freeze panes for the Excel sheet."""
        self.worksheet.freeze_panes = "E3"

    def apply_font_style(self):
        """Apply a font style to all cells in the worksheet."""
        font = Font(name="Bookman Old Style", size=12)
        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.font = font

    def apply_currency_formatting(self):
        """Apply currency formatting to specific columns."""
        currency_style = NamedStyle(
            name="currency",
            font=Font(name="Bookman Old Style", size=12),
            number_format=f"₹ #,##0.00;[Red]₹ -#,##0.00",
        )

        for col in range(4, self.last_col + 1):
            col_letter = get_column_letter(col)
            column = self.worksheet[col_letter]
            for cell in column:
                cell.style = currency_style

    def adjust_column_width(self):
        for col in range(1, self.last_col + 1):
            col_letter = get_column_letter(col)
            max_length = max(
                (
                    len(str(cell.value))
                    for cell in self.worksheet[col_letter]
                    if cell.value
                ),
                default=10,
            )
            self.worksheet.column_dimensions[col_letter].width = max_length + 10

    def create_table(self):
        """Create a table in the worksheet from data range."""
        data_range = f"A1:{self.worksheet.cell(row=self.last_row, column=self.last_col).coordinate}"
        # data_range = "A1:AM690"

        table = Table(displayName="Table1", ref=data_range)
        self.worksheet.add_table(table)

        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

    def apply_header_format(self):
        """Apply specific formatting to the header cells."""
        last_col_letter = get_column_letter(self.last_col)
        # lastPos= last_col_letter+str(self.last_row)

        # Define Black Border
        black_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        cell_range_str = f"A1:{self.worksheet.cell(row=self.last_row, column=self.last_col).coordinate}"
        cell_range = self.worksheet[cell_range_str]

        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        skyBlue_fill = PatternFill(
            start_color="87CEEB", end_color="87CEEB", fill_type="solid"
        )

        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        rowCtr = 0
        # White text
        bold_font = Font(name="Bookman Old Style", size=12, color="000000", bold=True)

        for row in cell_range:

            for cell in row:
                if rowCtr in [0, 1]:
                    cell.fill = yellow_fill
                    cell.border = black_border
                    cell.font = bold_font
                else:
                    if rowCtr % 2 == 0:
                        cell.fill = skyBlue_fill
                        cell.border = black_border
                    else:
                        cell.fill = white_fill
                        cell.border = black_border
                        # cell.font = white_font
            rowCtr += 1

    def highlight_rows_by_value(self, values_list):
        """
        For each cell in column 'D', if the cell's value is in values_list,
        highlight the entire row with a yellow fill.

        Parameters:
        self.worksheet          : openpyxl worksheet object
        values_list : list of values to check against
        """
        light_green_fill = PatternFill(
            start_color="90EE90", end_color="90EE90", fill_type="solid"
        )

        # self.worksheet['D'] returns all cells in column D
        for cell in self.worksheet["D"]:
            if cell.value in values_list:
                # Iterate through all cells in the current row and apply the yellow fill
                for row_cell in self.worksheet[cell.row]:
                    row_cell.fill = light_green_fill

    def save(self):
        """Save the final formatted workbook."""
        self.workbook.save(self.output_file)


def main():
    """
    Main execution function for the Budget Report Processor.
    """
    config = Config()
    logger = SAPReportLogger("BudgetReport")
    
    # Step 1: File selection
    file_path, file_name = FileHandler.get_file_path_and_name()
    if not file_path:
        return

    input_file = os.path.join(file_path, file_name)
    output_file = os.path.join(file_path, Path(file_name).with_suffix(".XLSX"))
    cleaned_DAT = os.path.join(file_path, config.TEMP_DAT_FILE)
    master_file = config.MASTER_WBS_FILE

    # Step 2: Data processing
    data_processor = DataProcessor(input_file)
    data_processor.clean_data(cleaned_DAT)
    data_processor.transform_data(master_file, output_file, cleaned_DAT)
    data_processor.add_heading(output_file)

    # Step 3: Excel formatting
    excel_formatter = ExcelFormatter(output_file)
    excel_formatter.apply_freeze_panes()
    excel_formatter.apply_font_style()
    excel_formatter.apply_currency_formatting()
    excel_formatter.adjust_column_width()
    # excel_formatter.create_table()
    excel_formatter.apply_header_format()

    excel_formatter.highlight_rows_by_value(data_processor.summary_wbs_list)
    excel_formatter.save()

    logger.log_info(f"Excel file '{output_file}' has been created with the specified formatting.")



if __name__ == "__main__":
    main()
