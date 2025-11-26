"""
************************************************************************************************************
SAP PROJECT SYSTEM - BUDGET UPDATES PROCESSOR
************************************************************************************************************

SYSTEM REQUIREMENTS:
- Python 3.7+
- pandas >= 1.3.0
- openpyxl >= 3.0.0
- PySide6 >= 6.0.0
- pathlib (built-in)
- os (built-in)
- sys (built-in)
- re (built-in)

PURPOSE:
Processes SAP DAT files containing budget update data and converts them to professionally
formatted Excel files. Specializes in handling budget modification reports with enhanced
WBS element processing and conditional formatting for summary vs transaction elements.

KEY DIFFERENCES FROM BudgetReport.py:
- Modified data cleaning removes lines 1, 4, and last line (different pattern)
- Enhanced WBS processing with regex-based child element detection
- Improved string manipulation for WBS element extraction
- Streamlined master file integration workflow

INPUT REQUIREMENTS:
- DAT file: Tab-delimited SAP export with budget update data
- WBS_NAMES.XLSX: Master reference file for WBS descriptions
- Files must be in the same working directory

OUTPUT SPECIFICATIONS:
- Excel file (.xlsx) with same base name as input DAT file
- Professional formatting with currency columns
- Color-coded rows and conditional highlighting
- Frozen panes for better navigation

MAIN WORKFLOW:
1. File Selection: Interactive DAT file selection via dialog
2. Data Cleaning: Intelligent removal of SAP metadata lines
3. WBS Processing: Advanced parsing and classification of WBS elements
4. Data Transformation: Structure optimization for Excel presentation
5. Master Data Integration: WBS description mapping and validation
6. Excel Generation: Comprehensive formatting and styling application

COMPLEX ALGORITHMS:
- WBS Classification: Uses regex patterns to identify parent-child relationships
- Dynamic Column Reorganization: Maintains data integrity during structure changes
- Conditional Highlighting: Summary WBS elements highlighted for quick identification

AUTHOR: [Original Author]
CREATION DATE: [Creation Date]
LAST MODIFIED: [Modification Date]
************************************************************************************************************
"""

import os, sys, re
import pandas as pd
from pathlib import Path
from PySide6.QtWidgets import QFileDialog, QApplication

import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# from config import Config
# from error_handler import SAPReportLogger


class FileHandler:
    """Class to handle file selection and directory management."""

    @staticmethod
    def get_file_path_and_name():
        """Opens a file dialog and returns the selected file path and name."""
        logger = SAPReportLogger("FileHandler")
        filepath = QFileDialog.getOpenFileName(
            None, "Select a DAT file", "", "DAT Files(*.DAT)"
        )
        logger.log_info(f"File selected: {filepath[0]}")

        if filepath:
            file_path, file_name = os.path.split(filepath[0])
            return file_path, file_name
        else:
            return None, None


class DataProcessor:
    """Class to handle reading, cleaning, and transforming the data."""

    def __init__(self, file_full_path):
        self.file_full_path = file_full_path
        self.df = None
        self.summary_wbs_list = []
        self.transaction_wbs_list = []
        self.logger = SAPReportLogger("DataProcessor")

    def read_data(self, cleaned_DAT, delimiter="\t"):
        """Reads the DAT file into a DataFrame."""
        self.df = pd.read_csv(
            cleaned_DAT, sep=delimiter, header=[0, 1], encoding="iso-8859-1"
        )
        self.df.columns = self.df.columns.map(lambda x: tuple(map(str, x)))
        return self.df

    def clean_data(self, cleaned_DAT):
        # Read the file
        with open(self.file_full_path, "r") as file:
            lines = file.readlines()

        # Calculate total number of lines
        total_lines = len(lines)

        # Remove the first, second, fifth, and last line
        lines_to_delete = [0, 3, total_lines - 1]
        lines = [line for i, line in enumerate(lines) if i not in lines_to_delete]

        lines = [line.replace("*", " ") for line in lines]

        # Write the modified content back to the file
        with open(cleaned_DAT, "w") as file:
            file.writelines(lines)

    def transform_data(self, master_file, output_file, cleaned_DAT):
        master_file = master_file
        output_file = output_file

        df_cleaned = self.read_data(cleaned_DAT)
        df_columns = list(df_cleaned.columns[1:])  # New Line of Code

        # Rename the first column header
        df_cleaned = df_cleaned.rename(
            columns={"Unnamed: 0_level_0": "WBS_Elements_Info."}
        )

        # Split "WBS Element Details" column into separate columns
        split_details = df_cleaned[("WBS_Elements_Info.", "Object")].str.split()

        """Transforms data by splitting the WBS Element Details and renaming columns to a multi-index format."""
        level, description, id_no = [], [], []

        for details in split_details:
            Level = details[0]
            Description = "".join(details[2:])
            ID = details[1]

            if Level not in ["*", "**", "***", "4*", "5*"]:
                Description = Level.strip() + Description.strip()
                Level = " "

            level.append(Level)
            description.append(Description)
            id_no.append(ID)

        # Add the new columns to the DataFrame
        df_cleaned[("WBS_Elements_Info.", "Level")] = level
        df_cleaned[("WBS_Elements_Info.", "Description")] = description
        df_cleaned[("WBS_Elements_Info.", "ID_No")] = id_no

        # Re-arrange columns to place new columns at the beginning
        df_cleaned = df_cleaned[
            [
                ("WBS_Elements_Info.", "Level"),
                ("WBS_Elements_Info.", "Description"),
                ("WBS_Elements_Info.", "ID_No"),
            ]
            + df_columns
        ]

        # Updating the Description of WBS elements to full description
        master_df = pd.read_excel(master_file)
        transaction_df = df_cleaned

        master_df['WBS_element'] = master_df['WBS_element'].str.strip()
        transaction_df[('WBS_Elements_Info.', 'ID_No')] = transaction_df[(
            'WBS_Elements_Info.', 'ID_No')].str.strip()

        transaction_df[('WBS_Elements_Info.', 'Description')] = transaction_df[('WBS_Elements_Info.', 'ID_No')].map(
            master_df.set_index('WBS_element')['Name'])

        transaction_df = transaction_df.dropna(how="all")

        # Process WBS classification before flattening columns
        self.summary_wbs_list, self.transaction_wbs_list = self.process_wbs(
            transaction_df[("WBS_Elements_Info.", "ID_No")].to_list()
        )

        # Flatten MultiIndex columns to strings for Excel compatibility
        transaction_df.columns = [' - '.join(col).strip() if isinstance(col, tuple) else str(col) for col in transaction_df.columns]

        # Add serial number column at the beginning
        transaction_df.insert(0, 'Sl No.', range(1, len(transaction_df) + 1))

        # self.logger.log_info("Summary WBS Elements: ",self.summary_wbs_list,"\n\n Transaction WBS Elements : ",self.transaction_wbs_list)
        transaction_df.to_excel(output_file, index=False)
        self.logger.log_info(f"Updated {output_file} saved successfully!")

    def process_wbs(self, wbs_list):
        summary_wbs = []
        transaction_wbs = []
        # Iterate over the WBS IDs
        for wbs in wbs_list:
            # Define a regular expression pattern to match a WBS element with child elements
            pattern = re.compile(re.escape(wbs) + r"-\d{2}")

            # Check if any WBS starts with the parent WBS followed by a hyphen and two digits
            has_child = any(pattern.fullmatch(item) for item in wbs_list if item != wbs)

            if has_child:
                summary_wbs.append(wbs)  # Add child WBS to summary WBS
            else:
                transaction_wbs.append(wbs)  # Add WBS to transaction WBS

        # print("Summary WBS Elements: ",summary_wbs,"\n\n Transaction WBS Elements : ",transaction_wbs)
        return summary_wbs, transaction_wbs

    def add_heading(self, output_file):
        """
        Placeholder method - serial number column is now added during data transformation.
        Kept for backward compatibility.
        """
        return output_file


class ExcelFormatter:
    """Class to handle Excel file formatting."""

    def __init__(self, output_file):
        self.output_file = output_file
        self.workbook = openpyxl.load_workbook(output_file)
        self.worksheet = self.workbook.active
        self.last_col = self.worksheet.max_column
        self.last_row = self.worksheet.max_row

    def apply_freeze_panes(self):
        """Freeze panes for the Excel sheet."""
        self.worksheet.freeze_panes = "E3"

    def apply_font_style(self):
        """Apply a font style to all cells in the worksheet."""
        font = Font(name="Bookman Old Style", size=12)
        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.font = font

    def apply_currency_formatting(self):
        """Apply currency formatting to specific columns."""
        currency_style = NamedStyle(
            name="currency",
            font=Font(name="Bookman Old Style", size=12),
            number_format=f"₹ #,##0.00;[Red]₹ -#,##0.00",
        )

        for col in range(4, self.last_col + 1):
            col_letter = get_column_letter(col)
            column = self.worksheet[col_letter]
            for cell in column:
                cell.style = currency_style

    def adjust_column_width(self):
        for col in range(1, self.last_col + 1):
            col_letter = get_column_letter(col)
            max_length = max(
                (
                    len(str(cell.value))
                    for cell in self.worksheet[col_letter]
                    if cell.value
                ),
                default=10,
            )
            self.worksheet.column_dimensions[col_letter].width = max_length + 10

    def create_table(self):
        """Create a table in the worksheet from data range."""
        data_range = f"A1:{self.worksheet.cell(row=self.last_row, column=self.last_col).coordinate}"
        # data_range = "A1:AM690"

        table = Table(displayName="Table1", ref=data_range)
        self.worksheet.add_table(table)

        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

    def apply_header_format(self):
        """Apply specific formatting to the header cells."""
        last_col_letter = get_column_letter(self.last_col)
        # lastPos= last_col_letter+str(self.last_row)

        # Define Black Border
        black_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        cell_range_str = f"A1:{self.worksheet.cell(row=self.last_row, column=self.last_col).coordinate}"
        cell_range = self.worksheet[cell_range_str]

        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        skyBlue_fill = PatternFill(
            start_color="87CEEB", end_color="87CEEB", fill_type="solid"
        )

        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        rowCtr = 0
        # White text
        bold_font = Font(name="Bookman Old Style", size=12, color="000000", bold=True)

        for row in cell_range:

            for cell in row:
                if rowCtr in [0, 1]:
                    cell.fill = yellow_fill
                    cell.border = black_border
                    cell.font = bold_font
                else:
                    if rowCtr % 2 == 0:
                        cell.fill = skyBlue_fill
                        cell.border = black_border
                    else:
                        cell.fill = white_fill
                        cell.border = black_border
                        # cell.font = white_font
            rowCtr += 1

    def highlight_rows_by_value(self, values_list):
        """
        For each cell in column 'D', if the cell's value is in values_list,
        highlight the entire row with a yellow fill.

        Parameters:
        self.worksheet          : openpyxl worksheet object
        values_list : list of values to check against
        """
        light_green_fill = PatternFill(
            start_color="90EE90", end_color="90EE90", fill_type="solid"
        )

        # self.worksheet['D'] returns all cells in column D
        for cell in self.worksheet["D"]:
            if cell.value in values_list:
                # Iterate through all cells in the current row and apply the yellow fill
                for row_cell in self.worksheet[cell.row]:
                    row_cell.fill = light_green_fill

    def save(self):
        """Save the final formatted workbook."""
        self.workbook.save(self.output_file)


# Main program execution
def main():
    # config = Config()
    # logger = SAPReportLogger("BudgetUpdates")
    # app = QApplication(sys.argv)
    
    # Step 1: File selection
    file_path, file_name = FileHandler.get_file_path_and_name()
    if not file_path:
        return

    input_file = os.path.join(file_path, file_name)
    output_file = os.path.join(file_path, Path(file_name).with_suffix(".XLSX"))
    cleaned_DAT = os.path.join(file_path, config.TEMP_DAT_FILE)
    master_file = config.MASTER_WBS_FILE # This is now the full path from the project root

    # Step 2: Data processing
    data_processor = DataProcessor(input_file)
    data_processor.clean_data(cleaned_DAT)
    data_processor.transform_data(master_file, output_file, cleaned_DAT)
    data_processor.add_heading(output_file)

    # Step 3: Excel formatting
    excel_formatter = ExcelFormatter(output_file)
    excel_formatter.apply_freeze_panes()
    excel_formatter.apply_font_style()
    excel_formatter.apply_currency_formatting()
    excel_formatter.adjust_column_width()
    # excel_formatter.create_table()
    excel_formatter.apply_header_format()

    # logger.log_info(data_processor.summary_wbs_list)
    excel_formatter.highlight_rows_by_value(data_processor.summary_wbs_list)
    excel_formatter.save()

    logger.log_info(f"Excel file '{output_file}' has been created with the specified formatting.")


if __name__ == "__main__":
    main()
# ****************************************************************
