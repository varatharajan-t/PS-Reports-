# ************************************************************************************************************
# This program takes input HTML file and outputs Excel file. Select the HTML file of large size to proces
# ***************************************************************************************************************

import os
import sys
import pandas as pd
from pathlib import Path
from PySide6.QtWidgets import QFileDialog, QApplication
import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from config import Config
from error_handler import SAPReportLogger


class FileHandler:
    """Handles file selection for processing."""

    @staticmethod
    def get_file_path_and_name():
        """Opens a file dialog and returns the selected file path and name."""
        filepath, _ = QFileDialog.getOpenFileName(
            None, "Select an HTML file", "", "HTML Files (*.html)"
        )
        return os.path.split(filepath) if filepath else (None, None)


class DataProcessor:
    """Handles data reading, cleaning, and transformation."""

    def __init__(self, file_full_path):
        self.file_full_path = file_full_path
        self.df = None
        self.logger = SAPReportLogger("DataProcessor")

    def clean_data(self, cleaned_html):
        """Removes the first two lines from the HTML file and saves it as a cleaned version."""
        with open(self.file_full_path, "r", encoding="utf-8") as file:
            lines = file.readlines()

        if len(lines) > 2:
            with open(cleaned_html, "w", encoding="utf-8") as file:
                file.writelines(lines[2:])  # Remove first two lines
            self.logger.log_info(f"Cleaned HTML saved as {cleaned_html}")
        else:
            self.logger.log_warning("File doesn't have enough lines to clean.")

    def read_data(self, cleaned_html):
        """Reads the cleaned HTML into a Pandas DataFrame."""
        self.df = pd.read_html(cleaned_html, header=[0, 1], encoding="utf-8")[0]
        self.df.columns = self.df.columns.map(lambda x: " ".join(map(str, x)).strip())

        # Drop the last row of the DataFrame
        self.df = self.df.drop(self.df.index[-1])

        return self.df

    def transform_data(self, cleaned_html, master_file, output_file):
        """Processes WBS elements and updates descriptions."""
        df_cleaned = self.read_data(cleaned_html)
        column_mapping = {
            df_cleaned.columns[0]: "WBS_element",
            df_cleaned.columns[1]: "Description",
        }
        df_cleaned = df_cleaned.rename(columns=column_mapping)
        df_cleaned["WBS_element"] = (
            df_cleaned["WBS_element"].astype(str).str.strip().fillna("")
        )

        master_df = pd.read_excel(master_file)
        master_df.rename(
            columns={"Actual Column Name": "Name", "Another Column": "Name"},
            inplace=False,
        )
        df_cleaned["Description"] = (
            df_cleaned["WBS_element"]
            .map(master_df.set_index("WBS_element")["Name"])
            .fillna(df_cleaned["Description"])
        )
        df_cleaned = df_cleaned.dropna(how="all")

        # Add the "SI_NO" column for serial numbers starting at 1
        df_cleaned.insert(0, "SI_NO", range(1, len(df_cleaned) + 1))

        summary_wbs_list, _ = self.process_wbs(df_cleaned["WBS_element"].tolist())
        df_cleaned.to_excel(output_file, index=False)
        self.logger.log_info(f"Updated {output_file} saved successfully!")
        return summary_wbs_list

    def process_wbs(self, wbs_list):
        """Classifies WBS elements into summary and transaction WBS."""
        summary_wbs = []
        transaction_wbs = set(wbs_list)

        for wbs in wbs_list:
            # Check for any child WBS in the range -01 to -99
            is_parent = False
            for i in range(1, 100):  # Check for -01 to -99 (starting from 1)
                child_wbs = (
                    f"{wbs}-{i:02d}"  # Format the child WBS as -01, -02, ..., -99
                )
                if child_wbs in transaction_wbs:
                    is_parent = True
                    break  # Stop once we find a child WBS for the current parent WBS
            if is_parent:
                summary_wbs.append(wbs)

        return summary_wbs, list(transaction_wbs)


class ExcelFormatter:
    """Handles Excel file formatting."""

    def __init__(self, output_file):
        self.output_file = output_file
        self.workbook = load_workbook(output_file)
        self.worksheet = self.workbook.active
        self.last_col = self.worksheet.max_column
        self.last_row = self.worksheet.max_row

    def apply_formatting(self, summary_wbs_list):
        """Applies all formatting options to the Excel file."""
        self.worksheet.freeze_panes = "D3"
        self.apply_font_style()
        self.apply_currency_formatting()
        self.adjust_column_width()
        self.apply_header_format()  # Ensure this is correctly placed
        self.highlight_summary_wbs(summary_wbs_list)
        self.create_table()
        self.save()

    def apply_font_style(self):
        font = Font(name="Bookman Old Style", size=12)
        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.font = font

    def apply_currency_formatting(self):
        currency_style = NamedStyle(
            name="currency",
            font=Font(name="Bookman Old Style", size=12),
            number_format="₹ #,##0.00;[Red]₹ -#,##0.00",
        )
        for col in range(3, self.last_col + 1):
            col_letter = get_column_letter(col)
            for cell in self.worksheet[col_letter]:
                if isinstance(cell.value, (int, float)):
                    cell.style = currency_style

    def adjust_column_width(self):
        for col in range(1, self.last_col + 1):
            col_letter = get_column_letter(col)
            max_length = max(
                (
                    len(str(cell.value))
                    for cell in self.worksheet[col_letter]
                    if cell.value
                ),
                default=10,
            )
            self.worksheet.column_dimensions[col_letter].width = max_length + 10

    def highlight_summary_wbs(self, summary_wbs_list):
        """Highlight entire row and column for summary WBS."""
        highlight_fill = PatternFill(
            start_color="FFA500", end_color="FFA500", fill_type="solid"
        )

        for row in self.worksheet.iter_rows(
            min_row=1,
            max_row=self.worksheet.max_row,
            min_col=1,
            max_col=self.worksheet.max_column,
        ):
            for cell in row:
                if cell.value in summary_wbs_list:
                    row_num = cell.row
                    col_num = cell.column

                    # Highlight entire row
                    for col in range(1, self.worksheet.max_column + 1):
                        self.worksheet.cell(row=row_num, column=col).fill = (
                            highlight_fill
                        )

    def create_table(self):
        data_range = f"A1:{self.worksheet.cell(row=self.last_row, column=self.last_col).coordinate}"
        table = Table(displayName="Table1", ref=data_range)
        self.worksheet.add_table(table)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

    def apply_header_format(self):
        """Applies header formatting and removes '1' from column headers ending with '1'."""
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        black_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        bold_font = Font(name="Bookman Old Style", size=12, bold=True)

        # Loop through the first row (headers)
        for col in range(1, self.last_col + 1):
            cell = self.worksheet.cell(row=1, column=col)

            # Check if the header ends with '1' and remove it
            if str(cell.value).endswith("1"):
                cell.value = str(cell.value).rstrip("1").strip()

            # Apply the formatting
            cell.fill = yellow_fill
            cell.border = black_border
            cell.font = bold_font

    def save(self):
        self.workbook.save(self.output_file)


def main():
    config = Config()
    logger = SAPReportLogger("BudgetVariance")
    # app = QApplication(sys.argv)
    
    file_path, file_name = FileHandler.get_file_path_and_name()
    if not file_path or not file_name:
        logger.log_info("No file selected. Exiting.")
        return

    input_file = os.path.join(file_path, file_name)
    output_file = os.path.join(file_path, Path(file_name).with_suffix(".xlsx"))
    cleaned_html = os.path.join(file_path, config.TEMP_HTML_FILE)
    master_file = config.MASTER_WBS_FILE

    data_processor = DataProcessor(input_file)
    data_processor.clean_data(cleaned_html)
    summary_wbs_list = data_processor.transform_data(
        cleaned_html, master_file, output_file
    )
    excel_formatter = ExcelFormatter(output_file)
    excel_formatter.apply_formatting(summary_wbs_list)
    logger.log_info(f"Excel file '{output_file}' has been created with the specified formatting.")


if __name__ == "__main__":
    main()
