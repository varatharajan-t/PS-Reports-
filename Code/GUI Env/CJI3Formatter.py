"""
CJI3 Formatter - NLCIL SAP CJI3 Report Formatter
PySide6 Desktop Application
Converts raw SAP CJI3 export (hidden Sheet1) into a formatted workbook
with Summary sheet and per-plant detail sheets.
"""

import sys
import os
from collections import defaultdict
from pathlib import Path

# Load document type descriptions from config
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from config import Config
_DOC_TYPE_MAP = Config.DOCUMENT_TYPES


def _load_project_names() -> dict:
    """Load project definition → name mapping from All_Projects.XLSX."""
    xlsx_path = Path(__file__).resolve().parent.parent / Config.MASTER_PROJECTS_FILE
    if not xlsx_path.exists():
        return {}
    try:
        import openpyxl as _openpyxl
        wb = _openpyxl.load_workbook(str(xlsx_path), data_only=True)
        ws = wb.active
        mapping = {}
        headers = [str(ws.cell(1, c).value).strip() for c in range(1, ws.max_column + 1)]
        try:
            def_idx  = headers.index('Project definition') + 1
            name_idx = headers.index('Name') + 1
        except ValueError:
            return {}
        for r in range(2, ws.max_row + 1):
            proj_def  = ws.cell(r, def_idx).value
            proj_name = ws.cell(r, name_idx).value
            if proj_def and proj_name:
                mapping[str(proj_def).strip()] = str(proj_name).strip()
        return mapping
    except Exception:
        return {}


_PROJECT_NAME_MAP = _load_project_names()

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTextEdit, QProgressBar,
    QFrame, QSizePolicy, QMessageBox
)
from PySide6.QtCore import Qt, QThread, Signal, QObject
from PySide6.QtGui import QFont, QColor, QPalette, QIcon, QDragEnterEvent, QDropEvent

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ─────────────────────────────────────────────
#  STYLE CONSTANTS
# ─────────────────────────────────────────────
POSITIVE_TYPES = {'WE', 'ZA'}


def make_fill(hex_rgb: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_rgb)


def make_font(bold=False, color='000000', name='Arial', size=10) -> Font:
    return Font(bold=bold, color=color, name=name, size=size)


TITLE_FILL  = make_fill('BDD7EE')
HEADER_FILL = make_fill('4472C4')
SUBTOT_FILL = make_fill('D9E2F3')
GTOTAL_FILL = make_fill('4472C4')

TITLE_FONT  = make_font(bold=True,  color='1F4E79')
HEADER_FONT = make_font(bold=True,  color='FFFFFF')
SUBTOT_FONT = make_font(bold=True)
GTOTAL_FONT = make_font(bold=True,  color='FFFFFF')
TICK_FONT   = make_font(bold=True,  color='008000')
CROSS_FONT  = make_font(bold=True,  color='FF0000')
DATA_FONT   = make_font()


# ─────────────────────────────────────────────
#  CORE PROCESSING LOGIC
# ─────────────────────────────────────────────
def status_for_doctype(doc_type: str) -> str:
    return '✔' if doc_type in POSITIVE_TYPES else '✘'


def write_detail_sheet(wb, sheet_name: str, title: str, plant_pivot: dict):
    ws = wb.create_sheet(sheet_name)

    # Row 1 – Title
    ws.row_dimensions[1].height = 30
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = Alignment(vertical='center')
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}1'].fill = TITLE_FILL

    # Row 2 – Headers
    ws.row_dimensions[2].height = 13
    headers = ['WBS Group (First 12)', 'Project Name', 'Document Type', 'Description', 'Sum of Amount', 'Count', 'Status']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 3
    grand_amount, grand_count = 0.0, 0

    for wbs_group in sorted(plant_pivot.keys()):
        doc_types = plant_pivot[wbs_group]
        sub_amount, sub_count = 0.0, 0
        proj_name = _PROJECT_NAME_MAP.get(wbs_group, '')

        pos_types = sorted(d for d in doc_types if d in POSITIVE_TYPES)
        neg_types = sorted(d for d in doc_types if d not in POSITIVE_TYPES)

        for doc_type in pos_types + neg_types:
            amt = round(doc_types[doc_type]['amount'], 2)
            cnt = doc_types[doc_type]['count']
            status = status_for_doctype(doc_type)
            desc = _DOC_TYPE_MAP.get(doc_type, '')

            ws.cell(row=row, column=1, value=wbs_group).font = DATA_FONT
            ws.cell(row=row, column=2, value=proj_name).font = DATA_FONT
            ws.cell(row=row, column=3, value=doc_type).font = DATA_FONT
            ws.cell(row=row, column=4, value=desc).font = DATA_FONT
            c5 = ws.cell(row=row, column=5, value=amt)
            c5.font = DATA_FONT
            c5.number_format = '#,##0.00'
            c6 = ws.cell(row=row, column=6, value=cnt)
            c6.font = DATA_FONT
            c6.number_format = '#,##0'
            c7 = ws.cell(row=row, column=7, value=status)
            c7.font = TICK_FONT if status == '✔' else CROSS_FONT
            c7.alignment = Alignment(horizontal='center')

            ws.row_dimensions[row].height = 15.5
            sub_amount += amt
            sub_count += cnt
            row += 1

        # Subtotal
        sub_label = ws.cell(row=row, column=1, value=f'{wbs_group} Total')
        sub_label.font = SUBTOT_FONT
        sub_label.fill = SUBTOT_FILL
        for col in range(2, 8):
            ws.cell(row=row, column=col).fill = SUBTOT_FILL
        c5 = ws.cell(row=row, column=5, value=round(sub_amount, 2))
        c5.font = SUBTOT_FONT
        c5.fill = SUBTOT_FILL
        c5.number_format = '#,##0.00'
        c6 = ws.cell(row=row, column=6, value=sub_count)
        c6.font = SUBTOT_FONT
        c6.fill = SUBTOT_FILL
        c6.number_format = '#,##0'
        ws.row_dimensions[row].height = 16
        grand_amount += sub_amount
        grand_count += sub_count
        row += 1

    # Grand Total
    gt = ws.cell(row=row, column=1, value='Grand Total')
    gt.font = GTOTAL_FONT
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.fill = GTOTAL_FILL
        cell.font = GTOTAL_FONT
    c5 = ws.cell(row=row, column=5, value=round(grand_amount, 2))
    c5.number_format = '#,##0.00'
    c5.fill = GTOTAL_FILL
    c5.font = GTOTAL_FONT
    c6 = ws.cell(row=row, column=6, value=grand_count)
    c6.number_format = '#,##0'
    c6.fill = GTOTAL_FILL
    c6.font = GTOTAL_FONT
    ws.row_dimensions[row].height = 16

    # Column widths
    ws.column_dimensions['A'].width = 32.7
    ws.column_dimensions['B'].width = 45.0
    ws.column_dimensions['C'].width = 14.0
    ws.column_dimensions['D'].width = 27.3
    ws.column_dimensions['E'].width = 27.3
    ws.column_dimensions['F'].width = 12.7
    ws.column_dimensions['G'].width = 10.0


def write_summary_sheet(wb, title: str, summary_data: dict):
    ws = wb.create_sheet('Summary', 0)

    # Row 1 – Title
    ws.row_dimensions[1].height = 30
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = Alignment(vertical='center')
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}1'].fill = TITLE_FILL

    # Row 2 – Headers
    ws.row_dimensions[2].height = 13
    headers = ['WBS Group', 'Project Name', 'Status', 'Sum of Amount', 'Count']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 3
    grand_amount, grand_count = 0.0, 0

    for wbs_group in sorted(summary_data.keys()):
        statuses = summary_data[wbs_group]
        sub_amount, sub_count = 0.0, 0
        proj_name = _PROJECT_NAME_MAP.get(wbs_group, '')

        for status in ['✔', '✘']:
            if status not in statuses:
                continue
            amt = round(statuses[status]['amount'], 2)
            cnt = statuses[status]['count']

            ws.cell(row=row, column=1, value=wbs_group).font = DATA_FONT
            ws.cell(row=row, column=2, value=proj_name).font = DATA_FONT
            c3 = ws.cell(row=row, column=3, value=status)
            c3.font = TICK_FONT if status == '✔' else CROSS_FONT
            c3.alignment = Alignment(horizontal='center')
            c4 = ws.cell(row=row, column=4, value=amt)
            c4.font = DATA_FONT
            c4.number_format = '#,##0.00'
            c5 = ws.cell(row=row, column=5, value=cnt)
            c5.font = DATA_FONT
            c5.number_format = '#,##0'

            ws.row_dimensions[row].height = 15.5
            sub_amount += amt
            sub_count += cnt
            row += 1

        # Subtotal
        sub_label = ws.cell(row=row, column=1, value=f'{wbs_group} Total')
        sub_label.font = SUBTOT_FONT
        sub_label.fill = SUBTOT_FILL
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = SUBTOT_FILL
        c4 = ws.cell(row=row, column=4, value=round(sub_amount, 2))
        c4.font = SUBTOT_FONT
        c4.fill = SUBTOT_FILL
        c4.number_format = '#,##0.00'
        c5 = ws.cell(row=row, column=5, value=sub_count)
        c5.font = SUBTOT_FONT
        c5.fill = SUBTOT_FILL
        c5.number_format = '#,##0'
        ws.row_dimensions[row].height = 16
        grand_amount += sub_amount
        grand_count += sub_count
        row += 1

    # Grand Total
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.fill = GTOTAL_FILL
        cell.font = GTOTAL_FONT
    ws.cell(row=row, column=1, value='Grand Total')
    c4 = ws.cell(row=row, column=4, value=round(grand_amount, 2))
    c4.number_format = '#,##0.00'
    c4.fill = GTOTAL_FILL
    c4.font = GTOTAL_FONT
    c5 = ws.cell(row=row, column=5, value=grand_count)
    c5.number_format = '#,##0'
    c5.fill = GTOTAL_FILL
    c5.font = GTOTAL_FONT
    ws.row_dimensions[row].height = 16

    ws.column_dimensions['A'].width = 32.7
    ws.column_dimensions['B'].width = 45.0
    ws.column_dimensions['C'].width = 10.0
    ws.column_dimensions['D'].width = 27.3
    ws.column_dimensions['E'].width = 12.7


def process_cji3(input_path: str, output_path: str, progress_cb=None, log_cb=None) -> dict:
    def log(msg):
        if log_cb:
            log_cb(msg)

    log(f"Loading workbook: {input_path}")
    wb = load_workbook(input_path, data_only=True)

    if 'Sheet1' not in wb.sheetnames:
        raise ValueError("Input file does not contain a sheet named 'Sheet1'.")

    ws1 = wb['Sheet1']

    # ── Read Sheet1 ──────────────────────────────
    log("Reading Sheet1 data...")
    plant_data   = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'amount': 0.0, 'count': 0})))
    summary_data = defaultdict(lambda: defaultdict(lambda: {'amount': 0.0, 'count': 0}))
    fy_counter   = defaultdict(int)
    row_count    = 0

    all_rows = list(ws1.iter_rows(min_row=2, values_only=True))
    total_rows = len(all_rows)

    for i, row in enumerate(all_rows):
        if progress_cb and i % 500 == 0:
            progress_cb(int(i / total_rows * 60))  # 0–60% for reading

        if len(row) < 15:
            continue
        wbs      = row[1]   # col B
        doc_type = row[4]   # col E
        amount   = row[8]   # col I
        fy       = row[14]  # col O

        if not wbs or not doc_type:
            continue
        if not isinstance(amount, (int, float)):
            continue

        wbs = str(wbs).strip()
        doc_type = str(doc_type).strip()

        if len(wbs) < 8:
            continue

        plant     = wbs[5:8]
        wbs_group = wbs[:12]
        status    = status_for_doctype(doc_type)

        plant_data[plant][wbs_group][doc_type]['amount'] += amount
        plant_data[plant][wbs_group][doc_type]['count']  += 1

        summary_data[wbs_group][status]['amount'] += amount
        summary_data[wbs_group][status]['count']  += 1

        if fy:
            fy_counter[fy] += 1

        row_count += 1

    log(f"Processed {row_count:,} data rows.")

    # Detect fiscal year
    fiscal_year = max(fy_counter, key=fy_counter.get) if fy_counter else 2025
    fiscal_year = int(fiscal_year)
    title = f"Analysis of Project Actual for the Year {fiscal_year} -{fiscal_year + 1}."
    log(f"Fiscal year detected: {fiscal_year}")

    # ── Build output workbook ────────────────────
    log("Building output workbook...")
    out_wb = openpyxl.Workbook()
    for sheet_name in list(out_wb.sheetnames):
        del out_wb[sheet_name]

    if progress_cb:
        progress_cb(65)

    write_summary_sheet(out_wb, title, summary_data)
    log("  → Summary sheet written.")

    if progress_cb:
        progress_cb(70)

    sorted_plants = sorted(plant_data.keys())
    for idx, plant in enumerate(sorted_plants):
        write_detail_sheet(out_wb, plant, title, plant_data[plant])
        log(f"  → Plant sheet '{plant}' written.")
        if progress_cb:
            progress_cb(70 + int((idx + 1) / len(sorted_plants) * 25))

    if progress_cb:
        progress_cb(95)

    out_wb.save(output_path)
    log(f"Saved: {output_path}")

    if progress_cb:
        progress_cb(100)

    return {
        'row_count':   row_count,
        'fiscal_year': fiscal_year,
        'plants':      sorted_plants,
        'output_path': output_path,
    }


# ─────────────────────────────────────────────
#  WORKER THREAD
# ─────────────────────────────────────────────
class WorkerSignals(QObject):
    progress = Signal(int)
    log      = Signal(str)
    finished = Signal(dict)
    error    = Signal(str)


class ProcessWorker(QThread):
    def __init__(self, input_path: str, output_path: str):
        super().__init__()
        self.input_path  = input_path
        self.output_path = output_path
        self.signals     = WorkerSignals()

    def run(self):
        try:
            result = process_cji3(
                self.input_path,
                self.output_path,
                progress_cb=self.signals.progress.emit,
                log_cb=self.signals.log.emit,
            )
            self.signals.finished.emit(result)
        except Exception as e:
            self.signals.error.emit(str(e))


# ─────────────────────────────────────────────
#  DROP ZONE WIDGET
# ─────────────────────────────────────────────
class DropZone(QLabel):
    file_dropped = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setAlignment(Qt.AlignCenter)
        self.setMinimumHeight(120)
        self._set_idle_style()
        self.setText("Drop CJI3 .xlsx file here\nor click Browse")

    def _set_idle_style(self):
        self.setStyleSheet("""
            QLabel {
                border: 2px dashed #4472C4;
                border-radius: 8px;
                background: #EEF4FB;
                color: #4472C4;
                font-size: 14px;
                padding: 20px;
            }
        """)

    def _set_hover_style(self):
        self.setStyleSheet("""
            QLabel {
                border: 2px solid #1F4E79;
                border-radius: 8px;
                background: #BDD7EE;
                color: #1F4E79;
                font-size: 14px;
                padding: 20px;
            }
        """)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if any(u.toLocalFile().lower().endswith('.xlsx') for u in urls):
                event.acceptProposedAction()
                self._set_hover_style()
                return
        event.ignore()

    def dragLeaveEvent(self, event):
        self._set_idle_style()

    def dropEvent(self, event: QDropEvent):
        self._set_idle_style()
        for url in event.mimeData().urls():
            path = url.toLocalFile()
            if path.lower().endswith('.xlsx'):
                self.file_dropped.emit(path)
                break


# ─────────────────────────────────────────────
#  MAIN WINDOW
# ─────────────────────────────────────────────
class CJI3FormatterWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CJI3 Formatter  –  NLCIL SAP Report Tool")
        self.setMinimumSize(720, 620)
        self.input_path  = None
        self.output_path = None
        self.worker      = None
        self._build_ui()

    # ── UI Layout ──────────────────────────────
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setContentsMargins(20, 20, 20, 20)
        root.setSpacing(12)

        # Header bar
        header = QLabel("  CJI3 Report Formatter")
        header.setFixedHeight(48)
        header.setStyleSheet("""
            QLabel {
                background: #1F4E79;
                color: white;
                font-size: 18px;
                font-weight: bold;
                border-radius: 6px;
                padding-left: 10px;
            }
        """)
        root.addWidget(header)

        sub = QLabel("Transforms raw SAP CJI3 export → formatted NLCIL workbook (Summary + per-plant sheets)")
        sub.setStyleSheet("color: #555; font-size: 11px;")
        root.addWidget(sub)

        # ── Drop zone ─────────────────────────
        self.drop_zone = DropZone()
        self.drop_zone.file_dropped.connect(self._set_input_file)
        root.addWidget(self.drop_zone)

        # ── File row ──────────────────────────
        file_row = QHBoxLayout()
        self.lbl_input = QLabel("No file selected.")
        self.lbl_input.setStyleSheet("color: #333; font-size: 11px;")
        self.lbl_input.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        file_row.addWidget(self.lbl_input)

        btn_browse = QPushButton("Browse…")
        btn_browse.setFixedWidth(90)
        btn_browse.setStyleSheet(self._btn_style('#4472C4'))
        btn_browse.clicked.connect(self._browse_input)
        file_row.addWidget(btn_browse)
        root.addLayout(file_row)

        # ── Output row ────────────────────────
        out_row = QHBoxLayout()
        out_lbl = QLabel("Save to:")
        out_lbl.setFixedWidth(55)
        out_lbl.setStyleSheet("color: #333; font-size: 11px;")
        out_row.addWidget(out_lbl)

        self.lbl_output = QLabel("(same folder as input file)")
        self.lbl_output.setStyleSheet("color: #666; font-size: 11px;")
        self.lbl_output.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        out_row.addWidget(self.lbl_output)

        btn_out = QPushButton("Change…")
        btn_out.setFixedWidth(90)
        btn_out.setStyleSheet(self._btn_style('#666'))
        btn_out.clicked.connect(self._browse_output)
        out_row.addWidget(btn_out)
        root.addLayout(out_row)

        # ── Divider ───────────────────────────
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("color: #CCC;")
        root.addWidget(line)

        # ── Log ───────────────────────────────
        log_lbl = QLabel("Processing Log")
        log_lbl.setStyleSheet("font-weight: bold; color: #1F4E79; font-size: 12px;")
        root.addWidget(log_lbl)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setFont(QFont("Consolas", 9))
        self.log_box.setStyleSheet("""
            QTextEdit {
                background: #F7F9FC;
                border: 1px solid #CCC;
                border-radius: 4px;
                color: #222;
            }
        """)
        self.log_box.setMinimumHeight(160)
        root.addWidget(self.log_box)

        # ── Progress ──────────────────────────
        self.progress = QProgressBar()
        self.progress.setValue(0)
        self.progress.setTextVisible(True)
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 1px solid #CCC;
                border-radius: 4px;
                text-align: center;
                height: 18px;
            }
            QProgressBar::chunk {
                background: #4472C4;
                border-radius: 3px;
            }
        """)
        root.addWidget(self.progress)

        # ── Action buttons ────────────────────
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        self.btn_run = QPushButton("▶  Format CJI3")
        self.btn_run.setFixedHeight(42)
        self.btn_run.setEnabled(False)
        self.btn_run.setStyleSheet(self._btn_style('#1F4E79', text_color='white', font_size=13))
        self.btn_run.clicked.connect(self._run)
        btn_row.addWidget(self.btn_run)

        btn_clear = QPushButton("Clear Log")
        btn_clear.setFixedHeight(42)
        btn_clear.setFixedWidth(110)
        btn_clear.setStyleSheet(self._btn_style('#888'))
        btn_clear.clicked.connect(self.log_box.clear)
        btn_row.addWidget(btn_clear)

        self.btn_open = QPushButton("Open Output")
        self.btn_open.setFixedHeight(42)
        self.btn_open.setFixedWidth(120)
        self.btn_open.setEnabled(False)
        self.btn_open.setStyleSheet(self._btn_style('#008000'))
        self.btn_open.clicked.connect(self._open_output)
        btn_row.addWidget(self.btn_open)

        root.addLayout(btn_row)

        # Status bar
        self.statusBar().showMessage("Ready — select a CJI3 .xlsx file to begin.")

    # ── Helpers ────────────────────────────────
    @staticmethod
    def _btn_style(bg, text_color='white', font_size=11):
        return f"""
            QPushButton {{
                background: {bg};
                color: {text_color};
                border: none;
                border-radius: 5px;
                font-size: {font_size}px;
                font-weight: bold;
                padding: 4px 12px;
            }}
            QPushButton:hover {{ background: {bg}CC; }}
            QPushButton:disabled {{ background: #AAA; color: #EEE; }}
        """

    def _set_input_file(self, path: str):
        self.input_path = path
        self.lbl_input.setText(path)
        # Default output to same folder
        folder = str(Path(path).parent)
        default_out = str(Path(folder) / "CJI3-Formatted.xlsx")
        self.output_path = default_out
        self.lbl_output.setText(default_out)
        self.btn_run.setEnabled(True)
        self.btn_open.setEnabled(False)
        self.progress.setValue(0)
        self.statusBar().showMessage(f"Loaded: {Path(path).name}")

    def _browse_input(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select CJI3 Excel File", "", "Excel Files (*.xlsx)"
        )
        if path:
            self._set_input_file(path)

    def _browse_output(self):
        default = self.output_path or "CJI3-Formatted.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Formatted File As", default, "Excel Files (*.xlsx)"
        )
        if path:
            self.output_path = path
            self.lbl_output.setText(path)

    def _log(self, msg: str):
        self.log_box.append(msg)

    def _run(self):
        if not self.input_path or not self.output_path:
            return

        self.btn_run.setEnabled(False)
        self.btn_open.setEnabled(False)
        self.progress.setValue(0)
        self.log_box.clear()
        self._log("═" * 60)
        self._log("CJI3 Formatter  –  NLCIL")
        self._log("═" * 60)

        self.worker = ProcessWorker(self.input_path, self.output_path)
        self.worker.signals.log.connect(self._log)
        self.worker.signals.progress.connect(self.progress.setValue)
        self.worker.signals.finished.connect(self._on_finished)
        self.worker.signals.error.connect(self._on_error)
        self.worker.start()
        self.statusBar().showMessage("Processing…")

    def _on_finished(self, result: dict):
        plants_str = ', '.join(result['plants']) if result['plants'] else '—'
        self._log("═" * 60)
        self._log(f"✔ Done!")
        self._log(f"   Rows processed : {result['row_count']:,}")
        self._log(f"   Fiscal year     : {result['fiscal_year']}–{result['fiscal_year']+1}")
        self._log(f"   Plants found    : {len(result['plants'])}  ({plants_str})")
        self._log(f"   Output saved to : {result['output_path']}")
        self._log("═" * 60)

        self.btn_run.setEnabled(True)
        self.btn_open.setEnabled(True)
        self.statusBar().showMessage(f"Complete — {len(result['plants'])} plant sheet(s) generated.")

    def _on_error(self, msg: str):
        self._log(f"✘ ERROR: {msg}")
        self.btn_run.setEnabled(True)
        self.statusBar().showMessage("Error — see log for details.")
        QMessageBox.critical(self, "Processing Error", msg)

    def _open_output(self):
        if self.output_path and Path(self.output_path).exists():
            import subprocess, platform
            if platform.system() == 'Windows':
                os.startfile(self.output_path)
            elif platform.system() == 'Darwin':
                subprocess.call(['open', self.output_path])
            else:
                subprocess.call(['xdg-open', self.output_path])


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────
def main():
    """
    Entry point — safe to call from within an existing QApplication
    (e.g. from MenuPySide1.py).  If no QApplication exists yet,
    one is created and the event loop is started.
    """
    app = QApplication.instance()
    standalone = app is None
    if standalone:
        app = QApplication(sys.argv)
        app.setStyle('Fusion')

    font = QFont("Segoe UI", 10)
    app.setFont(font)

    win = CJI3FormatterWindow()
    win.show()

    if standalone:
        sys.exit(app.exec())


if __name__ == '__main__':
    main()
