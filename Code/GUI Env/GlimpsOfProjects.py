"""
************************************************************************************************************
SAP PROJECT SYSTEM - PROJECT ANALYTICS AND CROSS-TABULATION DASHBOARD
************************************************************************************************************

SYSTEM REQUIREMENTS:
- Python 3.7+
- pandas >= 1.3.0
- openpyxl >= 3.0.0 (with chart support)
- PySide6 >= 6.0.0
- re (built-in)
- os, sys, pathlib (built-in)

PURPOSE:
Advanced analytics tool for SAP project data that creates interactive dashboards with
cross-tabulation reports, dynamic charts, and real-time data validation. Processes
project identifiers to generate company-wise and project-type-wise analytics with
sophisticated visualization capabilities.

UNIQUE FEATURES:
- Intelligent Project ID parsing using regex patterns
- Dynamic cross-tabulation with configurable dimensions
- Interactive 3D bar charts with gradient effects
- Data validation dropdowns for dynamic filtering
- Company and project type mapping with predefined dictionaries
- Multi-dimensional analysis capabilities

ANALYTICS CAPABILITIES:
1. Project Classification: Automatic categorization by company and type
2. Cross-Tabulation: Matrix analysis of projects across multiple dimensions
3. Dynamic Visualization: Real-time chart updates based on user selection
4. Interactive Filtering: Dropdown-based data exploration
5. Statistical Summary: Comprehensive project counting and grouping

DATA PROCESSING WORKFLOW:
1. Pattern Recognition: Extracts project metadata using regex
2. Data Mapping: Applies business logic for company/type classification
3. Cross-Tabulation: Generates statistical summary matrices
4. Visualization: Creates interactive charts with advanced styling
5. Dashboard Creation: Assembles complete analytics workbook

COMPLEX ALGORITHMS:
- Regex-based project ID parsing: r"PRJ\s+([A-Z0-9-]+)"
- Dynamic chart reference formulas using INDEX-MATCH patterns
- Color gradient application for enhanced visual appeal
- Real-time data validation with Excel formula integration

OUTPUT SPECIFICATIONS:
- Excel workbook with multiple analysis perspectives
- Interactive 3D charts with customizable data series
- Data validation controls for dynamic filtering
- Professional styling with color-coded visualizations

AUTHOR: [Varatha Rajan T]
CREATION DATE: [12/02/2024]
LAST MODIFIED: [24/09/2025]
************************************************************************************************************
"""

import re
import pandas as pd
import os
import sys
from pathlib import Path
from PySide6.QtWidgets import QFileDialog, QApplication
from config import Config
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart3D, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.axis import ChartLines


class ProjectDataProcessor:
    """
    Advanced project data extraction and processing engine.

    This class provides comprehensive project analytics capabilities including:
    - Intelligent parsing of SAP project identifiers using regex patterns
    - Business logic mapping for company codes and project types
    - Statistical analysis and cross-tabulation generation
    - Data transformation for visualization readiness

    Business Logic Mappings:
        Company Codes:
            NL -> NLCIL (Neyveli Lignite Corporation India Limited)
            NT -> NTPL (NLC Tamil Nadu Power Limited)
            NU -> NUPPL (NLC Upper Stage Power Projects)
            NR -> NIRL (Neyveli Infrastructure Resources Limited)
            NG -> NIGEL (Neyveli Green Energy Limited)

        Project Types:
            S -> Service Projects
            I -> Income/Revenue Projects
            N -> Non-Plan Projects
            C -> Capital Expenditure Projects
            E -> Excetra Projects
            F -> Feasibility Studies
            R -> Research & Development
            O -> Operational Expenditure
            M -> Material Management

    Pattern Recognition:
        Uses regex pattern "PRJ\s+([A-Z0-9-]+)" to extract project IDs
        Analyzes first two characters for company identification
        Uses fourth character for project type classification
    """

    def __init__(self):
        """Initialize processor with predefined business logic mappings."""
        self.config = Config()
        self.company_codes = self.config.COMPANY_CODES
        self.project_types = self.config.PROJECT_TYPES

    def get_file_path_and_name(self):
        """Opens a file dialog and returns the selected file path and name."""
        filepath = QFileDialog.getOpenFileName(
            None, "Select a DAT file", "", "DAT Files(*.DAT)"
        )
        return os.path.split(filepath[0]) if filepath[0] else (None, None)

    def extract_project_data(self, file_path):
        """Extracts project data from the given DAT file."""
        project_id_pattern = r"PRJ\s+([A-Z0-9-]+)"
        project_data = []

        with open(file_path, "r", encoding="utf-8", errors="ignore") as file:
            for line in file:
                match = re.search(project_id_pattern, line)
                if match:
                    project_id = match.group(1)
                    company_desc = self.company_codes.get(
                        project_id[:2], "Unknown Company"
                    )
                    project_desc = self.project_types.get(
                        project_id[3], "Unknown Project Type"
                    )
                    project_data.append(
                        {
                            "Project Type": project_desc,
                            "Company": company_desc,
                            "Project ID": project_id,
                        }
                    )

        return project_data

    def generate_cross_tab(self, project_data):
        """Generates a cross-tab report from project data."""
        df = pd.DataFrame(project_data)
        return (
            pd.crosstab(
                df["Project Type"], df["Company"], margins=True, margins_name="Total"
            )
            if not df.empty
            else None
        )


class ExcelStyler:
    """Handles styling and chart creation in the Excel report."""

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path)
        self.ws = self.wb.active
        self.column_colors = {
            "A": ("DCE6F1", "B0C4DE"),  # Project Type: Light & Medium Blue
            "B": ("EAF1DD", "A6CE39"),  # NIGEL: Light & Medium Green
            "C": ("FDEDEC", "E15B64"),  # NIRL: Light & Medium Red
            "D": ("EDE3F1", "9B6ED2"),  # NLCIL: Light & Medium Purple
            "E": ("E0F2F1", "68C3A3"),  # NTPL: Light & Medium Teal
            "F": ("F3E5D8", "C69C6D"),  # NUPPL: Light & Medium Brown
            "G": ("EAEAEA", "D1D3D4"),  # Total Column: Light & Medium Gray
        }

    def apply_styling(self):
        """Applies formatting and colors to the Excel report."""
        self.ws.column_dimensions["A"].width = 15

        for col_letter, (header_color, cell_color) in self.column_colors.items():
            header_fill = PatternFill(
                start_color=header_color, end_color=header_color, fill_type="solid"
            )
            cell_fill = PatternFill(
                start_color=cell_color, end_color=cell_color, fill_type="solid"
            )

            for row in range(1, self.ws.max_row + 1):
                cell = self.ws[f"{col_letter}{row}"]
                cell.fill = header_fill if row == 1 else cell_fill
                if row == 1:
                    cell.font = Font(bold=True, color="000000")

    def add_data_validation(self):
        """Adds data validation dropdown for company selection."""
        dv = DataValidation(type="list", formula1="B1:F1", allow_blank=False)
        dv.prompt, dv.promptTitle = (
            "Please select a company from the list",
            "Company List",
        )
        self.ws.add_data_validation(dv)
        self.ws["A10"].value = "Select Company"
        dv.add(self.ws["A10"])

    def add_dynamic_reference_section(self):
        """Adds a section below the cross-tab table to dynamically reference the selected company's data."""
        # Add a header for the dynamic section
        self.ws["A10"].font = Font(bold=True)
        self.ws["B10"].value = "No. of Projects"
        self.ws["B10"].font = Font(bold=True)

        # Using a loop
        for row in range(11, 17):
            self.ws[f"A{row}"].value = f"=A{row - 9}"

        # Add formulas to dynamically reference the selected company's data
        for row in range(11, 17):  # Rows 11 to 16 for project types
            self.ws[f"B{row}"].value = (
                f"=INDEX($B$2:$F$7, ROW()-10, MATCH($A$10, $B$1:$F$1, 0))"
            )

    def add_chart(self):
        """Adds a dynamic 3D Bar Chart to the Excel sheet."""
        chart = BarChart3D()
        # chart.title, chart.x_axis.title, chart.y_axis.title = "Project Type Vs No of Projects", "Project Type", "No. of Projects"
        chart.title = "Project Type Vs No of Projects"

        selected_col = Reference(self.ws, min_col=1, min_row=11, max_row=16)
        values = Reference(self.ws, min_col=2, min_row=11, max_row=16)
        chart.add_data(values, titles_from_data=False)
        chart.set_categories(selected_col)

        # Remove the legend
        chart.legend = None

        # Add data labels on top of the columns
        for series in chart.series:
            series.dLbls = openpyxl.chart.label.DataLabelList()
            series.dLbls.showVal = True  # Show values as data labels
            series.dLbls.showSerName = False  # Hide series names
            series.dLbls.showCatName = False  # Hide category names

        # Configure X-axis to display integers only
        chart.x_axis.majorGridlines = None
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.majorGridlines = ChartLines()
        chart.x_axis.delete = False

        chart.y_axis.numFmt = "0"  # Format X-axis values as integers
        chart.y_axis.tickLblPos = "nextTo"
        chart.y_axis.majorGridlines = ChartLines()
        chart.y_axis.delete = False

        # **Apply Grey to White Gradient Effect**
        chart.style = 34  # Predefined style with a gradient effect

        # Assign different colors to each bar
        colors = [
            "00FF00",
            "FF0000",
            "0000FF",
            "FFFF00",
            "FF00FF",
            "00FFFF",
        ]  # Green, Red, Blue, Yellow, Magenta, Cyan
        slices = [DataPoint(idx=i) for i in range(6)]
        for i, series in enumerate(chart.series):
            # Assign colors cyclically
            slices[i].graphicalProperties.solidFill = colors[i]

        # Make the chart bigger
        chart.width = 16  # Set chart width (in units of cell width)
        chart.height = 10  # Set chart height (in units of cell height)

        self.ws.add_chart(chart, "I3")

    def save_workbook(self):
        """Saves the workbook."""
        self.wb.save(self.file_path)
        print(f"Excel formatting applied successfully. Saved to: {self.file_path}")


def main():
    # Use existing QApplication instance if available (when called from menu)
    app = QApplication.instance()
    if app is None:
        app = QApplication(sys.argv)

    processor = ProjectDataProcessor()
    file_path, file_name = processor.get_file_path_and_name()

    if file_path:
        input_file = os.path.join(file_path, file_name)
        output_file = os.path.join(file_path, "CrossTab_Report.xlsx")
        project_data = processor.extract_project_data(input_file)
        report = processor.generate_cross_tab(project_data)

        if report is not None:
            report.to_excel(output_file)
            styler = ExcelStyler(output_file)
            styler.apply_styling()
            styler.add_data_validation()
            styler.add_dynamic_reference_section()
            styler.add_chart()
            styler.save_workbook()


if __name__ == "__main__":
    main()
