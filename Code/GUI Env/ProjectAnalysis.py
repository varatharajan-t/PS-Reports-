import pandas as pd, os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart3D, Reference
from config import Config

config = Config()




# Function to process files
def process_file(file_path, encoding, extract_pattern, last_row_prefix):
    df = pd.read_csv(file_path, delimiter="\t", header=[0, 1], encoding=encoding)
    df = df[~df.iloc[:, 0].str.startswith(last_row_prefix, na=False)]
    if os.path.basename(file_path) == config.PROJECT_ANALYSIS_FILES["budget"]:
        df[["ProjectName", "ProjectID"]] = (
            df[("Unnamed: 0_level_0", "Object")].str.extract(extract_pattern).fillna("")
        )
    else:
        df[["ProjectID", "ProjectName"]] = (
            df[("Unnamed: 0_level_0", "Object")].str.extract(extract_pattern).fillna("")
        )
    df["ProjectID"], df["ProjectName"] = (
        df["ProjectID"].str.strip(),
        df["ProjectName"].str.strip(),
    )
    return df  # df.drop(columns=[df.columns[1]])


# Process files
df_budget = process_file(
    os.path.join("data", config.PROJECT_ANALYSIS_FILES["budget"]),
    "iso-8859-1",
    config.PROJECT_ANALYSIS_REGEX["budget"],
    "Result",
)
df_actual = process_file(
    os.path.join("data", config.PROJECT_ANALYSIS_FILES["plan"]),
    "iso-8859-1",
    config.PROJECT_ANALYSIS_REGEX["plan"],
    "Result",
)
# print(df_actual,df_actual.columns)
# print(df_budget,df_budget.columns)

# Merge data on 'ProjectID'
df_combined = pd.merge(
    df_budget,
    df_actual,
    on=[("ProjectID", "")],
    how="outer",
    suffixes=("_Budget", "_Actual"),
)

# Create Excel workbook
wb = openpyxl.Workbook()
del wb["Sheet"]


# Function to write DataFrame to sheet
def write_df_to_sheet(df, sheet_name):
    ws = wb.create_sheet(sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    return ws


ws_budget = write_df_to_sheet(df_budget, "Budget")
ws_actual = write_df_to_sheet(df_actual, "Actual")
ws_data = write_df_to_sheet(df_combined, "Data")
ws_analysis = wb.create_sheet("Analysis")

# Create data validation for 'ProjectID' - using dynamic range based on actual data
# Find the column with ProjectID (assuming it's in column AL)
project_id_col = "AL"  # Adjust if needed
last_row = ws_budget.max_row
last_col_letter = openpyxl.utils.get_column_letter(ws_budget.max_column)

dv_project = DataValidation(
    type="list", formula1=f"=Budget!${project_id_col}$3:${project_id_col}${last_row}", allow_blank=False
)
dv_header = DataValidation(type="list", formula1=f"=Budget!$D$1:${last_col_letter}$1", allow_blank=False)
ws_analysis.add_data_validation(dv_project)
ws_analysis.add_data_validation(dv_header)
ws_analysis["A1"].value = "Select ProjectID:"
dv_project.add(ws_analysis["B1"])
ws_analysis["A2"].value = "Select Year range:"
dv_header.add(ws_analysis["B2"])

# Add formula for table - using dynamic ranges
headers = ["Plan", "Budget", "Actual", "Commitment", "Assigned", "Available"]
for i, header in enumerate(headers, start=3):
    ws_analysis.cell(row=3, column=i, value=header)
    ws_analysis.cell(
        row=4,
        column=i,
        value=f'=IFERROR(INDEX(Budget!$D$2:${last_col_letter}${last_row}, MATCH($B$1, Budget!$C$2:$C${last_row}, 0), MATCH($B$2, Budget!$D$1:${last_col_letter}$1, 0)), "")',
    )


# Function to create a chart
def create_chart(ws, title, data_col, position):
    chart = BarChart3D()
    chart.title = title
    data = Reference(ws, min_col=data_col, min_row=4, max_row=9)
    categories = Reference(ws, min_col=2, min_row=4, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, position)


# Add charts
chart_positions = [
    ("D10", 3, "Plan Vs Budget"),
    ("D25", 4, "Budget Vs Actual"),
    ("M10", 5, "Budget Vs Commitment"),
    ("M25", 6, "Budget Vs Available"),
]
for pos, col, title in chart_positions:
    create_chart(ws_analysis, title, col, pos)

# Save workbook
wb.save(config.PROJECT_ANALYSIS_OUTPUT)
