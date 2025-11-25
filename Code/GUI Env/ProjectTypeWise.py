# ***********************************************************************************
# This program reads the CN42N report data of a company and
# groups that into an Excel sheets and then do necessary formatting.
# Author: Varatharajan T. Date of Creation: 23/05/2024 Modiifed on: 23/05/2024
# ***********************************************************************************

import pandas as pd
import openpyxl, os
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import NamedStyle, PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from tkinter import filedialog, messagebox
import tkinter as tk
from config import Config


class WBSProcessor:
    def __init__(self, input_file, output_file):
        self.input_file = input_file
        self.output_file = output_file
        self.df = None
        self.summary = None
        self.config = Config()

    def read_excel(self):
        self.df = pd.read_excel(self.input_file)

    def group_project_definitions(self):
        self.df["Group"] = self.df["Project definition"].apply(
            lambda x: "-".join(x.split("-")[:2])
        )

    def map_values(self):
        mapping_dict = {**self.config.COMPANY_CODES, **self.config.PROJECT_TYPES}
        self.df["Group"] = self.df["Group"].apply(
            lambda x: "-".join([mapping_dict.get(val, val) for val in x.split("-")])
        )

    def create_summary(self):
        summary = (
            self.df.groupby("Group")["Project definition"]
            .nunique()
            .reset_index(name="Unique Projects")
        )
        self.summary = summary.rename(
            columns={"Group": "Project", "Unique Projects": "Number of Unique Projects"}
        )

    def add_hyperlinks(self, workbook):
        # Add hyperlinks in the summary sheet
        summary_sheet = workbook["Summary"]
        for row in range(2, summary_sheet.max_row + 1):
            project_name = summary_sheet.cell(row=row, column=1).value
            project_count = summary_sheet.cell(row=row, column=2).value
            link = Hyperlink(
                ref=f"B{row}",
                location=f"'{project_name}'!A1",
                display=str(project_count),
            )
            summary_sheet[f"B{row}"].hyperlink = link
            summary_sheet[f"B{row}"].style = "Hyperlink"

    def write_to_excel(self, output_file):
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            self.df.to_excel(writer, index=False, sheet_name="ProjectsView")
            self.summary.to_excel(writer, index=False, sheet_name="Summary")
            grouped = self.df.groupby("Group")
            for name, group in grouped:
                group.to_excel(writer, index=False, sheet_name=name)

        workbook = openpyxl.load_workbook(self.output_file)
        self.add_hyperlinks(workbook)
        workbook.save(self.output_file)


class ExcelFormatter:
    def __init__(self, output_file, workbook, sheet_name):
        self.file_name = output_file
        self.workbook = workbook
        self.worksheet = self.workbook[sheet_name]
        self.sheet_name = sheet_name

    def apply_common_style(self):
        font = Font(name="Bookman Old Style", size=12)
        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.font = font

    def auto_adjust_columns(self):
        for column in self.worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.5
            self.worksheet.column_dimensions[column[0].column_letter].width = (
                adjusted_width
            )

    def add_table(self):
        last_row = self.worksheet.max_row
        last_col = self.worksheet.max_column
        data_range = (
            f"A1:{self.worksheet.cell(row=last_row, column=last_col).coordinate}"
        )
        table = Table(displayName="Table" + self.sheet_name, ref=data_range)
        self.worksheet.add_table(table)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

    def apply_header_style(self):
        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        for cell in self.worksheet[1]:
            cell.fill = white_fill

    def freeze_panes(self):
        self.worksheet.freeze_panes = "A2"

    def save_formatted_excel(self):
        self.workbook.save(self.file_name)


def get_file_path():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select an Excel file to summarise", filetypes=[("Excel files", "*.xlsx")]
    )
    if file_path:
        file_path, file_name = os.path.split(file_path)
        return file_path, file_name
    else:
        return None, None


def main():
    # Input file ProjectList.xlsx CN42N report with N* as Project ID
    file_path, file_name = get_file_path()
    input_file = os.path.join(file_path, file_name)
    fname, extension = os.path.splitext(file_name)
    output_file = os.path.join(file_path, fname + "Summary.xlsx")

    processor = WBSProcessor(input_file, output_file)
    processor.read_excel()
    processor.group_project_definitions()
    processor.map_values()
    processor.create_summary()
    processor.write_to_excel(output_file)

    # Open the existing workbook
    workbook = openpyxl.load_workbook(output_file)
    # Format all sheets
    sheet_names = workbook.sheetnames
    for sheet_name in sheet_names:
        formatter = ExcelFormatter(output_file, workbook, sheet_name)
        formatter.apply_common_style()
        formatter.auto_adjust_columns()
        formatter.add_table()
        formatter.apply_header_style()
        formatter.freeze_panes()
        formatter.save_formatted_excel()

    messagebox.showinfo(
        "Project Type Wise", f"Excel file '{output_file}' has been created successfully"
    )


if __name__ == "__main__":
    main()
