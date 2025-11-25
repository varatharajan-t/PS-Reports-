import pandas as pd
import os
from openpyxl.styles import NamedStyle, Font
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from config import Config


class ExcelProcessor:
    def __init__(self, input_file, output_file, config):
        self.input_file = input_file
        self.output_file = output_file
        self.config = config
        self.df = pd.read_excel(input_file, header=[0, 1])
        self.currency_style = None

    def extract_and_save(self):
        reference_columns = self.df.iloc[:, :4]

        # Extract "Available" and "Assigned" data
        available_data = pd.concat(
            [reference_columns, self.df.loc[:, pd.IndexSlice[:, "Available"]]], axis=1
        )
        assigned_data = pd.concat(
            [reference_columns, self.df.loc[:, pd.IndexSlice[:, "Assigned"]]], axis=1
        )

        with pd.ExcelWriter(self.output_file) as writer:
            available_data.to_excel(writer, sheet_name="Available")
            assigned_data.to_excel(writer, sheet_name="Assigned")
        print(f"Data successfully written to '{self.output_file}'")

    def register_styles(self, wb):
        """Register currency style once at workbook level."""
        if "currency" not in wb.named_styles:
            self.currency_style = NamedStyle(
                name="currency",
                font=Font(**self.config.EXCEL_FONT),
                number_format=self.config.CURRENCY_FORMAT,
            )
            wb.add_named_style(self.currency_style)
        else:
            self.currency_style = wb.named_styles["currency"]

    def apply_currency_formatting(self, ws):
        for col in range(5, ws.max_column + 1):
            for cell in ws[get_column_letter(col)]:
                cell.style = self.currency_style

    def adjust_column_width(self, ws):
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_length = max(
                (len(str(cell.value)) for cell in ws[col_letter] if cell.value),
                default=10,
            )
            ws.column_dimensions[col_letter].width = max_length + 5

    def apply_font_style(self, ws):
        font = Font(**self.config.EXCEL_FONT)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font

    def clean_sheets(self):
        wb = load_workbook(self.output_file)
        self.register_styles(wb)  # Register style once

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(3)  # Delete row 3
            ws.delete_rows(3)  # After deleting row 3, row 4 becomes row 3
            ws.delete_cols(1)  # Delete the index column

            # Apply formatting consistently
            self.apply_currency_formatting(ws)
            self.apply_font_style(ws)
            self.adjust_column_width(ws)

        wb.save(self.output_file)
        print(f"Deleted rows and cleaned all sheets in '{self.output_file}'")

    def process(self):
        self.extract_and_save()
        self.clean_sheets()


if __name__ == "__main__":
    config = Config()
    processor = ExcelProcessor(config.YEAR_END_558_INPUT_FILE, config.YEAR_END_558_OUTPUT_FILE, config)
    processor.process()
