"""
Service layer for generating the Budget Report.
This module contains the specific data processing and Excel formatting logic
for the Budget Report, adapted from the original BudgetReport.py script.
"""
import os
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

from django.conf import settings
from .data_processing import BaseDataProcessor, WBSProcessor, MasterDataManager
from .error_handling import handle_error, ExcelGenerationError
import re

# This service will reuse the core WBSProcessor and MasterDataManager,
# but the data cleaning and transformation is specific to this report.


def format_indian_currency(value):
    """
    Format a number in Indian currency style with crore, lakh separators.
    Example: 12,34,56,789.00
    """
    if pd.isna(value) or value == '':
        return ''

    try:
        # Convert to float if it's not already
        num = float(value)

        # Handle negative numbers
        is_negative = num < 0
        num = abs(num)

        # Format to 2 decimal places
        num_str = f"{num:.2f}"
        integer_part, decimal_part = num_str.split('.')

        # Apply Indian numbering system
        if len(integer_part) <= 3:
            formatted = integer_part
        else:
            # Last 3 digits
            last_three = integer_part[-3:]
            remaining = integer_part[:-3]

            # Group remaining digits in pairs from right to left
            groups = []
            while remaining:
                groups.append(remaining[-2:])
                remaining = remaining[:-2]

            groups.reverse()
            formatted = ','.join(groups) + ',' + last_three

        result = f"₹ {formatted}.{decimal_part}"

        if is_negative:
            result = f"-{result}"

        return result
    except (ValueError, TypeError):
        return str(value)


def generate_formatted_html(df, summary_wbs_list):
    """
    Generate a professionally formatted HTML table with Indian currency formatting.
    """
    # Identify currency columns (skip first few non-currency columns)
    non_currency_cols = ['Sl No.']
    currency_columns = []

    for col in df.columns:
        col_lower = str(col).lower()
        if col not in non_currency_cols and not any(x in col_lower for x in ['level', 'description', 'id', 'wbs', 'object', 'name']):
            currency_columns.append(col)

    # Find the ID column for highlighting summary WBS
    id_column = None
    for col in df.columns:
        if 'id' in str(col).lower() or 'wbs' in str(col).lower():
            id_column = col
            break

    # Build HTML table
    html_parts = []

    # Add CSS styling
    html_parts.append('''
    <style>
        .budget-report-table {
            width: 100%;
            border-collapse: collapse;
            font-family: 'Bookman Old Style', 'Times New Roman', serif;
            font-size: 12px;
            margin: 20px 0;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .budget-report-table th {
            background-color: #FFFF00 !important;
            color: #000000;
            font-weight: bold;
            padding: 12px 8px;
            text-align: left;
            border: 1px solid #000000;
            position: -webkit-sticky;
            position: sticky;
            top: 0;
            z-index: 10;
            box-shadow: 0 2px 2px -1px rgba(0, 0, 0, 0.4);
        }
        .budget-report-table td {
            padding: 8px;
            border: 1px solid #000000;
        }
        .budget-report-table tbody tr:nth-child(even) {
            background-color: #87CEEB;
        }
        .budget-report-table tbody tr:nth-child(odd) {
            background-color: #FFFFFF;
        }
        .budget-report-table tbody tr.summary-wbs {
            background-color: #90EE90 !important;
            font-weight: bold;
        }
        .budget-report-table .currency-cell {
            text-align: right;
            font-family: 'Courier New', monospace;
        }
        .budget-report-table .text-cell {
            text-align: left;
        }
        .budget-report-table .center-cell {
            text-align: center;
        }
        .budget-report-table .negative {
            color: #FF0000;
        }
        .budget-report-table-container {
            overflow-x: auto;
            max-width: 100%;
        }
    </style>
    ''')

    html_parts.append('<div class="budget-report-table-container">')
    html_parts.append('<table class="budget-report-table">')

    # Table header
    html_parts.append('<thead><tr>')
    for col in df.columns:
        html_parts.append(f'<th>{col}</th>')
    html_parts.append('</tr></thead>')

    # Table body
    html_parts.append('<tbody>')

    for idx, row in df.iterrows():
        # Check if this row is a summary WBS
        row_class = ''
        if id_column and summary_wbs_list:
            wbs_value = row.get(id_column, '')
            if wbs_value in summary_wbs_list:
                row_class = ' class="summary-wbs"'

        html_parts.append(f'<tr{row_class}>')

        for col in df.columns:
            value = row[col]

            if col in currency_columns:
                # Format as Indian currency
                formatted_value = format_indian_currency(value)
                cell_class = 'currency-cell'
                if formatted_value.startswith('-'):
                    cell_class += ' negative'
                html_parts.append(f'<td class="{cell_class}">{formatted_value}</td>')
            elif col == 'Sl No.':
                # Center align serial numbers
                html_parts.append(f'<td class="center-cell">{value}</td>')
            else:
                # Text columns - left align
                html_parts.append(f'<td class="text-cell">{value if pd.notna(value) else ""}</td>')

        html_parts.append('</tr>')

    html_parts.append('</tbody>')
    html_parts.append('</table>')
    html_parts.append('</div>')

    return ''.join(html_parts)


class BudgetReportProcessor(BaseDataProcessor):
    """
    Handles all data processing operations for the SAP Budget Report.
    Inherits from the base processor but contains report-specific logic.
    """
    def __init__(self, input_file_path: str):
        super().__init__('BudgetReport')
        self.input_file_path = Path(input_file_path)
        self.wbs_processor = WBSProcessor()

    def validate_input(self, file_path: str) -> bool:
        """
        Validate the input DAT file format and content.

        Args:
            file_path: Path to the DAT file to validate

        Returns:
            bool: True if the file is valid, raises exception otherwise
        """
        file_path_obj = Path(file_path)

        # Check if file exists
        if not file_path_obj.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Check if file is empty
        if file_path_obj.stat().st_size == 0:
            raise ValueError("DAT file is empty")

        # Check if file has .dat extension
        if file_path_obj.suffix.lower() not in ['.dat', '.txt']:
            self.logger.warning(f"File does not have .dat extension: {file_path}")

        # Try to read first few lines to verify it's a valid DAT file
        try:
            with open(file_path, "r", encoding="iso-8859-1") as file:
                lines = file.readlines()
                if len(lines) < 5:
                    raise ValueError("DAT file has too few lines to be valid")
        except Exception as e:
            raise ValueError(f"Error reading DAT file: {str(e)}")

        self.logger.info(f"Input file validation passed: {file_path}")
        return True

    def clean_data(self):
        """
        Cleans the raw DAT file by removing unnecessary header and footer lines.
        Returns the path to the cleaned file.
        """
        cleaned_dat_path = self.input_file_path.parent / f"cleaned_{self.input_file_path.name}"
        
        with open(self.input_file_path, "r", encoding="iso-8859-1") as file:
            lines = file.readlines()

        total_lines = len(lines)
        lines_to_delete = {0, 1, 4, total_lines - 1}

        cleaned_lines = [line for i, line in enumerate(lines) if i not in lines_to_delete]

        with open(cleaned_dat_path, "w", encoding="utf-8") as file:
            file.writelines(cleaned_lines)
        
        self.logger.info(f"Cleaned data written to {cleaned_dat_path}")
        return str(cleaned_dat_path)

    def process_data(self, file_path: str) -> pd.DataFrame:
        """
        Processes the cleaned DAT file into a structured DataFrame.
        """
        df = self.read_dat_file(file_path, header_rows=[0, 1])
        df = df.rename(columns={"Unnamed: 0_level_0": "WBS_Elements_Info."})
        
        df[("WBS_Elements_Info.", "ID_No")] = df[("WBS_Elements_Info.", "Object")].str.split().str[-1]
        df[("WBS_Elements_Info.", "Level")] = df[("WBS_Elements_Info.", "Object")].str.split().str[0]
        df[("WBS_Elements_Info.", "Description")] = ''

        first_cols = [
            ("WBS_Elements_Info.", "Level"),
            ("WBS_Elements_Info.", "Description"),
            ("WBS_Elements_Info.", "ID_No"),
        ]
        other_cols = [col for col in df.columns if col[1] != 'Object' and col not in first_cols]
        df = df[first_cols + other_cols]

        df = df.dropna(how="all")
        self.df = df
        return df, self.wbs_processor.classify_wbs_elements(df[("WBS_Elements_Info.", "ID_No")].tolist())


class BudgetExcelFormatter:
    """Handles the Excel formatting for the Budget Report."""
    def __init__(self, df: pd.DataFrame, output_file: str):
        self.df = df
        self.output_file = output_file
        self.workbook = None
        self.worksheet = None

    @handle_error
    def format_and_save(self, summary_wbs_list: list):
        # Find ID_No column index before flattening
        id_col_idx = -1
        for i, col in enumerate(self.df.columns):
            if isinstance(col, tuple) and col[1] == 'ID_No':
                id_col_idx = i + 1  # Excel columns are 1-indexed
                break

        # Flatten MultiIndex columns to strings for Excel compatibility
        self.df.columns = [' - '.join(col).strip() if isinstance(col, tuple) else str(col) for col in self.df.columns]

        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            self.df.to_excel(writer, index=False, sheet_name='Budget Report')
            self.workbook = writer.book
            self.worksheet = writer.sheets['Budget Report']

            self._apply_styles()
            self._highlight_summary_rows(summary_wbs_list, id_col_idx)
            self._adjust_column_widths()
            self.worksheet.freeze_panes = "E3"

        self.workbook.save(self.output_file)
        return self.output_file

    def _apply_styles(self):
        header_font = Font(name="Bookman Old Style", size=12, bold=True)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        for cell in self.worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        currency_format = settings.CURRENCY_FORMAT
        for col_idx in range(4, self.worksheet.max_column + 1):
            for row_idx in range(2, self.worksheet.max_row + 1):
                self.worksheet.cell(row=row_idx, column=col_idx).number_format = currency_format
    
    def _highlight_summary_rows(self, summary_wbs_list: list, id_col_idx: int):
        if not summary_wbs_list or id_col_idx == -1:
            return

        summary_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        for row in self.worksheet.iter_rows(min_row=2, max_row=self.worksheet.max_row):
            cell = row[id_col_idx - 1]
            if cell.value in summary_wbs_list:
                for cell_in_row in row:
                    cell_in_row.fill = summary_fill

    def _adjust_column_widths(self):
        for col in self.worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.worksheet.column_dimensions[column].width = adjusted_width


@handle_error
def generate_budget_report(uploaded_file_path: str) -> dict:
    """
    Orchestrates the entire budget report generation process.
    Returns a dictionary containing the path to the formatted Excel report
    and an HTML representation of the data.
    """
    processor = BudgetReportProcessor(uploaded_file_path)

    # Validate the input file before processing
    processor.validate_input(uploaded_file_path)

    cleaned_path = processor.clean_data()
    df, (summary_wbs, transaction_wbs) = processor.process_data(cleaned_path)
    
    master_data_manager = MasterDataManager()
    df = master_data_manager.map_wbs_descriptions(
        transaction_df=df,
        wbs_column=("WBS_Elements_Info.", "ID_No"),
        description_column=("WBS_Elements_Info.", "Description")
    )

    output_filename = Path(uploaded_file_path).with_suffix(".xlsx").name
    reports_dir = settings.BASE_DIR / 'data' / 'reports'
    reports_dir.mkdir(parents=True, exist_ok=True)
    output_path = str(reports_dir / output_filename)
    
    formatter = BudgetExcelFormatter(df, output_path)
    formatted_file_path = formatter.format_and_save(summary_wbs)
    
    os.remove(cleaned_path)

    # Generate HTML table from DataFrame for web display with Indian currency formatting
    df_html = generate_formatted_html(df, summary_wbs)

    return {
        "file_path": formatted_file_path,
        "data_html": df_html
    }
