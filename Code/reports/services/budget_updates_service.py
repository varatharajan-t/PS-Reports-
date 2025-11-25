"""
Service layer for generating the Budget Updates Report.
This module contains the specific data processing and Excel formatting logic
for the Budget Updates Report, adapted from the original BudgetUpdates.py script.

KEY DIFFERENCES FROM BUDGET REPORT:
- Modified data cleaning removes lines [0, 3, last] (different pattern than Budget Report)
- Enhanced WBS processing with regex-based child element detection
- Asterisk replacement for cleaner presentation
- Improved string manipulation for WBS element extraction
"""
import os
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import re

from django.conf import settings
from .data_processing import BaseDataProcessor, WBSProcessor, MasterDataManager
from .error_handling import handle_error, ExcelGenerationError


def format_indian_currency(value):
    """
    Format a number in Indian currency style with crore, lakh separators.
    Example: 12,34,56,789.00
    """
    if pd.isna(value) or value == '':
        return ''

    try:
        # Convert to float if it's not already
        num = float(value)

        # Handle negative numbers
        is_negative = num < 0
        num = abs(num)

        # Format to 2 decimal places
        num_str = f"{num:.2f}"
        integer_part, decimal_part = num_str.split('.')

        # Apply Indian numbering system
        if len(integer_part) <= 3:
            formatted = integer_part
        else:
            # Last 3 digits
            last_three = integer_part[-3:]
            remaining = integer_part[:-3]

            # Group remaining digits in pairs from right to left
            groups = []
            while remaining:
                groups.append(remaining[-2:])
                remaining = remaining[:-2]

            groups.reverse()
            formatted = ','.join(groups) + ',' + last_three

        result = f"₹ {formatted}.{decimal_part}"

        if is_negative:
            result = f"-{result}"

        return result
    except (ValueError, TypeError):
        return str(value)


def generate_formatted_html(df, summary_wbs_list):
    """
    Generate a professionally formatted HTML table with Indian currency formatting.
    """
    # Identify currency columns (skip first few non-currency columns)
    non_currency_cols = ['Sl No.']
    currency_columns = []

    for col in df.columns:
        col_lower = str(col).lower()
        if col not in non_currency_cols and not any(x in col_lower for x in ['level', 'description', 'id', 'wbs', 'object', 'name']):
            currency_columns.append(col)

    # Find the ID column for highlighting summary WBS
    id_column = None
    for col in df.columns:
        if 'id' in str(col).lower() or 'wbs' in str(col).lower():
            id_column = col
            break

    # Build HTML table
    html_parts = []

    # Add CSS styling
    html_parts.append('''
    <style>
        .budget-updates-table {
            width: 100%;
            border-collapse: collapse;
            font-family: 'Bookman Old Style', 'Times New Roman', serif;
            font-size: 12px;
            margin: 20px 0;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .budget-updates-table th {
            background-color: #FFFF00 !important;
            color: #000000;
            font-weight: bold;
            padding: 12px 8px;
            text-align: left;
            border: 1px solid #000000;
            position: -webkit-sticky;
            position: sticky;
            top: 0;
            z-index: 10;
            box-shadow: 0 2px 2px -1px rgba(0, 0, 0, 0.4);
        }
        .budget-updates-table td {
            padding: 8px;
            border: 1px solid #000000;
        }
        .budget-updates-table tbody tr:nth-child(even) {
            background-color: #87CEEB;
        }
        .budget-updates-table tbody tr:nth-child(odd) {
            background-color: #FFFFFF;
        }
        .budget-updates-table tbody tr.summary-wbs {
            background-color: #90EE90 !important;
            font-weight: bold;
        }
        .budget-updates-table .currency-cell {
            text-align: right;
            font-family: 'Courier New', monospace;
        }
        .budget-updates-table .text-cell {
            text-align: left;
        }
        .budget-updates-table .center-cell {
            text-align: center;
        }
        .budget-updates-table .negative {
            color: #FF0000;
        }
        .budget-updates-table-container {
            overflow-x: auto;
            max-width: 100%;
        }
    </style>
    ''')

    html_parts.append('<div class="budget-updates-table-container">')
    html_parts.append('<table class="budget-updates-table">')

    # Table header
    html_parts.append('<thead><tr>')
    for col in df.columns:
        html_parts.append(f'<th>{col}</th>')
    html_parts.append('</tr></thead>')

    # Table body
    html_parts.append('<tbody>')

    for idx, row in df.iterrows():
        # Check if this row is a summary WBS
        row_class = ''
        if id_column and summary_wbs_list:
            wbs_value = row.get(id_column, '')
            if wbs_value in summary_wbs_list:
                row_class = ' class="summary-wbs"'

        html_parts.append(f'<tr{row_class}>')

        for col in df.columns:
            value = row[col]

            if col in currency_columns:
                # Format as Indian currency
                formatted_value = format_indian_currency(value)
                cell_class = 'currency-cell'
                if formatted_value.startswith('-'):
                    cell_class += ' negative'
                html_parts.append(f'<td class="{cell_class}">{formatted_value}</td>')
            elif col == 'Sl No.':
                # Center align serial numbers
                html_parts.append(f'<td class="center-cell">{value}</td>')
            else:
                # Text columns - left align
                html_parts.append(f'<td class="text-cell">{value if pd.notna(value) else ""}</td>')

        html_parts.append('</tr>')

    html_parts.append('</tbody>')
    html_parts.append('</table>')
    html_parts.append('</div>')

    return ''.join(html_parts)


class BudgetUpdatesProcessor(BaseDataProcessor):
    """
    Handles all data processing operations for the Budget Updates Report.
    Uses a different cleaning pattern than Budget Report.
    """
    def __init__(self, input_file_path: str):
        super().__init__('BudgetUpdates')
        self.input_file_path = Path(input_file_path)
        self.wbs_processor = WBSProcessor()
        self.summary_wbs_list = []
        self.transaction_wbs_list = []

    def validate_input(self, file_path: str) -> bool:
        """
        Validate the input DAT file format and content.

        Args:
            file_path: Path to the DAT file to validate

        Returns:
            bool: True if the file is valid, raises exception otherwise
        """
        file_path_obj = Path(file_path)

        # Check if file exists
        if not file_path_obj.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Check if file is empty
        if file_path_obj.stat().st_size == 0:
            raise ValueError("DAT file is empty")

        # Check if file has .dat extension
        if file_path_obj.suffix.lower() not in ['.dat', '.txt']:
            self.logger.warning(f"File does not have .dat extension: {file_path}")

        # Try to read first few lines to verify it's a valid DAT file
        try:
            with open(file_path, "r", encoding="iso-8859-1") as file:
                lines = file.readlines()
                if len(lines) < 5:
                    raise ValueError("DAT file has too few lines to be valid")
        except Exception as e:
            raise ValueError(f"Error reading DAT file: {str(e)}")

        self.logger.info(f"Input file validation passed: {file_path}")
        return True

    def clean_data(self):
        """
        Cleans the raw DAT file by removing unnecessary header and footer lines.
        BUDGET UPDATES specific cleaning: removes lines [0, 3, last].
        Also replaces asterisks with spaces for cleaner presentation.
        Returns the path to the cleaned file.
        """
        cleaned_dat_path = self.input_file_path.parent / f"cleaned_{self.input_file_path.name}"

        with open(self.input_file_path, "r", encoding="iso-8859-1") as file:
            lines = file.readlines()

        total_lines = len(lines)
        # Budget Updates specific pattern: remove lines 0, 3, and last
        lines_to_delete = {0, 3, total_lines - 1}

        cleaned_lines = [line for i, line in enumerate(lines) if i not in lines_to_delete]

        # Replace asterisks with spaces for cleaner presentation (Budget Updates specific)
        cleaned_lines = [line.replace("*", " ") for line in cleaned_lines]

        with open(cleaned_dat_path, "w", encoding="utf-8") as file:
            file.writelines(cleaned_lines)

        self.logger.info(f"Cleaned data written to {cleaned_dat_path}")
        return str(cleaned_dat_path)

    def process_data(self, file_path: str) -> pd.DataFrame:
        """
        Processes the cleaned DAT file into a structured DataFrame.
        This implements the Budget Updates specific transformation logic.
        """
        # Read the cleaned DAT file
        df = self.read_dat_file(file_path, header_rows=[0, 1])

        # Rename the first column header
        df = df.rename(columns={"Unnamed: 0_level_0": "WBS_Elements_Info."})

        # Get all columns except the first one for later reordering
        df_columns = list(df.columns[1:])

        # Split "WBS Element Details" column into separate columns
        split_details = df[("WBS_Elements_Info.", "Object")].str.split()

        """Transforms data by splitting the WBS Element Details and renaming columns to a multi-index format."""
        level, description, id_no = [], [], []

        for details in split_details:
            # Handle NaN or empty values
            if details is None or (isinstance(details, float) and pd.isna(details)):
                level.append(" ")
                description.append("")
                id_no.append("")
                continue

            # Handle case where details is not a list (shouldn't happen after split, but be safe)
            if not isinstance(details, list):
                level.append(" ")
                description.append("")
                id_no.append(str(details) if details else "")
                continue

            # Handle empty list
            if len(details) == 0:
                level.append(" ")
                description.append("")
                id_no.append("")
                continue

            Level = details[0] if len(details) > 0 else " "
            Description = "".join(details[2:]) if len(details) > 2 else ""
            ID = details[1] if len(details) > 1 else ""

            # If Level is not a valid level indicator, treat it as part of description
            if Level not in ["*", "**", "***", "4*", "5*"]:
                Description = Level.strip() + Description.strip()
                Level = " "

            level.append(Level)
            description.append(Description)
            id_no.append(ID)

        # Add the new columns to the DataFrame
        df[("WBS_Elements_Info.", "Level")] = level
        df[("WBS_Elements_Info.", "Description")] = description
        df[("WBS_Elements_Info.", "ID_No")] = id_no

        # Re-arrange columns to place new columns at the beginning
        df = df[
            [
                ("WBS_Elements_Info.", "Level"),
                ("WBS_Elements_Info.", "Description"),
                ("WBS_Elements_Info.", "ID_No"),
            ]
            + df_columns
        ]

        # Drop rows that are completely empty
        df = df.dropna(how="all")

        # Process WBS classification before flattening columns
        self.summary_wbs_list, self.transaction_wbs_list = self.process_wbs(
            df[("WBS_Elements_Info.", "ID_No")].to_list()
        )

        self.df = df
        return df, (self.summary_wbs_list, self.transaction_wbs_list)

    def process_wbs(self, wbs_list):
        """
        Classify WBS elements into summary and transaction categories.
        Uses regex pattern with fullmatch for precise child element detection.
        """
        summary_wbs = []
        transaction_wbs = []

        # Remove None and NaN values - properly handle pandas types
        cleaned_wbs_list = []
        for wbs in wbs_list:
            if wbs is None:
                continue
            if isinstance(wbs, float) and pd.isna(wbs):
                continue
            wbs_str = str(wbs).strip()
            if wbs_str and wbs_str.lower() != 'nan':
                cleaned_wbs_list.append(wbs_str)

        wbs_list = cleaned_wbs_list

        # Iterate over the WBS IDs
        for wbs in wbs_list:
            # Define a regular expression pattern to match a WBS element with child elements
            pattern = re.compile(re.escape(wbs) + r"-\d{2}")

            # Check if any WBS starts with the parent WBS followed by a hyphen and two digits
            has_child = any(pattern.fullmatch(item) for item in wbs_list if item != wbs)

            if has_child:
                summary_wbs.append(wbs)  # Add parent WBS to summary WBS
            else:
                transaction_wbs.append(wbs)  # Add WBS to transaction WBS

        return summary_wbs, transaction_wbs


class BudgetUpdatesExcelFormatter:
    """Handles the Excel formatting for the Budget Updates Report."""
    def __init__(self, df: pd.DataFrame, output_file: str):
        self.df = df
        self.output_file = output_file
        self.workbook = None
        self.worksheet = None

    @handle_error
    def format_and_save(self, summary_wbs_list: list):
        """Format and save the Excel file with Budget Updates specific styling."""
        # Find ID_No column index before flattening
        id_col_idx = -1
        for i, col in enumerate(self.df.columns):
            if isinstance(col, tuple) and col[1] == 'ID_No':
                id_col_idx = i + 1  # Excel columns are 1-indexed
                break

        # Flatten MultiIndex columns to strings for Excel compatibility
        self.df.columns = [' - '.join(col).strip() if isinstance(col, tuple) else str(col) for col in self.df.columns]

        # Add serial number column at the beginning
        self.df.insert(0, 'Sl No.', range(1, len(self.df) + 1))

        # Adjust id_col_idx for the new serial number column
        if id_col_idx != -1:
            id_col_idx += 1

        # Write to Excel
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            self.df.to_excel(writer, index=False, sheet_name='Budget Updates')
            self.workbook = writer.book
            self.worksheet = writer.sheets['Budget Updates']

            self._apply_freeze_panes()
            self._apply_font_style()
            self._apply_currency_formatting()
            self._apply_header_format()
            self._highlight_summary_rows(summary_wbs_list, id_col_idx)
            self._adjust_column_widths()

        self.workbook.save(self.output_file)
        return self.output_file

    def _apply_freeze_panes(self):
        """Freeze panes for the Excel sheet."""
        self.worksheet.freeze_panes = "E3"

    def _apply_font_style(self):
        """Apply Bookman Old Style font to all cells."""
        font = Font(name="Bookman Old Style", size=12)
        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.font = font

    def _apply_currency_formatting(self):
        """Apply Indian currency formatting to numeric columns."""
        currency_style = NamedStyle(
            name="currency_updates",
            font=Font(name="Bookman Old Style", size=12),
            number_format=f"₹ #,##0.00;[Red]₹ -#,##0.00",
        )

        # Apply currency format to columns starting from column 5 (E)
        for col in range(5, self.worksheet.max_column + 1):
            col_letter = get_column_letter(col)
            for cell in self.worksheet[col_letter][1:]:  # Skip header
                cell.style = currency_style

    def _apply_header_format(self):
        """Apply specific formatting to header and data rows."""
        last_col = self.worksheet.max_column
        last_row = self.worksheet.max_row

        # Define border style
        black_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        # Define fill colors
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        skyBlue_fill = PatternFill(
            start_color="87CEEB", end_color="87CEEB", fill_type="solid"
        )
        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )

        # Bold font for headers
        bold_font = Font(name="Bookman Old Style", size=12, color="000000", bold=True)

        # Apply formatting to all cells
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=1, max_row=last_row), start=1):
            for cell in row:
                cell.border = black_border

                if row_idx in [1, 2]:  # Header rows
                    cell.fill = yellow_fill
                    cell.font = bold_font
                else:  # Data rows - alternating colors
                    if row_idx % 2 == 0:
                        cell.fill = skyBlue_fill
                    else:
                        cell.fill = white_fill

    def _highlight_summary_rows(self, summary_wbs_list: list, id_col_idx: int):
        """Highlight rows containing summary WBS elements in light green."""
        if not summary_wbs_list or id_col_idx == -1:
            return

        light_green_fill = PatternFill(
            start_color="90EE90", end_color="90EE90", fill_type="solid"
        )

        # Iterate through data rows (skip header)
        for row in self.worksheet.iter_rows(min_row=3, max_row=self.worksheet.max_row):
            cell = row[id_col_idx - 1]  # Get the ID column cell
            if cell.value in summary_wbs_list:
                # Apply light green fill to entire row
                for cell_in_row in row:
                    cell_in_row.fill = light_green_fill

    def _adjust_column_widths(self):
        """Auto-adjust column widths based on content."""
        for col in range(1, self.worksheet.max_column + 1):
            col_letter = get_column_letter(col)
            max_length = max(
                (
                    len(str(cell.value))
                    for cell in self.worksheet[col_letter]
                    if cell.value
                ),
                default=10,
            )
            self.worksheet.column_dimensions[col_letter].width = max_length + 10


@handle_error
def generate_budget_updates_report(uploaded_file_path: str) -> dict:
    """
    Orchestrates the entire budget updates report generation process.
    Returns a dictionary containing the path to the formatted Excel report
    and an HTML representation of the data.
    """
    processor = BudgetUpdatesProcessor(uploaded_file_path)

    # Validate the input file before processing
    processor.validate_input(uploaded_file_path)

    # Clean the data (Budget Updates specific pattern)
    cleaned_path = processor.clean_data()

    # Process the data
    df, (summary_wbs, transaction_wbs) = processor.process_data(cleaned_path)

    # Map WBS descriptions from master data
    master_data_manager = MasterDataManager()
    df = master_data_manager.map_wbs_descriptions(
        transaction_df=df,
        wbs_column=("WBS_Elements_Info.", "ID_No"),
        description_column=("WBS_Elements_Info.", "Description")
    )

    # Generate output filename and path
    output_filename = Path(uploaded_file_path).with_suffix(".xlsx").name
    reports_dir = settings.BASE_DIR / 'data' / 'reports'
    reports_dir.mkdir(parents=True, exist_ok=True)
    output_path = str(reports_dir / output_filename)

    # Format and save Excel file
    formatter = BudgetUpdatesExcelFormatter(df.copy(), output_path)
    formatted_file_path = formatter.format_and_save(summary_wbs)

    # Clean up temporary file
    os.remove(cleaned_path)

    # Generate HTML table from DataFrame for web display with Indian currency formatting
    # Flatten columns for HTML display
    df_html = df.copy()
    df_html.columns = [' - '.join(col).strip() if isinstance(col, tuple) else str(col) for col in df_html.columns]
    df_html.insert(0, 'Sl No.', range(1, len(df_html) + 1))

    html_output = generate_formatted_html(df_html, summary_wbs)

    return {
        "file_path": formatted_file_path,
        "data_html": html_output
    }
