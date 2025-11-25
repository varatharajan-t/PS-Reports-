"""
Service layer for generating the Budget Variance Report.
This module processes HTML files (not DAT) containing budget variance data
and converts them to professionally formatted Excel files with HTML preview.

KEY DIFFERENCES FROM OTHER REPORTS:
- Processes HTML input files (uses pd.read_html instead of pd.read_csv)
- Removes first TWO lines from HTML file
- Uses ORANGE highlighting for summary WBS (not light green)
- Removes trailing '1' characters from column headers
- Freeze panes at D3 (not E3)
"""
import os
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import re

from django.conf import settings
from .data_processing import BaseDataProcessor, WBSProcessor, MasterDataManager
from .error_handling import handle_error, ExcelGenerationError


def format_indian_currency(value):
    """Format a number in Indian currency style with crore, lakh separators."""
    if pd.isna(value) or value == '':
        return ''

    try:
        num = float(value)
        is_negative = num < 0
        num = abs(num)

        num_str = f"{num:.2f}"
        integer_part, decimal_part = num_str.split('.')

        if len(integer_part) <= 3:
            formatted = integer_part
        else:
            last_three = integer_part[-3:]
            remaining = integer_part[:-3]

            groups = []
            while remaining:
                groups.append(remaining[-2:])
                remaining = remaining[:-2]

            groups.reverse()
            formatted = ','.join(groups) + ',' + last_three

        result = f"₹ {formatted}.{decimal_part}"
        if is_negative:
            result = f"-{result}"

        return result
    except (ValueError, TypeError):
        return str(value)


def generate_formatted_html(df, summary_wbs_list):
    """Generate a professionally formatted HTML table with Indian currency formatting."""
    # Identify currency columns
    non_currency_cols = ['SI_NO']
    currency_columns = []

    for col in df.columns:
        col_lower = str(col).lower()
        if col not in non_currency_cols and not any(x in col_lower for x in ['wbs', 'description', 'element']):
            currency_columns.append(col)

    # Find the WBS column for highlighting
    wbs_column = None
    for col in df.columns:
        if 'wbs' in str(col).lower() or 'element' in str(col).lower():
            wbs_column = col
            break

    # Build HTML table
    html_parts = []

    # Add CSS styling - ORANGE highlighting for Budget Variance
    html_parts.append('''
    <style>
        .budget-variance-table {
            width: 100%;
            border-collapse: collapse;
            font-family: 'Bookman Old Style', 'Times New Roman', serif;
            font-size: 12px;
            margin: 20px 0;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .budget-variance-table th {
            background-color: #FFFF00 !important;
            color: #000000;
            font-weight: bold;
            padding: 12px 8px;
            text-align: left;
            border: 1px solid #000000;
            position: -webkit-sticky;
            position: sticky;
            top: 0;
            z-index: 10;
            box-shadow: 0 2px 2px -1px rgba(0, 0, 0, 0.4);
        }
        .budget-variance-table td {
            padding: 8px;
            border: 1px solid #000000;
        }
        .budget-variance-table tbody tr:nth-child(even) {
            background-color: #F0F0F0;
        }
        .budget-variance-table tbody tr:nth-child(odd) {
            background-color: #FFFFFF;
        }
        .budget-variance-table tbody tr.summary-wbs {
            background-color: #FFA500 !important;
            font-weight: bold;
        }
        .budget-variance-table .currency-cell {
            text-align: right;
            font-family: 'Courier New', monospace;
        }
        .budget-variance-table .text-cell {
            text-align: left;
        }
        .budget-variance-table .center-cell {
            text-align: center;
        }
        .budget-variance-table .negative {
            color: #FF0000;
        }
        .budget-variance-table-container {
            overflow-x: auto;
            max-width: 100%;
        }
    </style>
    ''')

    html_parts.append('<div class="budget-variance-table-container">')
    html_parts.append('<table class="budget-variance-table">')

    # Table header
    html_parts.append('<thead><tr>')
    for col in df.columns:
        html_parts.append(f'<th>{col}</th>')
    html_parts.append('</tr></thead>')

    # Table body
    html_parts.append('<tbody>')

    for idx, row in df.iterrows():
        # Check if this row is a summary WBS
        row_class = ''
        if wbs_column and summary_wbs_list:
            wbs_value = row.get(wbs_column, '')
            if wbs_value in summary_wbs_list:
                row_class = ' class="summary-wbs"'

        html_parts.append(f'<tr{row_class}>')

        for col in df.columns:
            value = row[col]

            if col in currency_columns:
                formatted_value = format_indian_currency(value)
                cell_class = 'currency-cell'
                if formatted_value.startswith('-'):
                    cell_class += ' negative'
                html_parts.append(f'<td class="{cell_class}">{formatted_value}</td>')
            elif col == 'SI_NO':
                html_parts.append(f'<td class="center-cell">{value}</td>')
            else:
                html_parts.append(f'<td class="text-cell">{value if pd.notna(value) else ""}</td>')

        html_parts.append('</tr>')

    html_parts.append('</tbody>')
    html_parts.append('</table>')
    html_parts.append('</div>')

    return ''.join(html_parts)


class BudgetVarianceProcessor(BaseDataProcessor):
    """Handles all data processing operations for the Budget Variance Report."""

    def __init__(self, input_file_path: str):
        super().__init__('BudgetVariance')
        self.input_file_path = Path(input_file_path)
        self.wbs_processor = WBSProcessor()
        self.summary_wbs_list = []
        self.transaction_wbs_list = []

    def validate_input(self, file_path: str) -> bool:
        """Validate the input HTML file format and content."""
        file_path_obj = Path(file_path)

        if not file_path_obj.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if file_path_obj.stat().st_size == 0:
            raise ValueError("HTML file is empty")

        if file_path_obj.suffix.lower() not in ['.html', '.htm']:
            self.logger.warning(f"File does not have .html extension: {file_path}")

        try:
            with open(file_path, "r", encoding="utf-8") as file:
                lines = file.readlines()
                if len(lines) < 3:
                    raise ValueError("HTML file has too few lines to be valid")
        except Exception as e:
            raise ValueError(f"Error reading HTML file: {str(e)}")

        self.logger.info(f"Input file validation passed: {file_path}")
        return True

    def clean_data(self):
        """Removes the first TWO lines from the HTML file (Budget Variance specific)."""
        cleaned_html_path = self.input_file_path.parent / f"cleaned_{self.input_file_path.name}"

        with open(self.input_file_path, "r", encoding="utf-8") as file:
            lines = file.readlines()

        if len(lines) > 2:
            with open(cleaned_html_path, "w", encoding="utf-8") as file:
                file.writelines(lines[2:])  # Remove first two lines
            self.logger.info(f"Cleaned HTML saved as {cleaned_html_path}")
        else:
            self.logger.warning("File doesn't have enough lines to clean.")
            # Write original content if not enough lines
            with open(cleaned_html_path, "w", encoding="utf-8") as file:
                file.writelines(lines)

        return str(cleaned_html_path)

    def read_html_data(self, cleaned_html):
        """Reads the cleaned HTML into a Pandas DataFrame."""
        # pd.read_html returns a list of DataFrames, take the first one
        df = pd.read_html(cleaned_html, header=[0, 1], encoding="utf-8")[0]

        # Flatten multi-level columns
        df.columns = df.columns.map(lambda x: " ".join(map(str, x)).strip())

        # Drop the last row if it exists (usually totals)
        if len(df) > 0:
            df = df.drop(df.index[-1])

        return df

    def process_data(self, file_path: str) -> pd.DataFrame:
        """Processes the cleaned HTML file into a structured DataFrame."""
        df = self.read_html_data(file_path)

        # Rename first two columns
        if len(df.columns) >= 2:
            column_mapping = {
                df.columns[0]: "WBS_element",
                df.columns[1]: "Description",
            }
            df = df.rename(columns=column_mapping)

            # Clean WBS elements
            df["WBS_element"] = (
                df["WBS_element"].astype(str).str.strip().fillna("")
            )

        # Drop completely empty rows
        df = df.dropna(how="all")

        # Process WBS classification
        self.summary_wbs_list, self.transaction_wbs_list = self.process_wbs(
            df["WBS_element"].tolist()
        )

        # Add serial number column
        df.insert(0, "SI_NO", range(1, len(df) + 1))

        self.df = df
        return df, (self.summary_wbs_list, self.transaction_wbs_list)

    def process_wbs(self, wbs_list):
        """Classifies WBS elements into summary and transaction WBS."""
        summary_wbs = []
        transaction_wbs_set = set(wbs for wbs in wbs_list if wbs and str(wbs).strip())

        for wbs in wbs_list:
            if not wbs or not str(wbs).strip():
                continue

            wbs_str = str(wbs).strip()

            # Check for any child WBS in the range -01 to -99
            is_parent = False
            for i in range(1, 100):
                child_wbs = f"{wbs_str}-{i:02d}"
                if child_wbs in transaction_wbs_set:
                    is_parent = True
                    break

            if is_parent:
                summary_wbs.append(wbs_str)

        return summary_wbs, list(transaction_wbs_set)


class BudgetVarianceExcelFormatter:
    """Handles the Excel formatting for the Budget Variance Report."""

    def __init__(self, df: pd.DataFrame, output_file: str):
        self.df = df
        self.output_file = output_file
        self.workbook = None
        self.worksheet = None

    @handle_error
    def format_and_save(self, summary_wbs_list: list):
        """Format and save the Excel file with Budget Variance specific styling."""
        # Find WBS column index
        wbs_col_idx = -1
        for i, col in enumerate(self.df.columns):
            if 'wbs' in str(col).lower() or 'element' in str(col).lower():
                wbs_col_idx = i + 1  # Excel columns are 1-indexed
                break

        # Remove trailing '1' from column headers (Budget Variance specific)
        cleaned_columns = []
        for col in self.df.columns:
            col_str = str(col)
            if col_str.endswith('1'):
                col_str = col_str.rstrip('1').strip()
            cleaned_columns.append(col_str)
        self.df.columns = cleaned_columns

        # Write to Excel
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            self.df.to_excel(writer, index=False, sheet_name='Budget Variance')
            self.workbook = writer.book
            self.worksheet = writer.sheets['Budget Variance']

            self._apply_freeze_panes()
            self._apply_font_style()
            self._apply_currency_formatting()
            self._apply_header_format()
            self._highlight_summary_rows(summary_wbs_list, wbs_col_idx)
            self._adjust_column_widths()
            self._create_table()

        self.workbook.save(self.output_file)
        return self.output_file

    def _apply_freeze_panes(self):
        """Freeze panes at D3 (Budget Variance specific)."""
        self.worksheet.freeze_panes = "D3"

    def _apply_font_style(self):
        """Apply Bookman Old Style font to all cells."""
        font = Font(name="Bookman Old Style", size=12)
        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.font = font

    def _apply_currency_formatting(self):
        """Apply Indian currency formatting to numeric columns (starting from column 3)."""
        currency_style = NamedStyle(
            name="currency_variance",
            font=Font(name="Bookman Old Style", size=12),
            number_format="₹ #,##0.00;[Red]₹ -#,##0.00",
        )

        for col in range(3, self.worksheet.max_column + 1):
            col_letter = get_column_letter(col)
            for cell in self.worksheet[col_letter][1:]:  # Skip header
                if isinstance(cell.value, (int, float)):
                    cell.style = currency_style

    def _apply_header_format(self):
        """Apply yellow header format with borders."""
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        black_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        bold_font = Font(name="Bookman Old Style", size=12, bold=True)

        # Apply to header row
        for col in range(1, self.worksheet.max_column + 1):
            cell = self.worksheet.cell(row=1, column=col)
            cell.fill = yellow_fill
            cell.border = black_border
            cell.font = bold_font

    def _highlight_summary_rows(self, summary_wbs_list: list, wbs_col_idx: int):
        """Highlight rows containing summary WBS elements in ORANGE (Budget Variance specific)."""
        if not summary_wbs_list or wbs_col_idx == -1:
            return

        # ORANGE highlighting for Budget Variance
        orange_fill = PatternFill(
            start_color="FFA500", end_color="FFA500", fill_type="solid"
        )

        for row in self.worksheet.iter_rows(min_row=2, max_row=self.worksheet.max_row):
            cell = row[wbs_col_idx - 1]
            if cell.value in summary_wbs_list:
                for cell_in_row in row:
                    cell_in_row.fill = orange_fill

    def _adjust_column_widths(self):
        """Auto-adjust column widths based on content."""
        for col in range(1, self.worksheet.max_column + 1):
            col_letter = get_column_letter(col)
            max_length = max(
                (
                    len(str(cell.value))
                    for cell in self.worksheet[col_letter]
                    if cell.value
                ),
                default=10,
            )
            self.worksheet.column_dimensions[col_letter].width = max_length + 10

    def _create_table(self):
        """Create a formatted table."""
        data_range = f"A1:{self.worksheet.cell(row=self.worksheet.max_row, column=self.worksheet.max_column).coordinate}"
        from openpyxl.worksheet.table import Table, TableStyleInfo

        table = Table(displayName="Table1", ref=data_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        self.worksheet.add_table(table)


@handle_error
def generate_budget_variance_report(uploaded_file_path: str) -> dict:
    """
    Orchestrates the entire budget variance report generation process.
    Returns a dictionary containing the path to the formatted Excel report
    and an HTML representation of the data.
    """
    processor = BudgetVarianceProcessor(uploaded_file_path)

    # Validate the input file before processing
    processor.validate_input(uploaded_file_path)

    # Clean the data (HTML specific - removes first 2 lines)
    cleaned_path = processor.clean_data()

    # Process the data
    df, (summary_wbs, transaction_wbs) = processor.process_data(cleaned_path)

    # Map WBS descriptions from master data
    master_data_manager = MasterDataManager()
    df = master_data_manager.map_wbs_descriptions(
        transaction_df=df,
        wbs_column="WBS_element",
        description_column="Description"
    )

    # Generate output filename and path
    output_filename = Path(uploaded_file_path).with_suffix(".xlsx").name
    reports_dir = settings.BASE_DIR / 'data' / 'reports'
    reports_dir.mkdir(parents=True, exist_ok=True)
    output_path = str(reports_dir / output_filename)

    # Format and save Excel file
    formatter = BudgetVarianceExcelFormatter(df.copy(), output_path)
    formatted_file_path = formatter.format_and_save(summary_wbs)

    # Clean up temporary file
    os.remove(cleaned_path)

    # Generate HTML table for web display
    html_output = generate_formatted_html(df, summary_wbs)

    return {
        "file_path": formatted_file_path,
        "data_html": html_output
    }
