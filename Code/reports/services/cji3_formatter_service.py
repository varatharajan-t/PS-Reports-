"""
Service layer for generating the CJI3 Formatted Report.
Adapted from cji3_formatter.py for the Django web interface.

Processes SAP CJI3 Excel exports and generates a formatted workbook
with a Summary sheet and per-plant detail sheets.

Input  : .xlsx exported from SAP CJI3 transaction (must contain 'Sheet1')
Output : Formatted .xlsx + HTML preview of the Summary sheet
"""
import os
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from django.conf import settings
from reports.models import DocumentType, ProjectName

# ── Style constants ───────────────────────────────────────────────────────────
POSITIVE_TYPES = {'WE', 'ZA'}


def _make_fill(hex_rgb: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_rgb)


def _make_font(bold=False, color='000000', name='Arial', size=10) -> Font:
    return Font(bold=bold, color=color, name=name, size=size)


_TITLE_FILL  = _make_fill('BDD7EE')
_HEADER_FILL = _make_fill('4472C4')
_SUBTOT_FILL = _make_fill('D9E2F3')
_GTOTAL_FILL = _make_fill('4472C4')

_TITLE_FONT  = _make_font(bold=True, color='1F4E79')
_HEADER_FONT = _make_font(bold=True, color='FFFFFF')
_SUBTOT_FONT = _make_font(bold=True)
_GTOTAL_FONT = _make_font(bold=True, color='FFFFFF')
_TICK_FONT   = _make_font(bold=True, color='008000')
_CROSS_FONT  = _make_font(bold=True, color='FF0000')
_DATA_FONT   = _make_font()


# ── Internal helpers ──────────────────────────────────────────────────────────

def _status_for_doctype(doc_type: str) -> str:
    return '✔' if doc_type in POSITIVE_TYPES else '✘'


def _write_summary_sheet(wb, title: str, summary_data: dict, project_name_map: dict):
    ws = wb.create_sheet('Summary', 0)

    ws.row_dimensions[1].height = 30
    ws['A1'] = title
    ws['A1'].font = _TITLE_FONT
    ws['A1'].fill = _TITLE_FILL
    ws['A1'].alignment = Alignment(vertical='center')
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}1'].fill = _TITLE_FILL

    ws.row_dimensions[2].height = 13
    headers = ['WBS Group', 'Project Name', 'Status', 'Sum of Amount', 'Count']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 3
    grand_amount, grand_count = 0.0, 0

    for wbs_group in sorted(summary_data.keys()):
        statuses = summary_data[wbs_group]
        sub_amount, sub_count = 0.0, 0
        proj_name = project_name_map.get(wbs_group, '')

        for status in ['✔', '✘']:
            if status not in statuses:
                continue
            amt = round(statuses[status]['amount'], 2)
            cnt = statuses[status]['count']

            ws.cell(row=row, column=1, value=wbs_group).font = _DATA_FONT
            ws.cell(row=row, column=2, value=proj_name).font = _DATA_FONT
            c3 = ws.cell(row=row, column=3, value=status)
            c3.font = _TICK_FONT if status == '✔' else _CROSS_FONT
            c3.alignment = Alignment(horizontal='center')
            c4 = ws.cell(row=row, column=4, value=amt)
            c4.font = _DATA_FONT
            c4.number_format = '#,##0.00'
            c5 = ws.cell(row=row, column=5, value=cnt)
            c5.font = _DATA_FONT
            c5.number_format = '#,##0'

            ws.row_dimensions[row].height = 15.5
            sub_amount += amt
            sub_count  += cnt
            row += 1

        sub_label = ws.cell(row=row, column=1, value=f'{wbs_group} Total')
        sub_label.font = _SUBTOT_FONT
        sub_label.fill = _SUBTOT_FILL
        for col in range(2, 6):
            ws.cell(row=row, column=col).fill = _SUBTOT_FILL
        c4 = ws.cell(row=row, column=4, value=round(sub_amount, 2))
        c4.font = _SUBTOT_FONT
        c4.fill = _SUBTOT_FILL
        c4.number_format = '#,##0.00'
        c5 = ws.cell(row=row, column=5, value=sub_count)
        c5.font = _SUBTOT_FONT
        c5.fill = _SUBTOT_FILL
        c5.number_format = '#,##0'
        ws.row_dimensions[row].height = 16
        grand_amount += sub_amount
        grand_count  += sub_count
        row += 1

    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.fill = _GTOTAL_FILL
        cell.font = _GTOTAL_FONT
    ws.cell(row=row, column=1, value='Grand Total')
    c4 = ws.cell(row=row, column=4, value=round(grand_amount, 2))
    c4.number_format = '#,##0.00'
    c4.fill = _GTOTAL_FILL
    c4.font = _GTOTAL_FONT
    c5 = ws.cell(row=row, column=5, value=grand_count)
    c5.number_format = '#,##0'
    c5.fill = _GTOTAL_FILL
    c5.font = _GTOTAL_FONT
    ws.row_dimensions[row].height = 16

    ws.column_dimensions['A'].width = 32.7
    ws.column_dimensions['B'].width = 45.0
    ws.column_dimensions['C'].width = 10.0
    ws.column_dimensions['D'].width = 27.3
    ws.column_dimensions['E'].width = 12.7


def _write_detail_sheet(wb, sheet_name: str, title: str, plant_pivot: dict, doc_type_map: dict, project_name_map: dict):
    ws = wb.create_sheet(sheet_name)

    ws.row_dimensions[1].height = 30
    ws['A1'] = title
    ws['A1'].font = _TITLE_FONT
    ws['A1'].fill = _TITLE_FILL
    ws['A1'].alignment = Alignment(vertical='center')
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}1'].fill = _TITLE_FILL

    ws.row_dimensions[2].height = 13
    headers = ['WBS Group (First 12)', 'Project Name', 'Document Type', 'Description', 'Sum of Amount', 'Count', 'Status']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row = 3
    grand_amount, grand_count = 0.0, 0

    for wbs_group in sorted(plant_pivot.keys()):
        doc_types = plant_pivot[wbs_group]
        sub_amount, sub_count = 0.0, 0
        proj_name = project_name_map.get(wbs_group, '')

        pos_types = sorted(d for d in doc_types if d in POSITIVE_TYPES)
        neg_types = sorted(d for d in doc_types if d not in POSITIVE_TYPES)

        for doc_type in pos_types + neg_types:
            amt    = round(doc_types[doc_type]['amount'], 2)
            cnt    = doc_types[doc_type]['count']
            status = _status_for_doctype(doc_type)
            desc   = doc_type_map.get(doc_type, '')

            ws.cell(row=row, column=1, value=wbs_group).font = _DATA_FONT
            ws.cell(row=row, column=2, value=proj_name).font = _DATA_FONT
            ws.cell(row=row, column=3, value=doc_type).font  = _DATA_FONT
            ws.cell(row=row, column=4, value=desc).font      = _DATA_FONT
            c5 = ws.cell(row=row, column=5, value=amt)
            c5.font = _DATA_FONT
            c5.number_format = '#,##0.00'
            c6 = ws.cell(row=row, column=6, value=cnt)
            c6.font = _DATA_FONT
            c6.number_format = '#,##0'
            c7 = ws.cell(row=row, column=7, value=status)
            c7.font = _TICK_FONT if status == '✔' else _CROSS_FONT
            c7.alignment = Alignment(horizontal='center')

            ws.row_dimensions[row].height = 15.5
            sub_amount += amt
            sub_count  += cnt
            row += 1

        sub_label = ws.cell(row=row, column=1, value=f'{wbs_group} Total')
        sub_label.font = _SUBTOT_FONT
        sub_label.fill = _SUBTOT_FILL
        for col in range(2, 8):
            ws.cell(row=row, column=col).fill = _SUBTOT_FILL
        c5 = ws.cell(row=row, column=5, value=round(sub_amount, 2))
        c5.font = _SUBTOT_FONT
        c5.fill = _SUBTOT_FILL
        c5.number_format = '#,##0.00'
        c6 = ws.cell(row=row, column=6, value=sub_count)
        c6.font = _SUBTOT_FONT
        c6.fill = _SUBTOT_FILL
        c6.number_format = '#,##0'
        ws.row_dimensions[row].height = 16
        grand_amount += sub_amount
        grand_count  += sub_count
        row += 1

    gt = ws.cell(row=row, column=1, value='Grand Total')
    gt.font = _GTOTAL_FONT
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.fill = _GTOTAL_FILL
        cell.font = _GTOTAL_FONT
    c5 = ws.cell(row=row, column=5, value=round(grand_amount, 2))
    c5.number_format = '#,##0.00'
    c5.fill = _GTOTAL_FILL
    c5.font = _GTOTAL_FONT
    c6 = ws.cell(row=row, column=6, value=grand_count)
    c6.number_format = '#,##0'
    c6.fill = _GTOTAL_FILL
    c6.font = _GTOTAL_FONT
    ws.row_dimensions[row].height = 16

    ws.column_dimensions['A'].width = 32.7
    ws.column_dimensions['B'].width = 45.0
    ws.column_dimensions['C'].width = 14.0
    ws.column_dimensions['D'].width = 27.3
    ws.column_dimensions['E'].width = 27.3
    ws.column_dimensions['F'].width = 12.7
    ws.column_dimensions['G'].width = 10.0


def _read_sheet1(file_path: str) -> dict:
    """
    Read and aggregate Sheet1 data.
    Returns plant_data, summary_data, row_count, fiscal_year.
    """
    wb = load_workbook(file_path, data_only=True)

    if 'Sheet1' not in wb.sheetnames:
        raise ValueError(
            "The uploaded file does not contain a sheet named 'Sheet1'. "
            "Please export from SAP CJI3 without renaming the default sheet."
        )

    ws1 = wb['Sheet1']

    plant_data   = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'amount': 0.0, 'count': 0})))
    summary_data = defaultdict(lambda: defaultdict(lambda: {'amount': 0.0, 'count': 0}))
    fy_counter   = defaultdict(int)
    row_count    = 0

    for row in ws1.iter_rows(min_row=2, values_only=True):
        if len(row) < 15:
            continue
        wbs      = row[1]   # col B
        doc_type = row[4]   # col E
        amount   = row[8]   # col I
        fy       = row[14]  # col O

        if not wbs or not doc_type:
            continue
        if not isinstance(amount, (int, float)):
            continue

        wbs      = str(wbs).strip()
        doc_type = str(doc_type).strip()

        if len(wbs) < 8:
            continue

        plant     = wbs[5:8]
        wbs_group = wbs[:12]
        status    = _status_for_doctype(doc_type)

        plant_data[plant][wbs_group][doc_type]['amount'] += amount
        plant_data[plant][wbs_group][doc_type]['count']  += 1

        summary_data[wbs_group][status]['amount'] += amount
        summary_data[wbs_group][status]['count']  += 1

        if fy:
            fy_counter[fy] += 1

        row_count += 1

    if row_count == 0:
        raise ValueError(
            "No valid data rows found in Sheet1. "
            "Ensure the file is a CJI3 export with WBS elements in column B."
        )

    fiscal_year = int(max(fy_counter, key=fy_counter.get)) if fy_counter else 2025

    return {
        'plant_data':   plant_data,
        'summary_data': summary_data,
        'row_count':    row_count,
        'fiscal_year':  fiscal_year,
    }


def _generate_preview_html(summary_data: dict, fiscal_year: int, project_name_map: dict) -> str:
    """Build an HTML table showing the Summary sheet contents for web preview."""

    rows_html = []
    grand_amount = 0.0
    grand_count  = 0

    for wbs_group in sorted(summary_data.keys()):
        statuses   = summary_data[wbs_group]
        sub_amount = 0.0
        sub_count  = 0
        proj_name  = project_name_map.get(wbs_group, '')

        for status in ['✔', '✘']:
            if status not in statuses:
                continue
            amt = round(statuses[status]['amount'], 2)
            cnt = statuses[status]['count']
            color = '#008000' if status == '✔' else '#CC0000'
            rows_html.append(
                f'<tr>'
                f'<td style="padding:6px 10px;border:1px solid #ddd;">{wbs_group}</td>'
                f'<td style="padding:6px 10px;border:1px solid #ddd;">{proj_name}</td>'
                f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:center;'
                f'color:{color};font-weight:bold;">{status}</td>'
                f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:right;">{amt:,.2f}</td>'
                f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:right;">{cnt:,}</td>'
                f'</tr>'
            )
            sub_amount += amt
            sub_count  += cnt

        rows_html.append(
            f'<tr style="background:#D9E2F3;font-weight:bold;">'
            f'<td style="padding:6px 10px;border:1px solid #ddd;">{wbs_group} Total</td>'
            f'<td style="padding:6px 10px;border:1px solid #ddd;"></td>'
            f'<td style="padding:6px 10px;border:1px solid #ddd;"></td>'
            f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:right;">{round(sub_amount,2):,.2f}</td>'
            f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:right;">{sub_count:,}</td>'
            f'</tr>'
        )
        grand_amount += sub_amount
        grand_count  += sub_count

    rows_html.append(
        f'<tr style="background:#4472C4;color:white;font-weight:bold;">'
        f'<td style="padding:8px 10px;border:1px solid #4472C4;">Grand Total</td>'
        f'<td style="padding:8px 10px;border:1px solid #4472C4;"></td>'
        f'<td style="padding:8px 10px;border:1px solid #4472C4;"></td>'
        f'<td style="padding:8px 10px;border:1px solid #4472C4;text-align:right;">{round(grand_amount,2):,.2f}</td>'
        f'<td style="padding:8px 10px;border:1px solid #4472C4;text-align:right;">{grand_count:,}</td>'
        f'</tr>'
    )

    title = f"Analysis of Project Actual for the Year {fiscal_year}\u2013{fiscal_year + 1}"

    return (
        f'<div style="font-family:Arial,sans-serif;">'
        f'<p style="color:#1F4E79;font-weight:bold;margin-bottom:10px;">{title}</p>'
        f'<table style="border-collapse:collapse;width:100%;font-size:13px;">'
        f'<thead>'
        f'<tr style="background:#4472C4;color:white;">'
        f'<th style="padding:8px 10px;border:1px solid #4472C4;text-align:left;">WBS Group</th>'
        f'<th style="padding:8px 10px;border:1px solid #4472C4;text-align:left;">Project Name</th>'
        f'<th style="padding:8px 10px;border:1px solid #4472C4;text-align:center;">Status</th>'
        f'<th style="padding:8px 10px;border:1px solid #4472C4;text-align:right;">Sum of Amount</th>'
        f'<th style="padding:8px 10px;border:1px solid #4472C4;text-align:right;">Count</th>'
        f'</tr>'
        f'</thead>'
        f'<tbody>{"".join(rows_html)}</tbody>'
        f'</table>'
        f'<p style="color:#666;font-size:11px;margin-top:8px;">'
        f'&#10004; = Positive document types (WE, ZA) &nbsp;|&nbsp; &#10008; = Other document types'
        f'</p>'
        f'</div>'
    )


# ── Public API ────────────────────────────────────────────────────────────────

def generate_cji3_report(file_path: str) -> dict:
    """
    Process a CJI3 Excel file and produce a formatted output workbook.

    Args:
        file_path: Absolute path to the uploaded .xlsx file.

    Returns:
        {
            'file_path':   str  — absolute path of the saved output .xlsx,
            'data_html':   str  — HTML preview of the Summary sheet,
            'row_count':   int,
            'fiscal_year': int,
            'plants':      list[str],
        }
    """
    output_dir = settings.BASE_DIR / 'data' / 'reports'
    os.makedirs(str(output_dir), exist_ok=True)

    timestamp       = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f"CJI3_Formatted_{timestamp}.xlsx"
    output_path     = str(output_dir / output_filename)

    # 1 — Read & aggregate data
    result       = _read_sheet1(file_path)
    plant_data   = result['plant_data']
    summary_data = result['summary_data']
    fiscal_year  = result['fiscal_year']
    row_count    = result['row_count']

    title = f"Analysis of Project Actual for the Year {fiscal_year} -{fiscal_year + 1}."

    # 2 — Load document type descriptions and project names from DB
    doc_type_map     = dict(DocumentType.objects.values_list('code', 'description'))
    project_name_map = dict(ProjectName.objects.values_list('project_definition', 'name'))

    # 3 — Build formatted workbook
    out_wb = openpyxl.Workbook()
    for sheet_name in list(out_wb.sheetnames):
        del out_wb[sheet_name]

    _write_summary_sheet(out_wb, title, summary_data, project_name_map)

    sorted_plants = sorted(plant_data.keys())
    for plant in sorted_plants:
        _write_detail_sheet(out_wb, plant, title, plant_data[plant], doc_type_map, project_name_map)

    out_wb.save(output_path)

    # 4 — Generate HTML preview
    data_html = _generate_preview_html(summary_data, fiscal_year, project_name_map)

    return {
        'file_path':   output_path,
        'data_html':   data_html,
        'row_count':   row_count,
        'fiscal_year': fiscal_year,
        'plants':      sorted_plants,
    }
