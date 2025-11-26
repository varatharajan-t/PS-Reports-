"""
Excel Formatting Framework for Django SAP Reports
Provides consistent, professional Excel formatting across all report services
"""

import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart3D, Reference
from typing import List, Dict, Optional, Union
from abc import ABC, abstractmethod
import logging
import pandas as pd

from django.conf import settings
from .error_handling import handle_error, ExcelGenerationError


def format_indian_currency(value):
    """
    Format a number in Indian currency style with crore, lakh separators.
    Example: 12,34,56,789.00

    Args:
        value: Number to format (int, float, or string)

    Returns:
        str: Formatted currency string with ₹ symbol
    """
    if pd.isna(value) or value == '':
        return ''

    try:
        # Convert to float if it's not already
        num = float(value)

        # Handle negative numbers
        is_negative = num < 0
        num = abs(num)

        # Format to 2 decimal places
        num_str = f"{num:.2f}"
        integer_part, decimal_part = num_str.split('.')

        # Apply Indian numbering system
        if len(integer_part) <= 3:
            formatted = integer_part
        else:
            # Last 3 digits
            last_three = integer_part[-3:]
            remaining = integer_part[:-3]

            # Group remaining digits in pairs from right to left
            groups = []
            while remaining:
                groups.append(remaining[-2:])
                remaining = remaining[:-2]

            groups.reverse()
            formatted = ','.join(groups) + ',' + last_three

        result = f"₹ {formatted}.{decimal_part}"

        if is_negative:
            result = f"-{result}"

        return result
    except (ValueError, TypeError):
        return str(value)


class BaseExcelFormatter(ABC):
    """Abstract base class for Excel formatting operations."""

    def __init__(self, output_file: str, module_name: str = "ExcelFormatter"):
        self.output_file = output_file
        self.logger = logging.getLogger(module_name)

        self._workbook: Optional[openpyxl.Workbook] = None
        self._worksheet: Optional[openpyxl.Worksheet] = None
        self._styles_registered = False

    @property
    def workbook(self) -> openpyxl.Workbook:
        """Lazy loading of workbook."""
        if self._workbook is None:
            self._workbook = openpyxl.load_workbook(self.output_file)
            self._register_styles()
        return self._workbook

    @property
    def worksheet(self) -> Worksheet:
        """Get active worksheet."""
        if self._worksheet is None:
            self._worksheet = self.workbook.active
        return self._worksheet

    @handle_error
    def _register_styles(self):
        """Register all named styles at workbook level."""
        if self._styles_registered:
            return

        # Currency style
        if "currency_style" not in self.workbook.named_styles:
            currency_style = NamedStyle(
                name="currency_style",
                font=Font(**settings.EXCEL_FONT),
                number_format=settings.CURRENCY_FORMAT,
                alignment=Alignment(horizontal="right"),
            )
            self.workbook.add_named_style(currency_style)

        # Header style
        if "header_style" not in self.workbook.named_styles:
            header_style = NamedStyle(
                name="header_style",
                font=Font(
                    name=settings.EXCEL_FONT["name"],
                    size=settings.EXCEL_FONT["size"],
                    bold=True,
                    color=settings.COLORS["header_bold"],
                ),
                fill=PatternFill(
                    start_color=settings.COLORS["header_yellow"],
                    end_color=settings.COLORS["header_yellow"],
                    fill_type="solid",
                ),
                alignment=Alignment(horizontal="center", vertical="center"),
                border=self._create_border(),
            )
            self.workbook.add_named_style(header_style)

        # Data style
        if "data_style" not in self.workbook.named_styles:
            data_style = NamedStyle(
                name="data_style",
                font=Font(**settings.EXCEL_FONT),
                alignment=Alignment(horizontal="left", vertical="center"),
                border=self._create_border(),
            )
            self.workbook.add_named_style(data_style)

        self._styles_registered = True
        self.logger.info("Excel styles registered successfully")

    def _create_border(self) -> Border:
        """Create standardized border style."""
        return Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

    @handle_error
    def apply_freeze_panes(self, freeze_cell: str = None):
        """Apply freeze panes to worksheet."""
        if freeze_cell is None:
            freeze_cell = settings.FREEZE_PANES["default"]

        self.worksheet.freeze_panes = freeze_cell
        self.logger.info(f"Freeze panes applied at {freeze_cell}")

    @handle_error
    def apply_font_style(self, cell_range: str = None):
        """Apply consistent font styling to entire worksheet."""
        font = Font(**settings.EXCEL_FONT)

        if cell_range:
            for row in self.worksheet[cell_range]:
                for cell in row:
                    cell.font = font
        else:
            for row in self.worksheet.iter_rows():
                for cell in row:
                    cell.font = font

        self.logger.info(f"Font styling applied to {cell_range or 'entire sheet'}")

    @handle_error
    def apply_currency_formatting(self, start_col: int = 4, end_col: int = None):
        """Apply currency formatting to specified columns."""
        if end_col is None:
            end_col = self.worksheet.max_column

        currency_style = self.workbook.named_styles["currency_style"]

        for col in range(start_col, end_col + 1):
            col_letter = get_column_letter(col)
            for cell in self.worksheet[col_letter]:
                if isinstance(cell.value, (int, float)) and cell.value != 0:
                    cell.style = currency_style

        self.logger.info(f"Currency formatting applied to columns {start_col}-{end_col}")

    @handle_error
    def auto_adjust_column_widths(self, min_width: int = 10, max_width: int = 50):
        """Auto-adjust column widths based on content."""
        for col in range(1, self.worksheet.max_column + 1):
            col_letter = get_column_letter(col)

            # Calculate max length in column
            max_length = max(
                (
                    len(str(cell.value))
                    for cell in self.worksheet[col_letter]
                    if cell.value
                ),
                default=min_width,
            )

            # Apply width with constraints
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            self.worksheet.column_dimensions[col_letter].width = adjusted_width

        self.logger.info(f"Column widths auto-adjusted (min={min_width}, max={max_width})")

    @handle_error
    def apply_alternating_row_colors(self, start_row: int = 3):
        """Apply alternating row colors for better readability."""
        sky_blue_fill = PatternFill(
            start_color=settings.COLORS["row_sky_blue"],
            end_color=settings.COLORS["row_sky_blue"],
            fill_type="solid",
        )
        white_fill = PatternFill(
            start_color=settings.COLORS["row_white"],
            end_color=settings.COLORS["row_white"],
            fill_type="solid",
        )

        for row_num in range(start_row, self.worksheet.max_row + 1):
            fill = sky_blue_fill if row_num % 2 == 0 else white_fill

            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row_num, column=col)
                cell.fill = fill
                cell.border = self._create_border()

        self.logger.info(f"Alternating row colors applied from row {start_row}")

    @handle_error
    def apply_header_formatting(self, header_rows: List[int] = [1, 2]):
        """Apply header formatting to specified rows."""
        header_style = self.workbook.named_styles["header_style"]

        for row_num in header_rows:
            for col in range(1, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=row_num, column=col)
                cell.style = header_style

        self.logger.info(f"Header formatting applied to rows {header_rows}")

    @handle_error
    def highlight_summary_rows(
        self,
        values_to_highlight: List[str],
        search_column: int = 4,
        highlight_color: str = None,
    ):
        """Highlight entire rows containing specific values."""
        if highlight_color is None:
            highlight_color = settings.COLORS["summary_light_green"]

        highlight_fill = PatternFill(
            start_color=highlight_color, end_color=highlight_color, fill_type="solid"
        )

        highlighted_count = 0
        search_col_letter = get_column_letter(search_column)

        for cell in self.worksheet[search_col_letter]:
            if cell.value in values_to_highlight:
                # Highlight entire row
                for col in range(1, self.worksheet.max_column + 1):
                    row_cell = self.worksheet.cell(row=cell.row, column=col)
                    row_cell.fill = highlight_fill
                highlighted_count += 1

        self.logger.info(f"Highlighted {highlighted_count} summary rows")

    @handle_error
    def create_data_table(
        self, table_name: str = "DataTable", table_style: str = "TableStyleMedium9"
    ):
        """Create a formatted data table."""
        if self.worksheet.max_row < 2:
            self.logger.warning("Insufficient data for table creation")
            return

        data_range = (
            f"A1:{get_column_letter(self.worksheet.max_column)}{self.worksheet.max_row}"
        )

        table = Table(displayName=table_name, ref=data_range)
        table.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        self.worksheet.add_table(table)
        self.logger.info(f"Data table '{table_name}' created with style {table_style}")

    @handle_error
    def save(self):
        """Save the formatted workbook."""
        self.workbook.save(self.output_file)
        self.logger.info(f"Excel file saved: {self.output_file}")

    @abstractmethod
    def apply_all_formatting(self):
        """Apply all formatting operations in the correct order."""
        pass


class StandardReportFormatter(BaseExcelFormatter):
    """Standard formatter for budget and variance reports."""

    def apply_all_formatting(self, summary_wbs_list: List[str] = None):
        """Apply comprehensive formatting for standard reports."""
        try:
            self.logger.info("Applying standard report formatting")

            # Step 1: Basic formatting
            self.apply_freeze_panes()
            self.apply_font_style()
            self.auto_adjust_column_widths()

            # Step 2: Header formatting
            self.apply_header_formatting()

            # Step 3: Data formatting
            self.apply_currency_formatting()
            self.apply_alternating_row_colors()

            # Step 4: Conditional highlighting
            if summary_wbs_list:
                self.highlight_summary_rows(summary_wbs_list)

            self.logger.info("Standard report formatting completed successfully")

        except Exception as e:
            raise ExcelGenerationError(
                f"Error applying formatting: {str(e)}", "FORMATTING_ERROR", e
            )


class AnalyticsReportFormatter(BaseExcelFormatter):
    """Enhanced formatter for analytics reports with charts."""

    @handle_error
    def add_data_validation(
        self,
        cell_range: str,
        validation_formula: str,
        prompt_title: str = "Selection",
        prompt_message: str = "Select a value",
    ):
        """Add data validation to specified range."""
        dv = DataValidation(type="list", formula1=validation_formula, allow_blank=False)
        dv.prompt = prompt_message
        dv.promptTitle = prompt_title

        self.worksheet.add_data_validation(dv)
        dv.add(self.worksheet[cell_range])

        self.logger.info(f"Data validation added to {cell_range}")

    @handle_error
    def create_3d_chart(
        self,
        chart_title: str,
        data_range: str,
        categories_range: str,
        chart_position: str = "H5",
    ):
        """Create a 3D bar chart with professional styling."""
        chart = BarChart3D()
        chart.title = chart_title
        chart.style = 10  # Professional style

        data = Reference(self.worksheet, range_string=data_range)
        categories = Reference(self.worksheet, range_string=categories_range)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Remove legend if not needed
        chart.legend = None

        # Set chart size
        chart.width = 16
        chart.height = 10

        self.worksheet.add_chart(chart, chart_position)
        self.logger.info(f"3D chart '{chart_title}' created at {chart_position}")

    def apply_all_formatting(self, enable_charts: bool = True):
        """Apply comprehensive formatting for analytics reports."""
        try:
            self.logger.info("Applying analytics report formatting")

            # Basic formatting
            self.apply_freeze_panes("D3")
            self.apply_font_style()
            self.auto_adjust_column_widths()
            self.apply_header_formatting()

            # Create data table
            self.create_data_table("AnalyticsTable", "TableStyleMedium2")

            if enable_charts:
                # Charts will be added by specific analytics modules
                pass

            self.logger.info("Analytics report formatting completed successfully")

        except Exception as e:
            raise ExcelGenerationError(
                f"Error applying analytics formatting: {str(e)}",
                "ANALYTICS_FORMATTING_ERROR",
                e,
            )
