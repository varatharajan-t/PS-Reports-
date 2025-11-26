"""
Service layer for generating the Glimps of Projects Report.
This module creates interactive analytics dashboards with cross-tabulation,
dynamic charts, and statistical analysis of SAP project data.

KEY FEATURES:
- Regex-based project ID parsing from DAT files
- Cross-tabulation matrix analysis
- Interactive 3D charts (Excel) and Chart.js charts (HTML)
- Dynamic data validation dropdowns
- Company and project type mapping
- Statistical summary with multiple dimensions
"""
import os
import re
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart3D, Reference
from openpyxl.chart.series import DataPoint

from django.conf import settings
from .data_processing import BaseDataProcessor
from .error_handling import handle_error


def generate_formatted_html_with_charts(crosstab_df):
    """
    Generate a professionally formatted HTML dashboard with interactive charts.
    Uses Chart.js for web-based visualization.
    Returns a dictionary with separate HTML and JavaScript components.
    """
    html_parts = []
    script_parts = []

    # Add CSS only (Chart.js will be loaded separately in template)
    html_parts.append('''
    <style>
        .glimps-container {
            font-family: 'Bookman Old Style', 'Times New Roman', serif;
            font-size: 12px;
            margin: 20px 0;
        }
        .crosstab-table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .crosstab-table th {
            font-weight: bold;
            padding: 12px 8px;
            text-align: center;
            border: 1px solid #000000;
        }
        .crosstab-table td {
            padding: 8px;
            text-align: center;
            border: 1px solid #000000;
        }
        .chart-container {
            margin-top: 30px;
            padding: 20px;
            background-color: #f9f9f9;
            border-radius: 8px;
        }
        .company-selector {
            margin: 20px 0;
            padding: 15px;
            background-color: #e8f4f8;
            border-radius: 8px;
        }
        .company-selector label {
            font-weight: bold;
            margin-right: 10px;
        }
        .company-selector select {
            padding: 8px 15px;
            font-size: 14px;
            border: 2px solid #3498db;
            border-radius: 4px;
            background-color: white;
        }
    </style>
    ''')

    html_parts.append('<div class="glimps-container">')

    # Extract column colors
    column_colors = {
        "NIGEL": "#A6CE39",
        "NIRL": "#E15B64",
        "NLCIL": "#9B6ED2",
        "NTPL": "#68C3A3",
        "NUPPL": "#C69C6D",
        "Total": "#D1D3D4"
    }

    # Cross-tab table with colors
    html_parts.append('<h2>Project Cross-Tabulation Analysis</h2>')
    html_parts.append('<table class="crosstab-table">')

    # Header row
    html_parts.append('<thead><tr>')
    html_parts.append('<th style="background-color: #B0C4DE;">Project Type</th>')

    for col in crosstab_df.columns:
        color = column_colors.get(col, "#EAEAEA")
        header_color_light = {
            "NIGEL": "#EAF1DD",
            "NIRL": "#FDEDEC",
            "NLCIL": "#EDE3F1",
            "NTPL": "#E0F2F1",
            "NUPPL": "#F3E5D8",
            "Total": "#EAEAEA"
        }.get(col, "#EAEAEA")

        html_parts.append(f'<th style="background-color: {header_color_light};">{col}</th>')

    html_parts.append('</tr></thead>')

    # Data rows
    html_parts.append('<tbody>')
    for idx, row in crosstab_df.iterrows():
        html_parts.append('<tr>')
        html_parts.append(f'<td style="background-color: #DCE6F1;"><strong>{idx}</strong></td>')

        for col_idx, (col, value) in enumerate(row.items()):
            color = column_colors.get(col, "#EAEAEA")
            html_parts.append(f'<td style="background-color: {color};">{value}</td>')

        html_parts.append('</tr>')
    html_parts.append('</tbody>')

    html_parts.append('</table>')

    # Company selector
    html_parts.append('<div class="company-selector">')
    html_parts.append('<label for="companySelect">Select Company:</label>')
    html_parts.append('<select id="companySelect" onchange="updateChart()">')

    for col in crosstab_df.columns:
        if col != 'Total':
            html_parts.append(f'<option value="{col}">{col}</option>')

    html_parts.append('</select>')
    html_parts.append('</div>')

    # Chart container
    html_parts.append('<div class="chart-container">')
    html_parts.append('<canvas id="projectChart" width="800" height="400"></canvas>')
    html_parts.append('</div>')

    html_parts.append('</div>')

    # Prepare data for JavaScript
    project_types = [str(idx) for idx in crosstab_df.index if idx != 'Total']
    chart_data = {}

    for col in crosstab_df.columns:
        if col != 'Total':
            chart_data[col] = [int(crosstab_df.loc[pt, col]) if pt in crosstab_df.index else 0 for pt in project_types]

    # JavaScript for interactive chart (to be loaded separately)
    script_parts.append(f'''
    const projectTypes = {project_types};
    const chartData = {chart_data};
    const colors = ["#00FF00", "#FF0000", "#0000FF", "#FFFF00", "#FF00FF", "#00FFFF"];

    let chartInstance = null;

    function updateChart() {{
        const company = document.getElementById("companySelect").value;
        const data = chartData[company];

        if (chartInstance) {{
            chartInstance.destroy();
        }}

        const ctx = document.getElementById("projectChart").getContext("2d");
        chartInstance = new Chart(ctx, {{
            type: "bar",
            data: {{
                labels: projectTypes,
                datasets: [{{
                    label: company + " Projects",
                    data: data,
                    backgroundColor: colors,
                    borderColor: colors.map(c => c),
                    borderWidth: 1
                }}]
            }},
            options: {{
                responsive: true,
                plugins: {{
                    title: {{
                        display: true,
                        text: "Project Type vs Number of Projects - " + company,
                        font: {{
                            size: 18
                        }}
                    }},
                    legend: {{
                        display: false
                    }},
                    datalabels: {{
                        display: false
                    }}
                }},
                scales: {{
                    y: {{
                        beginAtZero: true,
                        ticks: {{
                            stepSize: 1,
                            precision: 0
                        }},
                        title: {{
                            display: true,
                            text: "Number of Projects"
                        }}
                    }},
                    x: {{
                        title: {{
                            display: true,
                            text: "Project Type"
                        }}
                    }}
                }}
            }}
        }});
    }}

    // Initialize chart with first company
    document.addEventListener('DOMContentLoaded', function() {{
        updateChart();
    }});
    ''')

    return {
        'html': ''.join(html_parts),
        'script': ''.join(script_parts)
    }


class GlimpsOfProjectsProcessor(BaseDataProcessor):
    """Handles all data processing operations for the Glimps of Projects Report."""

    def __init__(self, input_file_path: str):
        super().__init__('GlimpsOfProjects')
        self.input_file_path = Path(input_file_path)
        self.project_data = []

    def validate_input(self, file_path: str) -> bool:
        """Validate the input DAT file format and content."""
        file_path_obj = Path(file_path)

        if not file_path_obj.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if file_path_obj.stat().st_size == 0:
            raise ValueError("DAT file is empty")

        if file_path_obj.suffix.lower() not in ['.dat', '.txt']:
            self.logger.warning(f"File does not have .dat extension: {file_path}")

        self.logger.info(f"Input file validation passed: {file_path}")
        return True

    def process_data(self, file_path: str):
        """
        Extracts project data from DAT file using regex patterns.
        Returns crosstab DataFrame for analysis.
        """
        project_id_pattern = settings.REGEX_PATTERNS["project_id"]  # r"PRJ\s+([A-Z0-9-]+)"
        company_codes = settings.COMPANY_CODES
        project_types = settings.PROJECT_TYPES

        self.project_data = []

        with open(file_path, "r", encoding="utf-8", errors="ignore") as file:
            for line in file:
                match = re.search(project_id_pattern, line)
                if match:
                    project_id = match.group(1)

                    # Extract company code (first 2 characters)
                    company_code = project_id[:2]
                    company_desc = company_codes.get(company_code, "Unknown Company")

                    # Extract project type (4th character, index 3)
                    if len(project_id) > 3:
                        project_type_code = project_id[3]
                        project_desc = project_types.get(project_type_code, "Unknown Project Type")
                    else:
                        project_desc = "Unknown Project Type"

                    self.project_data.append(
                        {
                            "Project Type": project_desc,
                            "Company": company_desc,
                            "Project ID": project_id,
                        }
                    )

        # Generate cross-tabulation
        if self.project_data:
            df = pd.DataFrame(self.project_data)
            crosstab = pd.crosstab(
                df["Project Type"],
                df["Company"],
                margins=True,
                margins_name="Total"
            )
            return crosstab
        else:
            raise ValueError("No project data found in the DAT file")


class GlimpsExcelFormatter:
    """Handles the Excel formatting for the Glimps of Projects Report."""

    def __init__(self, crosstab_df: pd.DataFrame, output_file: str):
        self.crosstab_df = crosstab_df
        self.output_file = output_file
        self.workbook = None
        self.worksheet = None

    @handle_error
    def format_and_save(self):
        """Format and save the Excel file with charts and data validation."""
        # Write crosstab to Excel
        self.crosstab_df.to_excel(self.output_file)

        # Load workbook for advanced formatting
        self.workbook = openpyxl.load_workbook(self.output_file)
        self.worksheet = self.workbook.active

        # Apply styling
        self._apply_column_colors()
        self._add_data_validation()
        self._add_dynamic_reference_section()
        self._add_chart()

        self.workbook.save(self.output_file)
        return self.output_file

    def _apply_column_colors(self):
        """Applies color coding to columns based on company."""
        self.worksheet.column_dimensions["A"].width = 15

        column_colors = {
            "A": ("DCE6F1", "B0C4DE"),  # Project Type
            "B": ("EAF1DD", "A6CE39"),  # NIGEL
            "C": ("FDEDEC", "E15B64"),  # NIRL
            "D": ("EDE3F1", "9B6ED2"),  # NLCIL
            "E": ("E0F2F1", "68C3A3"),  # NTPL
            "F": ("F3E5D8", "C69C6D"),  # NUPPL
            "G": ("EAEAEA", "D1D3D4"),  # Total
        }

        for col_letter, (header_color, cell_color) in column_colors.items():
            header_fill = PatternFill(
                start_color=header_color, end_color=header_color, fill_type="solid"
            )
            cell_fill = PatternFill(
                start_color=cell_color, end_color=cell_color, fill_type="solid"
            )

            for row in range(1, self.worksheet.max_row + 1):
                cell = self.worksheet[f"{col_letter}{row}"]
                cell.fill = header_fill if row == 1 else cell_fill
                if row == 1:
                    cell.font = Font(bold=True, color="000000")

    def _add_data_validation(self):
        """Adds data validation dropdown for company selection."""
        dv = DataValidation(type="list", formula1="B1:F1", allow_blank=False)
        dv.prompt = "Please select a company from the list"
        dv.promptTitle = "Company List"

        self.worksheet.add_data_validation(dv)
        self.worksheet["A10"].value = "Select Company"
        self.worksheet["A10"].font = Font(bold=True)
        dv.add(self.worksheet["A10"])

    def _add_dynamic_reference_section(self):
        """Adds dynamic reference section with INDEX-MATCH formulas."""
        self.worksheet["B10"].value = "No. of Projects"
        self.worksheet["B10"].font = Font(bold=True)

        # Copy project types to column A (rows 11-16)
        for row in range(11, 17):
            self.worksheet[f"A{row}"].value = f"=A{row - 9}"

        # Add INDEX-MATCH formulas for dynamic data
        for row in range(11, 17):
            self.worksheet[f"B{row}"].value = (
                f"=INDEX($B$2:$F$7, ROW()-10, MATCH($A$10, $B$1:$F$1, 0))"
            )

    def _add_chart(self):
        """Adds a 3D bar chart to the Excel sheet."""
        chart = BarChart3D()
        chart.title = "Project Type Vs No of Projects"
        chart.style = 34  # Gradient style

        # Data references
        selected_col = Reference(self.worksheet, min_col=1, min_row=11, max_row=16)
        values = Reference(self.worksheet, min_col=2, min_row=11, max_row=16)

        chart.add_data(values, titles_from_data=False)
        chart.set_categories(selected_col)

        # Remove legend
        chart.legend = None

        # Add data labels
        for series in chart.series:
            series.dLbls = openpyxl.chart.label.DataLabelList()
            series.dLbls.showVal = True
            series.dLbls.showSerName = False
            series.dLbls.showCatName = False

        # Configure axes
        chart.y_axis.numFmt = "0"
        chart.y_axis.tickLblPos = "nextTo"

        # Apply colors to bars
        colors = ["00FF00", "FF0000", "0000FF", "FFFF00", "FF00FF", "00FFFF"]
        for i in range(6):
            point = DataPoint(idx=i)
            point.graphicalProperties.solidFill = colors[i]
            if len(chart.series) > 0:
                if not hasattr(chart.series[0], 'dPt'):
                    chart.series[0].dPt = []

        # Chart size
        chart.width = 16
        chart.height = 10

        self.worksheet.add_chart(chart, "I3")


@handle_error
def generate_glimps_of_projects_report(uploaded_file_path: str) -> dict:
    """
    Orchestrates the entire glimps of projects report generation process.
    Returns a dictionary containing the path to the formatted Excel report
    and an HTML representation with interactive charts.
    """
    processor = GlimpsOfProjectsProcessor(uploaded_file_path)

    # Validate the input file before processing
    processor.validate_input(uploaded_file_path)

    # Process the data
    crosstab_df = processor.process_data(uploaded_file_path)

    # Generate output filename and path
    output_filename = "CrossTab_Report.xlsx"
    reports_dir = settings.BASE_DIR / 'data' / 'reports'
    reports_dir.mkdir(parents=True, exist_ok=True)
    output_path = str(reports_dir / output_filename)

    # Format and save Excel file
    formatter = GlimpsExcelFormatter(crosstab_df, output_path)
    formatted_file_path = formatter.format_and_save()

    # Generate HTML output with interactive charts
    chart_output = generate_formatted_html_with_charts(crosstab_df)

    return {
        "file_path": formatted_file_path,
        "data_html": chart_output['html'],
        "chart_script": chart_output['script']
    }
