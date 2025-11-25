"""
Service layer for generating the Plan Variance Report.
This module processes DAT files from SAP and generates formatted Excel reports
with WBS classification and currency formatting.

Logic derived from: GUI Env/PlanVariance.py
"""
import os
import re
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from django.conf import settings
from .data_processing import BaseDataProcessor
from .formatting import BaseExcelFormatter
from .wbs_master_data import MasterDataManager
from .error_handling import handle_error


def generate_formatted_html(df, summary_wbs_list):
    """
    Generate a professionally formatted HTML table with Indian currency formatting.
    """
    # Identify currency columns (skip first few non-currency columns)
    non_currency_cols = ['Sl No.']
    currency_columns = []

    for col in df.columns:
        col_lower = str(col).lower()
        if col not in non_currency_cols and not any(x in col_lower for x in ['level', 'description', 'id', 'wbs', 'object', 'name']):
            currency_columns.append(col)

    # Find the ID column for highlighting summary WBS
    id_column = None
    for col in df.columns:
        if 'id' in str(col).lower() or 'wbs' in str(col).lower():
            id_column = col
            break

    # Build HTML table
    html_parts = []

    # Add CSS styling
    html_parts.append('''
    <style>
        .plan-variance-table {
            width: 100%;
            border-collapse: collapse;
            font-family: 'Bookman Old Style', 'Times New Roman', serif;
            font-size: 12px;
            margin: 20px 0;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .plan-variance-table th {
            background-color: #FFFF00 !important;
            color: #000000;
            font-weight: bold;
            padding: 12px 8px;
            text-align: left;
            border: 1px solid #000000;
            position: -webkit-sticky;
            position: sticky;
            top: 0;
            z-index: 10;
            box-shadow: 0 2px 2px -1px rgba(0, 0, 0, 0.4);
        }
        .plan-variance-table td {
            padding: 8px;
            border: 1px solid #000000;
        }
        .plan-variance-table tbody tr:nth-child(even) {
            background-color: #87CEEB;
        }
        .plan-variance-table tbody tr:nth-child(odd) {
            background-color: #FFFFFF;
        }
        .plan-variance-table tbody tr.summary-wbs {
            background-color: #90EE90 !important;
            font-weight: bold;
        }
        .plan-variance-table .currency-cell {
            text-align: right;
            font-family: 'Courier New', monospace;
        }
        .plan-variance-table .text-cell {
            text-align: left;
        }
        .plan-variance-table .center-cell {
            text-align: center;
        }
        .plan-variance-table .negative {
            color: #FF0000;
        }
        .plan-variance-table-container {
            overflow-x: auto;
            max-width: 100%;
        }
    </style>
    ''')

    html_parts.append('<div class="plan-variance-table-container">')
    html_parts.append('<table class="plan-variance-table">')

    # Table header
    html_parts.append('<thead><tr>')
    for col in df.columns:
        html_parts.append(f'<th>{col}</th>')
    html_parts.append('</tr></thead>')

    # Table body
    html_parts.append('<tbody>')

    for idx, row in df.iterrows():
        # Check if this row is a summary WBS
        row_class = ''
        if id_column and summary_wbs_list:
            wbs_value = row.get(id_column, '')
            if wbs_value in summary_wbs_list:
                row_class = ' class="summary-wbs"'

        html_parts.append(f'<tr{row_class}>')

        for col in df.columns:
            value = row[col]

            if col in currency_columns:
                # Format as Indian currency
                from .formatting import format_indian_currency
                formatted_value = format_indian_currency(value)
                cell_class = 'currency-cell'
                if formatted_value.startswith('-'):
                    cell_class += ' negative'
                html_parts.append(f'<td class="{cell_class}">{formatted_value}</td>')
            elif col == 'Sl No.':
                # Center align serial numbers
                html_parts.append(f'<td class="center-cell">{value}</td>')
            else:
                # Text columns - left align
                html_parts.append(f'<td class="text-cell">{value if pd.notna(value) else ""}</td>')

        html_parts.append('</tr>')

    html_parts.append('</tbody>')
    html_parts.append('</table>')
    html_parts.append('</div>')

    return ''.join(html_parts)


class PlanVarianceProcessor(BaseDataProcessor):
    """Handles all data processing operations for the Plan Variance Report."""

    def __init__(self, input_file_path: str):
        super().__init__('PlanVariance')
        self.input_file_path = Path(input_file_path)
        self.df = None
        self.summary_wbs_list = []
        self.transaction_wbs_list = []

    def validate_input(self, file_path: str) -> bool:
        """Validate the input DAT file format and content."""
        file_path_obj = Path(file_path)

        if not file_path_obj.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if file_path_obj.stat().st_size == 0:
            raise ValueError("DAT file is empty")

        if file_path_obj.suffix.lower() not in ['.dat', '.txt']:
            self.logger.warning(f"File does not have .dat extension: {file_path}")

        self.logger.info(f"Input file validation passed: {file_path}")
        return True

    def clean_data(self, input_path: str, cleaned_path: str):
        """Clean the DAT file by removing unnecessary lines and asterisks."""
        with open(input_path, 'r', encoding='iso-8859-1') as file:
            lines = file.readlines()

        total_lines = len(lines)

        # Remove the first, fourth, and last line (0-indexed: 0, 3, total_lines-1)
        lines_to_delete = [0, 3, total_lines - 1]
        lines = [line for i, line in enumerate(lines) if i not in lines_to_delete]

        # Replace asterisks with spaces
        lines = [line.replace("*", " ") for line in lines]

        with open(cleaned_path, 'w', encoding='iso-8859-1') as file:
            file.writelines(lines)

        self.logger.info(f"Data cleaned and saved to {cleaned_path}")
        return cleaned_path

    def process_data(self, cleaned_path: str):
        """Process the cleaned DAT file into a formatted DataFrame."""
        # Read the DAT file with multi-level headers
        self.df = pd.read_csv(
            cleaned_path,
            sep="\t",
            header=[0, 1],
            encoding="iso-8859-1"
        )

        # Convert columns to tuples
        self.df.columns = self.df.columns.map(lambda x: tuple(map(str, x)))

        df_columns = list(self.df.columns[1:])

        # Rename the first column header
        self.df = self.df.rename(
            columns={"Unnamed: 0_level_0": "WBS_Elements_Info."}
        )

        # Split "WBS Element Details" column into separate columns
        split_details = self.df[("WBS_Elements_Info.", "Object")].str.split()

        level, description, id_no = [], [], []

        for details in split_details:
            Level = details[0]
            Description = " ".join(details[1:])
            ID = details[1]

            if Level not in ["*", "**", "***", "4*", "5*"]:
                Description = Level.strip() + Description.strip()
                Level = " "

            level.append(Level)
            description.append(Description)
            id_no.append(ID)

        # Add the new columns to the DataFrame
        self.df[("WBS_Elements_Info.", "Level")] = level
        self.df[("WBS_Elements_Info.", "Description")] = description
        self.df[("WBS_Elements_Info.", "ID_No")] = id_no

        # Re-arrange columns
        self.df = self.df[
            [
                ("WBS_Elements_Info.", "Level"),
                ("WBS_Elements_Info.", "Description"),
                ("WBS_Elements_Info.", "ID_No"),
            ]
            + df_columns
        ]

        # Drop rows with all NaN values
        self.df = self.df.dropna(how='all')

        # Process WBS classification
        self.summary_wbs_list, self.transaction_wbs_list = self.process_wbs(
            self.df[("WBS_Elements_Info.", "ID_No")].to_list()
        )

        return self.df, (self.summary_wbs_list, self.transaction_wbs_list)

    def process_wbs(self, wbs_list):
        """Classifies WBS elements into summary and transaction WBS dynamically."""
        summary_wbs = []
        transaction_wbs = []

        for wbs in wbs_list:
            # Define a regular expression pattern to match a WBS element with child elements
            pattern = re.compile(re.escape(wbs) + r"-\d{2}")

            # Check if any WBS starts with the parent WBS followed by a hyphen and two digits
            has_child = any(pattern.fullmatch(item) for item in wbs_list if item != wbs)

            if has_child:
                summary_wbs.append(wbs)
            else:
                transaction_wbs.append(wbs)

        return summary_wbs, transaction_wbs


class PlanVarianceExcelFormatter(BaseExcelFormatter):
    """Handles the Excel formatting for the Plan Variance Report."""

    def __init__(self, df: pd.DataFrame, output_file: str, summary_wbs_list: list):
        self.df = df
        self.output_file = output_file
        self.summary_wbs_list = summary_wbs_list
        self.workbook = None
        self.worksheet = None

    def format_and_save(self):
        """Apply all formatting and save the Excel file."""
        # Flatten MultiIndex columns to strings for Excel compatibility
        df_output = self.df.copy()
        df_output.columns = [' - '.join(col).strip() if isinstance(col, tuple) else str(col) for col in df_output.columns]

        # Add serial number column at the beginning
        df_output.insert(0, 'Sl No.', range(1, len(df_output) + 1))

        # Save to Excel
        df_output.to_excel(self.output_file, index=False)

        # Load workbook for formatting
        self.workbook = openpyxl.load_workbook(self.output_file)
        self.worksheet = self.workbook.active

        # Apply formatting
        self.apply_font_style()
        self.apply_currency_format()
        self.adjust_column_widths()
        self.apply_header_format()
        self.apply_table_formatting()
        self.highlight_summary_wbs()
        self.apply_freeze_panes()

        # Save formatted workbook
        self.workbook.save(self.output_file)
        self.logger.info(f"Formatted Excel file saved: {self.output_file}")
        return self.output_file

    def apply_freeze_panes(self):
        """Freeze panes for the Excel sheet."""
        self.worksheet.freeze_panes = "E3"

    def highlight_summary_wbs(self):
        """Highlight summary WBS rows in green."""
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        # Find the ID_No column
        id_col_index = None
        for col_idx, cell in enumerate(self.worksheet[1], 1):
            if 'ID' in str(cell.value) or 'WBS' in str(cell.value):
                id_col_index = col_idx
                break

        if id_col_index and self.summary_wbs_list:
            for row in self.worksheet.iter_rows(min_row=2, max_row=self.worksheet.max_row):
                cell_value = row[id_col_index - 1].value
                if cell_value in self.summary_wbs_list:
                    for cell in row:
                        cell.fill = green_fill


@handle_error
def generate_plan_variance_report(uploaded_file_path: str) -> dict:
    """
    Orchestrates the entire plan variance report generation process.
    Returns a dictionary containing the path to the formatted Excel report
    and an HTML representation of the data.
    """
    processor = PlanVarianceProcessor(uploaded_file_path)

    # Validate the input file before processing
    processor.validate_input(uploaded_file_path)

    # Clean the data
    cleaned_path = str(Path(uploaded_file_path).parent / 'cleaned_temp.dat')
    processor.clean_data(uploaded_file_path, cleaned_path)

    # Process the data
    df, (summary_wbs, transaction_wbs) = processor.process_data(cleaned_path)

    # Map WBS descriptions from master data
    master_data_manager = MasterDataManager()
    df = master_data_manager.map_wbs_descriptions(
        transaction_df=df,
        wbs_column=("WBS_Elements_Info.", "ID_No"),
        description_column=("WBS_Elements_Info.", "Description")
    )

    # Generate output filename and path
    output_filename = Path(uploaded_file_path).with_suffix(".xlsx").name
    reports_dir = settings.BASE_DIR / 'data' / 'reports'
    reports_dir.mkdir(parents=True, exist_ok=True)
    output_path = str(reports_dir / output_filename)

    # Format and save Excel file
    formatter = PlanVarianceExcelFormatter(df.copy(), output_path, summary_wbs)
    formatted_file_path = formatter.format_and_save()

    # Clean up temporary file
    os.remove(cleaned_path)

    # Generate HTML table from DataFrame for web display
    df_html = df.copy()
    df_html.columns = [' - '.join(col).strip() if isinstance(col, tuple) else str(col) for col in df_html.columns]
    df_html.insert(0, 'Sl No.', range(1, len(df_html) + 1))

    html_output = generate_formatted_html(df_html, summary_wbs)

    return {
        "file_path": formatted_file_path,
        "data_html": html_output
    }
