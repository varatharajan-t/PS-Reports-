"""
Service layer for generating the Project Analysis Report.
This module processes two DAT files (Budget and Plan) and creates an interactive
Excel report with data validation, formulas, and 3D charts.

Logic derived from: GUI Env/ProjectAnalysis.py
"""
import os
import re
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart3D, Reference
from openpyxl.styles import Font, PatternFill, Alignment
import logging

from django.conf import settings
from .data_processing import BaseDataProcessor
from .error_handling import handle_error, FileProcessingError


logger = logging.getLogger(__name__)


class ProjectAnalysisProcessor(BaseDataProcessor):
    """Processes DAT files for Project Analysis report."""

    def __init__(self):
        super().__init__("ProjectAnalysisProcessor")

    def validate_input(self, file_path: str) -> bool:
        """Validate that the input file exists and is a DAT file."""
        if not Path(file_path).exists():
            raise FileProcessingError(f"File not found: {file_path}", "FILE_NOT_FOUND")

        if not file_path.lower().endswith('.dat'):
            raise FileProcessingError(
                "Invalid file format. Expected .DAT file",
                "INVALID_FORMAT"
            )
        return True

    def process_data(self, file_path: str, file_type: str) -> pd.DataFrame:
        """
        Process a single DAT file (either budget or plan).

        Args:
            file_path: Path to the DAT file
            file_type: Either 'budget' or 'plan'

        Returns:
            Processed DataFrame with ProjectID and ProjectName columns
        """
        self.validate_input(file_path)

        # Read the DAT file with multi-level headers
        df = self.read_dat_file(
            file_path,
            delimiter="\t",
            header_rows=[0, 1],
            encoding="iso-8859-1"
        )

        # Remove result/summary rows
        df = df[~df.iloc[:, 0].astype(str).str.startswith("Result", na=False)]

        # Extract ProjectID and ProjectName based on file type
        if file_type == "budget":
            extract_pattern = settings.PROJECT_ANALYSIS_REGEX["budget"]
            # For budget file: pattern extracts (ProjectName, ProjectID)
            df[["ProjectName", "ProjectID"]] = (
                df[("Unnamed: 0_level_0", "Object")]
                .str.extract(extract_pattern)
                .fillna("")
            )
        else:  # plan
            extract_pattern = settings.PROJECT_ANALYSIS_REGEX["plan"]
            # For plan file: pattern extracts (ProjectID, ProjectName)
            df[["ProjectID", "ProjectName"]] = (
                df[("Unnamed: 0_level_0", "Object")]
                .str.extract(extract_pattern)
                .fillna("")
            )

        # Clean up the extracted data
        df["ProjectID"] = df["ProjectID"].str.strip()
        df["ProjectName"] = df["ProjectName"].str.strip()

        self.logger.info(f"Processed {file_type} file: {len(df)} rows")
        return df


def write_df_to_sheet(workbook, df, sheet_name):
    """
    Write a DataFrame to an Excel sheet.

    Args:
        workbook: openpyxl Workbook object
        df: pandas DataFrame
        sheet_name: Name for the sheet

    Returns:
        The created worksheet
    """
    ws = workbook.create_sheet(sheet_name)

    # Write data to sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Format header row
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True, size=12)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    logger.info(f"Created sheet '{sheet_name}' with {ws.max_row} rows")
    return ws


def create_analysis_sheet(workbook, budget_sheet, last_row, last_col_letter):
    """
    Create the interactive Analysis sheet with dropdowns and formulas.

    Args:
        workbook: openpyxl Workbook
        budget_sheet: Reference to the Budget worksheet
        last_row: Last row number in Budget sheet
        last_col_letter: Last column letter in Budget sheet
    """
    ws_analysis = workbook.create_sheet("Analysis")

    # ProjectID column (assumed to be in column AL based on legacy code)
    # We'll dynamically find it, but default to AL if not found
    project_id_col = "AL"

    # Create data validation for ProjectID dropdown
    dv_project = DataValidation(
        type="list",
        formula1=f"=Budget!${project_id_col}$3:${project_id_col}${last_row}",
        allow_blank=False
    )

    # Create data validation for Year/Header dropdown
    dv_header = DataValidation(
        type="list",
        formula1=f"=Budget!$D$1:${last_col_letter}$1",
        allow_blank=False
    )

    ws_analysis.add_data_validation(dv_project)
    ws_analysis.add_data_validation(dv_header)

    # Add dropdown labels and cells
    ws_analysis["A1"] = "Select ProjectID:"
    ws_analysis["A1"].font = Font(bold=True, size=12)
    dv_project.add(ws_analysis["B1"])

    ws_analysis["A2"] = "Select Year range:"
    ws_analysis["A2"].font = Font(bold=True, size=12)
    dv_header.add(ws_analysis["B2"])

    # Add analysis table headers
    headers = ["Plan", "Budget", "Actual", "Commitment", "Assigned", "Available"]
    header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    for i, header in enumerate(headers, start=3):
        cell = ws_analysis.cell(row=3, column=i, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")

        # Add INDEX-MATCH formula for dynamic lookup
        formula_cell = ws_analysis.cell(row=4, column=i)
        formula_cell.value = (
            f'=IFERROR(INDEX(Budget!$D$2:${last_col_letter}${last_row}, '
            f'MATCH($B$1, Budget!$C$2:$C${last_row}, 0), '
            f'MATCH($B$2, Budget!$D$1:${last_col_letter}$1, 0)), "")'
        )

    # Set column widths
    ws_analysis.column_dimensions['A'].width = 20
    ws_analysis.column_dimensions['B'].width = 25

    logger.info("Created Analysis sheet with interactive dropdowns")
    return ws_analysis


def create_3d_chart(worksheet, title, data_col, position):
    """
    Create a 3D bar chart.

    Args:
        worksheet: Worksheet to add chart to
        title: Chart title
        data_col: Column number for data
        position: Cell position for chart (e.g., "D10")
    """
    chart = BarChart3D()
    chart.title = title
    chart.style = 10  # Professional style

    # Data range (row 4 to row 9)
    data = Reference(worksheet, min_col=data_col, min_row=4, max_row=9)
    categories = Reference(worksheet, min_col=2, min_row=4, max_row=9)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Set chart size
    chart.width = 15
    chart.height = 10

    worksheet.add_chart(chart, position)
    logger.info(f"Created 3D chart '{title}' at position {position}")


@handle_error
def generate_project_analysis_report(budget_file_path: str, plan_file_path: str) -> dict:
    """
    Generate the Project Analysis report from two DAT files.

    Args:
        budget_file_path: Path to Budget DAT file (All_Projects.DAT)
        plan_file_path: Path to Plan DAT file (All_Projects_Plan.DAT)

    Returns:
        Dictionary with file_path and data_html
    """
    processor = ProjectAnalysisProcessor()

    # Process both files
    logger.info("Processing budget file...")
    df_budget = processor.process_data(budget_file_path, "budget")

    logger.info("Processing plan/actual file...")
    df_actual = processor.process_data(plan_file_path, "plan")

    # Merge data on ProjectID
    logger.info("Merging budget and actual data...")
    df_combined = pd.merge(
        df_budget,
        df_actual,
        on=[("ProjectID", "")],
        how="outer",
        suffixes=("_Budget", "_Actual"),
    )

    # Create Excel workbook
    wb = openpyxl.Workbook()
    del wb["Sheet"]  # Remove default sheet

    # Write DataFrames to sheets
    logger.info("Creating Excel sheets...")
    ws_budget = write_df_to_sheet(wb, df_budget, "Budget")
    ws_actual = write_df_to_sheet(wb, df_actual, "Actual")
    ws_data = write_df_to_sheet(wb, df_combined, "Data")

    # Create interactive Analysis sheet
    last_row = ws_budget.max_row
    last_col_letter = get_column_letter(ws_budget.max_column)

    ws_analysis = create_analysis_sheet(wb, ws_budget, last_row, last_col_letter)

    # Add 3D charts to Analysis sheet
    logger.info("Adding charts to Analysis sheet...")
    chart_definitions = [
        ("D10", 3, "Plan Vs Budget"),
        ("D25", 4, "Budget Vs Actual"),
        ("M10", 5, "Budget Vs Commitment"),
        ("M25", 6, "Budget Vs Available"),
    ]

    for position, col, title in chart_definitions:
        create_3d_chart(ws_analysis, title, col, position)

    # Generate output filename and path
    output_filename = "ProjectAnalysis.xlsx"
    reports_dir = settings.BASE_DIR / 'data' / 'reports'
    reports_dir.mkdir(parents=True, exist_ok=True)
    output_path = str(reports_dir / output_filename)

    # Save workbook
    wb.save(output_path)
    logger.info(f"Project Analysis report saved: {output_path}")

    # Generate HTML preview of combined data (first 100 rows)
    df_preview = df_combined.head(100)
    df_preview.columns = [
        ' - '.join(col).strip() if isinstance(col, tuple) else str(col)
        for col in df_preview.columns
    ]

    data_html = df_preview.to_html(
        classes='table table-striped table-bordered table-hover',
        index=False,
        na_rep='',
        float_format=lambda x: f'{x:,.2f}' if pd.notna(x) else ''
    )

    return {
        "file_path": output_path,
        "data_html": data_html,
        "rows_processed": len(df_combined),
        "budget_projects": len(df_budget),
        "actual_projects": len(df_actual),
    }
