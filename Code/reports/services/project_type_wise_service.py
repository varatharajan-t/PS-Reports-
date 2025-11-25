"""
Service layer for generating the Project Type Wise Report.
This module reads Excel files containing project data and creates a summary
with hyperlinked navigation to detailed project sheets grouped by company and type.

KEY FEATURES:
- Reads EXCEL input files (not DAT/HTML)
- Groups projects by Company Code and Project Type
- Creates multiple sheets with hyperlinks
- Summary sheet with navigation links
- Applies company/project type mappings from config
"""
import os
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from django.conf import settings
from .data_processing import BaseDataProcessor
from .error_handling import handle_error


def generate_formatted_html(summary_df, grouped_data):
    """
    Generate a professionally formatted HTML summary with expandable sections.
    """
    html_parts = []

    # Add CSS styling
    html_parts.append('''
    <style>
        .project-type-wise-container {
            font-family: 'Bookman Old Style', 'Times New Roman', serif;
            font-size: 12px;
            margin: 20px 0;
        }
        .summary-table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 30px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .summary-table th {
            background-color: #FFFF00;
            color: #000000;
            font-weight: bold;
            padding: 12px 8px;
            text-align: left;
            border: 1px solid #000000;
        }
        .summary-table td {
            padding: 8px;
            border: 1px solid #000000;
        }
        .summary-table tbody tr:hover {
            background-color: #f5f5f5;
            cursor: pointer;
        }
        .detail-section {
            margin-top: 30px;
            padding: 15px;
            background-color: #f9f9f9;
            border-radius: 8px;
        }
        .detail-section h3 {
            color: #2c3e50;
            margin-bottom: 15px;
        }
        .detail-table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
            background-color: white;
        }
        .detail-table th {
            background-color: #FFFF00;
            color: #000000;
            font-weight: bold;
            padding: 10px 8px;
            text-align: left;
            border: 1px solid #000000;
        }
        .detail-table td {
            padding: 8px;
            border: 1px solid #ddd;
        }
        .detail-table tbody tr:nth-child(even) {
            background-color: #ecf0f1;
        }
        .expand-link {
            color: #3498db;
            text-decoration: none;
            font-weight: bold;
        }
        .expand-link:hover {
            text-decoration: underline;
        }
    </style>
    ''')

    html_parts.append('<div class="project-type-wise-container">')

    # Summary Table
    html_parts.append('<h2>Project Summary</h2>')
    html_parts.append('<table class="summary-table">')
    html_parts.append('<thead><tr>')
    html_parts.append('<th>Project Type</th>')
    html_parts.append('<th>Number of Unique Projects</th>')
    html_parts.append('</tr></thead>')
    html_parts.append('<tbody>')

    for _, row in summary_df.iterrows():
        project_name = row['Project']
        project_count = row['Number of Unique Projects']
        # Create anchor link to detail section
        html_parts.append(f'<tr onclick="document.getElementById(\'{project_name}\').scrollIntoView({{behavior: \'smooth\'}})">')
        html_parts.append(f'<td>{project_name}</td>')
        html_parts.append(f'<td><a href="#{project_name}" class="expand-link">{project_count}</a></td>')
        html_parts.append('</tr>')

    html_parts.append('</tbody></table>')

    # Detail Sections
    for project_name, group_df in grouped_data:
        html_parts.append(f'<div class="detail-section" id="{project_name}">')
        html_parts.append(f'<h3>{project_name} - Projects</h3>')
        html_parts.append('<table class="detail-table">')

        # Table headers
        html_parts.append('<thead><tr>')
        for col in group_df.columns:
            html_parts.append(f'<th>{col}</th>')
        html_parts.append('</tr></thead>')

        # Table body
        html_parts.append('<tbody>')
        for _, row in group_df.iterrows():
            html_parts.append('<tr>')
            for val in row:
                html_parts.append(f'<td>{val if pd.notna(val) else ""}</td>')
            html_parts.append('</tr>')
        html_parts.append('</tbody>')

        html_parts.append('</table>')
        html_parts.append('</div>')

    html_parts.append('</div>')

    return ''.join(html_parts)


class ProjectTypeWiseProcessor(BaseDataProcessor):
    """Handles all data processing operations for the Project Type Wise Report."""

    def __init__(self, input_file_path: str):
        super().__init__('ProjectTypeWise')
        self.input_file_path = Path(input_file_path)
        self.df = None
        self.summary = None

    def validate_input(self, file_path: str) -> bool:
        """Validate the input Excel file format and content."""
        file_path_obj = Path(file_path)

        if not file_path_obj.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        if file_path_obj.stat().st_size == 0:
            raise ValueError("Excel file is empty")

        if file_path_obj.suffix.lower() not in ['.xlsx', '.xls']:
            raise ValueError("File must be an Excel file (.xlsx or .xls)")

        self.logger.info(f"Input file validation passed: {file_path}")
        return True

    def process_data(self, file_path: str):
        """
        Processes the Excel file into grouped project data.
        Returns DataFrame and summary information.
        """
        # Read Excel file
        self.df = pd.read_excel(file_path)

        # Validate required column exists
        if "Project definition" not in self.df.columns:
            available_columns = ", ".join(self.df.columns.tolist())
            raise ValueError(
                f"Required column 'Project definition' not found in Excel file. "
                f"Available columns: {available_columns}. "
                f"Please ensure you're uploading CN42N output saved as Excel."
            )

        # Create Group column from Project definition
        # Example: "NL-C-001-01" -> "NL-C"
        self.df["Group"] = self.df["Project definition"].apply(
            lambda x: "-".join(str(x).split("-")[:2])
        )

        # Map values using company codes and project types
        company_codes = settings.COMPANY_CODES
        project_types = settings.PROJECT_TYPES
        mapping_dict = {**company_codes, **project_types}

        self.df["Group"] = self.df["Group"].apply(
            lambda x: "-".join([mapping_dict.get(val, val) for val in str(x).split("-")])
        )

        # Create summary
        summary = (
            self.df.groupby("Group")["Project definition"]
            .nunique()
            .reset_index(name="Unique Projects")
        )
        self.summary = summary.rename(
            columns={"Group": "Project", "Unique Projects": "Number of Unique Projects"}
        )

        return self.df, self.summary


class ProjectTypeWiseExcelFormatter:
    """Handles the Excel formatting for the Project Type Wise Report."""

    def __init__(self, df: pd.DataFrame, summary: pd.DataFrame, output_file: str):
        self.df = df
        self.summary = summary
        self.output_file = output_file
        self.workbook = None

    @handle_error
    def format_and_save(self):
        """Format and save the Excel file with multiple sheets and hyperlinks."""
        # Write to Excel with multiple sheets
        with pd.ExcelWriter(self.output_file, engine="openpyxl") as writer:
            # Write main data
            self.df.to_excel(writer, index=False, sheet_name="ProjectsView")

            # Write summary
            self.summary.to_excel(writer, index=False, sheet_name="Summary")

            # Write grouped data
            grouped = self.df.groupby("Group")
            for name, group in grouped:
                # Sanitize sheet name (Excel has 31 char limit and restricted chars)
                sheet_name = str(name)[:31].replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-').replace('[', '').replace(']', '')
                group.to_excel(writer, index=False, sheet_name=sheet_name)

        # Load workbook to add hyperlinks and formatting
        self.workbook = openpyxl.load_workbook(self.output_file)

        # Add hyperlinks in summary sheet
        self._add_hyperlinks()

        # Format all sheets
        for sheet_name in self.workbook.sheetnames:
            self._format_sheet(sheet_name)

        self.workbook.save(self.output_file)
        return self.output_file

    def _add_hyperlinks(self):
        """Add hyperlinks in the summary sheet."""
        summary_sheet = self.workbook["Summary"]

        for row in range(2, summary_sheet.max_row + 1):
            project_name = summary_sheet.cell(row=row, column=1).value
            project_count = summary_sheet.cell(row=row, column=2).value

            # Sanitize sheet name for hyperlink
            sanitized_name = str(project_name)[:31].replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-').replace('[', '').replace(']', '')

            # Check if sheet exists
            if sanitized_name in self.workbook.sheetnames:
                link = Hyperlink(
                    ref=f"B{row}",
                    location=f"'{sanitized_name}'!A1",
                    display=str(project_count),
                )
                summary_sheet[f"B{row}"].hyperlink = link
                summary_sheet[f"B{row}"].style = "Hyperlink"

    def _format_sheet(self, sheet_name: str):
        """Apply common formatting to a sheet."""
        worksheet = self.workbook[sheet_name]

        # Apply Bookman Old Style font
        font = Font(name="Bookman Old Style", size=12)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Add table
        last_row = worksheet.max_row
        last_col = worksheet.max_column
        if last_row > 1 and last_col > 0:
            data_range = f"A1:{worksheet.cell(row=last_row, column=last_col).coordinate}"
            # Create unique table name
            table_name = f"Table{sheet_name.replace('-', '')[:20]}"
            try:
                table = Table(displayName=table_name, ref=data_range)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                worksheet.add_table(table)
            except:
                # If table name conflicts, skip
                pass

        # Apply white header style
        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        for cell in worksheet[1]:
            cell.fill = white_fill

        # Freeze panes
        worksheet.freeze_panes = "A2"


@handle_error
def generate_project_type_wise_report(uploaded_file_path: str) -> dict:
    """
    Orchestrates the entire project type wise report generation process.
    Returns a dictionary containing the path to the formatted Excel report
    and an HTML representation of the data.
    """
    processor = ProjectTypeWiseProcessor(uploaded_file_path)

    # Validate the input file before processing
    processor.validate_input(uploaded_file_path)

    # Process the data
    df, summary = processor.process_data(uploaded_file_path)

    # Generate output filename and path
    input_filename = Path(uploaded_file_path).stem
    output_filename = f"{input_filename}Summary.xlsx"
    reports_dir = settings.BASE_DIR / 'data' / 'reports'
    reports_dir.mkdir(parents=True, exist_ok=True)
    output_path = str(reports_dir / output_filename)

    # Format and save Excel file
    formatter = ProjectTypeWiseExcelFormatter(df, summary, output_path)
    formatted_file_path = formatter.format_and_save()

    # Generate HTML output
    grouped_data = list(df.groupby("Group"))
    html_output = generate_formatted_html(summary, grouped_data)

    return {
        "file_path": formatted_file_path,
        "data_html": html_output
    }
