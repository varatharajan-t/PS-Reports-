# ************************************************************************************************************************************
# Purpose: This module converts S_ALR_87013558 report output from SAP in the 'DAT' file into an Excel file with all formatting options.
# Author: Varatha Rajan T., DCE / ERP, NLC India Limited, Mobile: 9488997735
# Date: Created : 2024-11-10  Last Modified: 2025-03-12
# Version: 1.0
# Dependencies: re, os, sys, pandas, PySide6, Openpyxl, pathlib
# Usage: 1] Import the S_ALR_87013558 report and save it as a 'DAT' file for a single project.
# 	     2] Select the 'DAT' file using the file dialog. 
# 	     3] The script will process the file, clean and transform the data, and save it as an Excel file with formatting applied.
# ************************************************************************************************************************************
import os
import sys
import re
import pandas as pd
from pathlib import Path
from PySide6.QtWidgets import QFileDialog, QApplication
import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


class FileHandler:
    """Handles file selection and directory management."""

    @staticmethod
    def get_file_path_and_name():
        """Open a file dialog and return the selected file path and name."""
        filepath = QFileDialog.getOpenFileName(None, "Select a DAT file", "", "DAT Files(*.DAT)")
        if filepath[0]:
            return os.path.split(filepath[0])
        return None, None


class DataProcessor:
    """Handles reading, cleaning, and transforming data."""

    def __init__(self, file_path, file_name):
        self.file_path = file_path
        self.file_name = file_name
        self.df = None
        os.chdir(file_path)

    def clean_data(self, cleaned_DAT):
        """Clean the DAT file by removing unnecessary lines."""
        with open(self.file_name, 'r') as file:
            lines = file.readlines()
        lines_to_delete = {0, 1, 4, len(lines) - 1}
        with open(cleaned_DAT, 'w') as file:
            file.writelines(line for i, line in enumerate(lines) if i not in lines_to_delete)

    def read_data(self, cleaned_DAT, delimiter="\t"):
        """Read the cleaned DAT file into a DataFrame."""
        self.df = pd.read_csv(cleaned_DAT, sep=delimiter, header=[0, 1], encoding="iso-8859-1")
        self.df.columns = self.df.columns.map(lambda x: tuple(map(str, x)))
        return self.df

    def transform_data(self, master_file, output_file, cleaned_DAT):
        """Transform the data by splitting WBS elements and renaming columns."""
        df_cleaned = self.read_data(cleaned_DAT)
        df_cleaned = df_cleaned.rename(columns={'Unnamed: 0_level_0': 'WBS_Elements_Info.'})

        split_details = df_cleaned[('WBS_Elements_Info.', 'Object')].str.split()
        level, description, id_no = zip(*[(details[0], " ".join(details[1:-2]), details[-1]) for details in split_details])

        df_cleaned[('WBS_Elements_Info.', 'Level')] = level
        df_cleaned[('WBS_Elements_Info.', 'Description')] = description
        df_cleaned[('WBS_Elements_Info.', 'ID_No')] = id_no

        df_cleaned = df_cleaned[[('WBS_Elements_Info.', 'Level'), ('WBS_Elements_Info.', 'Description'), ('WBS_Elements_Info.', 'ID_No')] +
                                list(df_cleaned.columns[1:])]

        master_df = pd.read_excel(master_file)
        master_df['WBS_element'] = master_df['WBS_element'].str.strip()
        df_cleaned[('WBS_Elements_Info.', 'ID_No')] = df_cleaned[('WBS_Elements_Info.', 'ID_No')].str.strip()
        df_cleaned[('WBS_Elements_Info.', 'Description')] = df_cleaned[('WBS_Elements_Info.', 'ID_No')].map(
            master_df.set_index('WBS_element')['Name'])

        summary_wbs_list, transaction_wbs_list = self.process_wbs(df_cleaned[('WBS_Elements_Info.', 'ID_No')].to_list())
        df_cleaned.to_excel(output_file)
        print(f"Updated {output_file} saved successfully!")
        return summary_wbs_list

    @staticmethod
    def process_wbs(wbs_list):
        """Classify WBS elements into summary and transaction WBS dynamically."""
        summary_wbs, transaction_wbs = [], []
        for wbs in wbs_list:
            pattern = re.compile(re.escape(wbs) + r"-\d{2}")
            has_child = any(pattern.fullmatch(item) for item in wbs_list if item != wbs)
            summary_wbs.append(wbs) if has_child else transaction_wbs.append(wbs)
        return summary_wbs, transaction_wbs


class ExcelFormatter:
    """Handles Excel file formatting."""

    def __init__(self, output_file):
        self.output_file = output_file
        self.workbook = openpyxl.load_workbook(output_file)
        self.worksheet = self.workbook.active
        self.last_col = self.worksheet.max_column
        self.last_row = self.worksheet.max_row

    def apply_freeze_panes(self):
        """Freeze panes for the Excel sheet."""
        self.worksheet.freeze_panes = 'E3'

    def apply_font_style(self):
        """Apply a font style to all cells in the worksheet."""
        font = Font(name='Bookman Old Style', size=12)
        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.font = font

    def apply_currency_formatting(self):
        """Apply currency formatting to specific columns."""
        currency_style = NamedStyle(name='currency', font=Font(name='Bookman Old Style', size=12),
                                    number_format='₹ #,##0.00;[Red]₹ -#,##0.00')
        for col in range(4, self.last_col + 1):
            for cell in self.worksheet[get_column_letter(col)]:
                cell.style = currency_style

    def adjust_column_width(self):
        """Adjust column widths based on content."""
        for col in range(1, self.last_col + 1):
            col_letter = get_column_letter(col)
            max_length = max((len(str(cell.value)) for cell in self.worksheet[col_letter] if cell.value), default=10)
            self.worksheet.column_dimensions[col_letter].width = max_length + 10

    def apply_header_format(self):
        """Apply specific formatting to the header cells."""
        black_border = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"),
                             top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        skyBlue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        bold_font = Font(name='Bookman Old Style', size=12, color="000000", bold=True)

        for row in self.worksheet.iter_rows(min_row=1, max_row=self.last_row, max_col=self.last_col):
            for cell in row:
                if cell.row in [1, 2]:
                    cell.fill = yellow_fill
                    cell.font = bold_font
                else:
                    cell.fill = skyBlue_fill if cell.row % 2 == 0 else white_fill
                cell.border = black_border

    def highlight_rows_by_value(self, values_list):
        """Highlight rows where column D contains values from `values_list`."""
        light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        for cell in self.worksheet['D']:
            if cell.value in values_list:
                for row_cell in self.worksheet[cell.row]:
                    row_cell.fill = light_green_fill

    def save(self):
        """Save the final formatted workbook."""
        self.workbook.save(self.output_file)


def main():
    """Main function to execute the program."""
    app = QApplication(sys.argv)
    cleaned_DAT = 'temp.dat'
    master_file = 'WBS_NAMES.XLSX'

    file_path, file_name = FileHandler.get_file_path_and_name()
    if not file_path:
        print("No file selected. Exiting.")
        sys.exit()

    os.chdir(file_path)
    output_file = Path(file_name).with_suffix(".XLSX")

    data_processor = DataProcessor(file_path, file_name)
    data_processor.clean_data(cleaned_DAT)
    summary_wbs_list = data_processor.transform_data(master_file, output_file, cleaned_DAT)

    excel_formatter = ExcelFormatter(output_file)
    excel_formatter.apply_freeze_panes()
    excel_formatter.apply_font_style()
    excel_formatter.apply_currency_formatting()
    excel_formatter.adjust_column_width()
    excel_formatter.apply_header_format()
    excel_formatter.highlight_rows_by_value(summary_wbs_list)
    excel_formatter.save()

    print(f"Excel file '{output_file}' has been created with the specified formatting.")
    sys.exit(app.exec())


if __name__ == "__main__":
    main()