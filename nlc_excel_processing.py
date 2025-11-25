import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import platform
import subprocess

# Style utilities
def apply_header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

def apply_table_title(ws, title, row_idx, col_span):
    cell = ws.cell(row=row_idx, column=1, value=title)
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=col_span)

def apply_border_to_range(ws, start_row, end_row, start_col, end_col):
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border

def get_adjusted_width(value):
    if value is None:
        return 0
    string_val = str(value).strip().replace('₹', '')
    return len(string_val) * 1.2 + 2

def auto_adjust_column_widths(workbook):
    for ws in workbook.worksheets:
        col_widths = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    col_idx = cell.column
                    width = get_adjusted_width(cell.value)
                    if col_idx not in col_widths or width > col_widths[col_idx]:
                        col_widths[col_idx] = min(width, 60)
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

# Colors
fill_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
fill_total = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# Load input Excel
input_path = "NLC_CJI3_2025.xlsx"
df = pd.read_excel(input_path)
df.columns = df.columns.str.strip()
df = df[df['WPS_ID'].notna()]
df = df[df['WPS_ID'].astype(str).str.strip() != ""]

# ✅ FIX: Include 4 parts for proper grouping (e.g., NL-C-MTL-001)
df['Prefix'] = df['WPS_ID'].apply(lambda x: '-'.join(str(x).split('-')[:4]))

desc_map = {
    'AA': 'Asset Posting', 'WE': 'Goods Receipt', 'ZA': 'Value Transfer', 'AB': 'Accounting Document',
    'SA': 'G/L Account Doc', 'ZW': 'WBS Settlement', 'DR': 'Customer Invoice', 'RE': 'Invoice Gross',
    'IU': 'G/L Account Doc', 'KA': 'Vendor Document', 'KE': 'Emp. Vendor Inv.', 'KR': 'Vendor Invoice',
    'KZ': 'Vendor Payment', 'WA': 'Goods Issue', 'ZT': 'Value Reduction'
}
valid_types = ['WE', 'ZA']

wb = Workbook()
wb.remove(wb.active)

for prefix, group in df.groupby('Prefix'):
    parts = prefix.split('-')
    sheet_name = f"{parts[2]}_{parts[3]}" if len(parts) >= 4 else prefix
    ws = wb.create_sheet(title=sheet_name)

    group['Description'] = group['Type'].map(desc_map)

    # --- Summary Table ---
    summary = group.groupby('Type').agg({
        'WPS_ID': 'count',
        'Description': 'first'
    }).rename(columns={'WPS_ID': 'Count'}).reset_index()
    summary['Valid'] = summary['Type'].apply(lambda x: '✓' if x in valid_types else '✗')
    summary = summary[['Type', 'Count', 'Description', 'Valid']]

    apply_table_title(ws, "Document", 1, 4)
    summary_start_row = 2

    for i, row in enumerate(dataframe_to_rows(summary, index=False, header=True), start=summary_start_row):
        is_header = (i == summary_start_row)
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if is_header:
                apply_header_style(cell)
            else:
                cell.fill = fill_blue if (i - summary_start_row) % 2 == 1 else fill_white
                if j == 4:
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True, color="FF0000" if val == "✗" else "228B22")

    summary_end_row = summary_start_row + len(summary)
    apply_border_to_range(ws, summary_start_row, summary_end_row, 1, 4)

    # ✅ Add dropdowns to summary table
    summary_table_ref = f"A{summary_start_row}:D{summary_end_row}"
    summary_table_name = f"Summary_{sheet_name}_{ws.title}"
    summary_table = Table(displayName=summary_table_name, ref=summary_table_ref)
    summary_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                                   showRowStripes=True, showColumnStripes=False)
    summary_table.tableStyleInfo = summary_style
    ws.add_table(summary_table)

    # --- Pivot Table ---
    wps_types = group.pivot_table(index='WPS_ID', columns='Type', values='Val/COArea Crcy', aggfunc='sum').fillna(pd.NA)
    all_types = summary['Type'].unique()
    for col in all_types:
        if col not in wps_types.columns:
            wps_types[col] = pd.NA
    wps_types = wps_types[sorted(all_types)]
    wps_types.reset_index(inplace=True)
    wps_types['Grand Total'] = wps_types[all_types].apply(lambda row: pd.to_numeric(row, errors='coerce').fillna(0).sum(), axis=1)

    wps_types_numeric = wps_types.copy()
    for col in all_types:
        wps_types_numeric[col] = pd.to_numeric(wps_types_numeric[col], errors='coerce').fillna(0)
    col_totals = wps_types_numeric[all_types].sum()
    grand_total_row = {col: col_totals[col] for col in all_types}
    grand_total_row['WPS_ID'] = 'Grand Total'
    grand_total_row['Grand Total'] = col_totals.sum()
    value_table = pd.concat([wps_types, pd.DataFrame([grand_total_row])], ignore_index=True)

    start_col = 6
    value_start_row = 1
    for r_idx, row in enumerate(dataframe_to_rows(value_table, index=False, header=True), start=value_start_row):
        is_header = r_idx == value_start_row
        is_grand_total_row = row[0] == 'Grand Total'
        for c_idx, value in enumerate(row, start=start_col):
            col_name = value_table.columns[c_idx - start_col]
            cell = ws.cell(row=r_idx, column=c_idx)

            if is_header:
                cell.value = col_name
                apply_header_style(cell)
            else:
                if col_name == "WPS_ID":
                    cell.value = value
                elif pd.isna(value):
                    cell.value = ""
                elif isinstance(value, (int, float)):
                    formatted = f"₹ {value:15,.2f}"
                    cell.value = formatted
                    cell.font = Font(name='Courier New')
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.value = value

                if is_grand_total_row:
                    cell.font = Font(bold=True, color="FFFFFF", name='Courier New')
                    cell.fill = fill_total
                    cell.alignment = Alignment(horizontal="center")
                else:
                    fill = fill_blue if (r_idx - value_start_row) % 2 == 1 else fill_white
                    cell.fill = fill

            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

    # ✅ Add dropdowns to pivot table
    pivot_end_col = start_col + len(value_table.columns) - 1
    pivot_end_row = value_start_row + len(value_table)
    pivot_table_ref = f"{get_column_letter(start_col)}{value_start_row}:{get_column_letter(pivot_end_col)}{pivot_end_row}"
    pivot_table_name = f"Pivot_{sheet_name}_{ws.title}"
    pivot_table = Table(displayName=pivot_table_name, ref=pivot_table_ref)
    pivot_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                                 showRowStripes=True, showColumnStripes=False)
    pivot_table.tableStyleInfo = pivot_style
    ws.add_table(pivot_table)

# Finalize workbook
auto_adjust_column_widths(wb)

output_path = "output.xlsx"
if os.path.exists(output_path):
    try:
        os.remove(output_path)
    except PermissionError:
        print(f"\u274C Please close '{output_path}' and run again.")
        exit(1)

wb.save(output_path)
print(f"\u2705 Output saved to {output_path}")

if platform.system() == 'Windows':
    os.startfile(output_path)
elif platform.system() == 'Darwin':
    subprocess.call(["open", output_path])
else:
    subprocess.call(["xdg-open", output_path])
